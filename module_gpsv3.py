import openpyxl

import module_other_v3
import var_v3
import re
import caseid_v3







def ModuleTest():
    modetest = ''.join(re.findall(r'\d+', var_v3.modetest))
    for i in modetest:
        print("so1", i)
        if i == "0":
            run_all(self='')

        else:
            moduletest = ''.join(re.findall(r'\d+', var_v3.moduletest))
            print(type(moduletest))
            print(moduletest)
            for i in moduletest:
                print("so2", i)
                if i == "1":
                    login(self='')
                if i == "2":
                    monitor(self='')
                if i == "3":
                    route(self='')
                if i == "4":
                    administration(self='')
                if i == "5":
                    report(self='')
                if i == "6":
                    videoclip(self='')
                if i == "7":
                    image(self='')
                if i == "8":
                    utility(self='')
                if i == "9":
                    ai(self='')



def check_casenone():
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 60, 2, "0")
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 61, 2, "0")
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 62, 2, "0")
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 63, 2, "0")
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 64, 2, "0")
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 65, 2, "0")


    list_casefail = []
    wordbook = openpyxl.load_workbook(var_v3.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_v3.modetest))

    for i in modetest:
        print("i", i)
        if i == "1":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G"+rownum].value)
                print(sheet["H"+rownum].value)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức1: ", count)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 61, 2, count)


        if i == "2":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức2: ", count)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 62, 2, count)


        if i == "3":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức3: ", count)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 63, 2, count)


        if i == "4":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["K"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức4: ", count)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 64, 2, count)


        if i == "0":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G" + rownum].value)
                print(sheet["H" + rownum].value)
                if (sheet["H" + rownum].value == "x" or sheet["I" + rownum].value == "x" or sheet[
                    "J" + rownum].value == "x" or sheet["K" + rownum].value == "x") and sheet[
                    "G" + rownum].value == None:
                    print(sheet["A" + rownum].value)
                    case_fail = sheet["A" + rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức 0 : ", count)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 66, 2, count)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 60, 2, count)

        else:
            muc1 = int(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 61, 2))
            muc2 = int(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 62, 2))
            muc3 = int(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 63, 2))
            muc4 = int(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 64, 2))
            sumarry_case_none = muc1 + muc2 + muc3 + muc4
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 60, 2, sumarry_case_none)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 66, 2, sumarry_case_none)



    rownum = 9
    #Đặc biệt:
    list_case_nghiemtrong = []
    while (rownum < 1000):
        rownum += 1
        rownum = str(rownum)
        if sheet["L" + rownum].value == "x" and sheet["G" + rownum].value == "Fail":
            print(sheet["A" + rownum].value)
            case_fail = sheet["A" + rownum].value
            list_case_nghiemtrong.append(case_fail)
        rownum = int(rownum)
    print(list_case_nghiemtrong)
    count = len(list_case_nghiemtrong)
    print("Số case fail mức nghiêm trọng: ", count)
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 65, 2, count)



def check_casefail():
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 70, 2, "0")
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 71, 2, "0")
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 72, 2, "0")
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 73, 2, "0")
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 74, 2, "0")
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 75, 2, "0")


    list_casefail = []
    wordbook = openpyxl.load_workbook(var_v3.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_v3.modetest))

    for i in modetest:
        print("i", i)
        if i == "1":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G"+rownum].value)
                print(sheet["H"+rownum].value)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức1: ", count)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 71, 2, count)


        if i == "2":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức2: ", count)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 72, 2, count)


        if i == "3":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức3: ", count)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 73, 2, count)


        if i == "4":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["K"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức4: ", count)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 74, 2, count)


        if i == "0":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G" + rownum].value)
                print(sheet["H" + rownum].value)
                if (sheet["H" + rownum].value == "x" or sheet["I" + rownum].value == "x" or sheet[
                    "J" + rownum].value == "x" or sheet["K" + rownum].value == "x") and sheet[
                    "G" + rownum].value == "Fail":
                    print(sheet["A" + rownum].value)
                    case_fail = sheet["A" + rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức 0 : ", count)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 76, 2, count)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 70, 2, count)

        else:
            muc1 = int(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 71, 2))
            muc2 = int(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 72, 2))
            muc3 = int(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 73, 2))
            muc4 = int(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 74, 2))
            sumarry_case_none = muc1 + muc2 + muc3 + muc4
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 70, 2, sumarry_case_none)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 76, 2, sumarry_case_none)



    rownum = 9
    #Đặc biệt:
    list_case_nghiemtrong = []
    while (rownum < 1000):
        rownum += 1
        rownum = str(rownum)
        if sheet["L" + rownum].value == "x" and sheet["G" + rownum].value == "Fail":
            print(sheet["A" + rownum].value)
            case_fail = sheet["A" + rownum].value
            list_case_nghiemtrong.append(case_fail)
        rownum = int(rownum)
    print(list_case_nghiemtrong)
    count = len(list_case_nghiemtrong)
    print("Số case fail mức nghiêm trọng: ", count)
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 75, 2, count)


    if var_v3.modetest == "0":
        case_fail = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 70, 2))
        var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 77, 2, case_fail)
    if var_v3.modetest == "1":
        case_fail = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 71, 2))
        var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 77, 2, case_fail)
    if var_v3.modetest == "2":
        case_fail = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 72, 2))
        var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 77, 2, case_fail)
    if var_v3.modetest == "3":
        case_fail = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 73, 2))
        var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 77, 2, case_fail)
    if var_v3.modetest == "4":
        case_fail = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 74, 2))
        var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 77, 2, case_fail)



def check_casepass():
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 70, 2, "0")
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 71, 2, "0")
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 72, 2, "0")
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 73, 2, "0")
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 74, 2, "0")
    var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 75, 2, "0")


    list_casefail = []
    wordbook = openpyxl.load_workbook(var_v3.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_v3.modetest))

    for i in modetest:
        print("i", i)
        if i == "1":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G"+rownum].value)
                print(sheet["H"+rownum].value)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức1: ", count)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 81, 2, count)


        if i == "2":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức2: ", count)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 82, 2, count)


        if i == "3":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức3: ", count)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 83, 2, count)


        if i == "4":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["K"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức4: ", count)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 84, 2, count)


        if i == "0":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G" + rownum].value)
                print(sheet["H" + rownum].value)
                if (sheet["H" + rownum].value == "x" or sheet["I" + rownum].value == "x" or sheet[
                    "J" + rownum].value == "x" or sheet["K" + rownum].value == "x") and sheet[
                    "G" + rownum].value == "Pass":
                    print(sheet["A" + rownum].value)
                    case_fail = sheet["A" + rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức 0 : ", count)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 86, 2, count)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 80, 2, count)

        else:
            muc1 = int(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 71, 2))
            muc2 = int(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 72, 2))
            muc3 = int(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 73, 2))
            muc4 = int(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 74, 2))
            sumarry_case_none = muc1 + muc2 + muc3 + muc4
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 80, 2, sumarry_case_none)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 86, 2, sumarry_case_none)

        if var_v3.modetest == "0":
            case_fail = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 80, 2))
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 87, 2, case_fail)
        if var_v3.modetest == "1":
            case_fail = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 81, 2))
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 87, 2, case_fail)
        if var_v3.modetest == "2":
            case_fail = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 82, 2))
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 87, 2, case_fail)
        if var_v3.modetest == "3":
            case_fail = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 83, 2))
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 87, 2, case_fail)
        if var_v3.modetest == "4":
            case_fail = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 84, 2))
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 87, 2, case_fail)



def retest_casenone(self):
    list_casefail = []
    wordbook = openpyxl.load_workbook(var_v3.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9

    modetest = ''.join(re.findall(r'\d+', var_v3.modetest))
    for i in modetest:
        print("i", i)
        if i == "1":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G"+rownum].value)
                print(sheet["H"+rownum].value)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức1: ", count)

        if i == "2":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức2: ", count)


        if i == "3":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức3: ", count)

        if i == "4":
            while (rownum < 1000):
                rownum += 1
                rownum = str(rownum)
                if sheet["K"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức4: ", count)

        for case in list_casefail:
            try:
                if case == 'Login01':
                    caseid_v3.caseid_login01(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login02':
                    caseid_v3.caseid_login02(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login03':
                    caseid_v3.caseid_login03(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login04':
                    caseid_v3.caseid_login04(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login05':
                    caseid_v3.caseid_login05(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login06':
                    caseid_v3.caseid_login06(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login07':
                    caseid_v3.caseid_login07(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login08':
                    caseid_v3.caseid_login08(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login09':
                    caseid_v3.caseid_login09(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login10':
                    caseid_v3.caseid_login10(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login11':
                    caseid_v3.caseid_login11(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login12':
                    caseid_v3.caseid_login12(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login13':
                    caseid_v3.caseid_login13(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login14':
                    caseid_v3.caseid_login14(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login15':
                    caseid_v3.caseid_login15(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login16':
                    caseid_v3.caseid_login16(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login17':
                    caseid_v3.caseid_login17(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login18':
                    caseid_v3.caseid_login18(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login19':
                    caseid_v3.caseid_login19(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login20':
                    caseid_v3.caseid_login20(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login21':
                    caseid_v3.caseid_login21(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login22':
                    caseid_v3.caseid_login22(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login23':
                    caseid_v3.caseid_login23(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login24':
                    caseid_v3.caseid_login24(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login25':
                    caseid_v3.caseid_login25(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Login26':
                    caseid_v3.caseid_login26(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat01':
                    caseid_v3.caseid_giamsat01(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat02':
                    caseid_v3.caseid_giamsat02(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat03':
                    caseid_v3.caseid_giamsat03(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat04':
                    caseid_v3.caseid_giamsat04(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat05':
                    caseid_v3.caseid_giamsat05(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat06':
                    caseid_v3.caseid_giamsat06(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat07':
                    caseid_v3.caseid_giamsat07(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat08':
                    caseid_v3.caseid_giamsat08(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat09':
                    caseid_v3.caseid_giamsat09(self)
            except:
                module_other_v3.swich_tab_0()

            try:
                if case == 'GiamSat10':
                    caseid_v3.caseid_giamsat10(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat11':
                    caseid_v3.caseid_giamsat11(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat12':
                    caseid_v3.caseid_giamsat12(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat13':
                    caseid_v3.caseid_giamsat13(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat14':
                    caseid_v3.caseid_giamsat14(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat15':
                    caseid_v3.caseid_giamsat15(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat16':
                    caseid_v3.caseid_giamsat16(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat17':
                    caseid_v3.caseid_giamsat17(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat18':
                    caseid_v3.caseid_giamsat18(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat19':
                    caseid_v3.caseid_giamsat19(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat20':
                    caseid_v3.caseid_giamsat20(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat21':
                    caseid_v3.caseid_giamsat21(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat22':
                    caseid_v3.caseid_giamsat22(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat23':
                    caseid_v3.caseid_giamsat23(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat24':
                    caseid_v3.caseid_giamsat24(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat25':
                    caseid_v3.caseid_giamsat25(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat26':
                    caseid_v3.caseid_giamsat26(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat27':
                    caseid_v3.caseid_giamsat27(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat28':
                    caseid_v3.caseid_giamsat28(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat29':
                    caseid_v3.caseid_giamsat29(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat30':
                    caseid_v3.caseid_giamsat30(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat31':
                    caseid_v3.caseid_giamsat31(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat32':
                    caseid_v3.caseid_giamsat32(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat33':
                    caseid_v3.caseid_giamsat33(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat34':
                    caseid_v3.caseid_giamsat34(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat35':
                    caseid_v3.caseid_giamsat35(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat36':
                    caseid_v3.caseid_giamsat36(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat37':
                    caseid_v3.caseid_giamsat37(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat38':
                    caseid_v3.caseid_giamsat38(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat39':
                    caseid_v3.caseid_giamsat39(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat40':
                    caseid_v3.caseid_giamsat40(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat41':
                    caseid_v3.caseid_giamsat41(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat42':
                    caseid_v3.caseid_giamsat42(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat43':
                    caseid_v3.caseid_giamsat43(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat44':
                    caseid_v3.caseid_giamsat44(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat45':
                    caseid_v3.caseid_giamsat45(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat46':
                    caseid_v3.caseid_giamsat46(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat47':
                    caseid_v3.caseid_giamsat47(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat48':
                    caseid_v3.caseid_giamsat48(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat49':
                    caseid_v3.caseid_giamsat49(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat50':
                    caseid_v3.caseid_giamsat50(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat51':
                    caseid_v3.caseid_giamsat51(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat52':
                    caseid_v3.caseid_giamsat52(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat53':
                    caseid_v3.caseid_giamsat53(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat54':
                    caseid_v3.caseid_giamsat54(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat55':
                    caseid_v3.caseid_giamsat55(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat56':
                    caseid_v3.caseid_giamsat56(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat57':
                    caseid_v3.caseid_giamsat57(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat58':
                    caseid_v3.caseid_giamsat58(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat59':
                    caseid_v3.caseid_giamsat59(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat60':
                    caseid_v3.caseid_giamsat60(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat61':
                    caseid_v3.caseid_giamsat61(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat62':
                    caseid_v3.caseid_giamsat62(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat63':
                    caseid_v3.caseid_giamsat63(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat64':
                    caseid_v3.caseid_giamsat64(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat65':
                    caseid_v3.caseid_giamsat65(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat66':
                    caseid_v3.caseid_giamsat66(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat67':
                    caseid_v3.caseid_giamsat67(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat68':
                    caseid_v3.caseid_giamsat68(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat69':
                    caseid_v3.caseid_giamsat69(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat70':
                    caseid_v3.caseid_giamsat70(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat71':
                    caseid_v3.caseid_giamsat71(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat72':
                    caseid_v3.caseid_giamsat72(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat73':
                    caseid_v3.caseid_giamsat73(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat74':
                    caseid_v3.caseid_giamsat74(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat75':
                    caseid_v3.caseid_giamsat75(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat76':
                    caseid_v3.caseid_giamsat76(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat77':
                    caseid_v3.caseid_giamsat77(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat78':
                    caseid_v3.caseid_giamsat78(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat79':
                    caseid_v3.caseid_giamsat79(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat80':
                    caseid_v3.caseid_giamsat80(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat81':
                    caseid_v3.caseid_giamsat81(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat82':
                    caseid_v3.caseid_giamsat82(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat83':
                    caseid_v3.caseid_giamsat83(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat84':
                    caseid_v3.caseid_giamsat84(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat85':
                    caseid_v3.caseid_giamsat85(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat86':
                    caseid_v3.caseid_giamsat86(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat87':
                    caseid_v3.caseid_giamsat87(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat88':
                    caseid_v3.caseid_giamsat88(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat89':
                    caseid_v3.caseid_giamsat89(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat90':
                    caseid_v3.caseid_giamsat90(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat91':
                    caseid_v3.caseid_giamsat91(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat92':
                    caseid_v3.caseid_giamsat92(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat93':
                    caseid_v3.caseid_giamsat93(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat94':
                    caseid_v3.caseid_giamsat94(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat95':
                    caseid_v3.caseid_giamsat95(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat96':
                    caseid_v3.caseid_giamsat96(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat97':
                    caseid_v3.caseid_giamsat97(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat98':
                    caseid_v3.caseid_giamsat98(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat99':
                    caseid_v3.caseid_giamsat99(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat100':
                    caseid_v3.caseid_giamsat100(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat101':
                    caseid_v3.caseid_giamsat101(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat102':
                    caseid_v3.caseid_giamsat102(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat103':
                    caseid_v3.caseid_giamsat103(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat104':
                    caseid_v3.caseid_giamsat104(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat105':
                    caseid_v3.caseid_giamsat105(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat106':
                    caseid_v3.caseid_giamsat106(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat107':
                    caseid_v3.caseid_giamsat107(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat108':
                    caseid_v3.caseid_giamsat108(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat109':
                    caseid_v3.caseid_giamsat109(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat110':
                    caseid_v3.caseid_giamsat110(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat111':
                    caseid_v3.caseid_giamsat111(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat112':
                    caseid_v3.caseid_giamsat112(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat113':
                    caseid_v3.caseid_giamsat113(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat114':
                    caseid_v3.caseid_giamsat114(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat115':
                    caseid_v3.caseid_giamsat115(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat116':
                    caseid_v3.caseid_giamsat116(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat117':
                    caseid_v3.caseid_giamsat117(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat118':
                    caseid_v3.caseid_giamsat118(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat119':
                    caseid_v3.caseid_giamsat119(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat120':
                    caseid_v3.caseid_giamsat120(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat121':
                    caseid_v3.caseid_giamsat121(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat122':
                    caseid_v3.caseid_giamsat122(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat123':
                    caseid_v3.caseid_giamsat123(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat124':
                    caseid_v3.caseid_giamsat124(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat125':
                    caseid_v3.caseid_giamsat125(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat126':
                    caseid_v3.caseid_giamsat126(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat127':
                    caseid_v3.caseid_giamsat127(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat128':
                    caseid_v3.caseid_giamsat128(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat129':
                    caseid_v3.caseid_giamsat129(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat130':
                    caseid_v3.caseid_giamsat130(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat131':
                    caseid_v3.caseid_giamsat131(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat132':
                    caseid_v3.caseid_giamsat132(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat133':
                    caseid_v3.caseid_giamsat133(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat134':
                    caseid_v3.caseid_giamsat134(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat135':
                    caseid_v3.caseid_giamsat135(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat136':
                    caseid_v3.caseid_giamsat136(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat137':
                    caseid_v3.caseid_giamsat137(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat138':
                    caseid_v3.caseid_giamsat138(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat139':
                    caseid_v3.caseid_giamsat139(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat140':
                    caseid_v3.caseid_giamsat140(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat141':
                    caseid_v3.caseid_giamsat141(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat142':
                    caseid_v3.caseid_giamsat142(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat143':
                    caseid_v3.caseid_giamsat143(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat144':
                    caseid_v3.caseid_giamsat144(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat145':
                    caseid_v3.caseid_giamsat145(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat146':
                    caseid_v3.caseid_giamsat146(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat147':
                    caseid_v3.caseid_giamsat147(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat148':
                    caseid_v3.caseid_giamsat148(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat149':
                    caseid_v3.caseid_giamsat149(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat150':
                    caseid_v3.caseid_giamsat150(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat151':
                    caseid_v3.caseid_giamsat151(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat152':
                    caseid_v3.caseid_giamsat152(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat153':
                    caseid_v3.caseid_giamsat153(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat154':
                    caseid_v3.caseid_giamsat154(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat155':
                    caseid_v3.caseid_giamsat155(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat156':
                    caseid_v3.caseid_giamsat156(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat157':
                    caseid_v3.caseid_giamsat157(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat158':
                    caseid_v3.caseid_giamsat158(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat159':
                    caseid_v3.caseid_giamsat159(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat160':
                    caseid_v3.caseid_giamsat160(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat161':
                    caseid_v3.caseid_giamsat161(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat162':
                    caseid_v3.caseid_giamsat162(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat163':
                    caseid_v3.caseid_giamsat163(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat164':
                    caseid_v3.caseid_giamsat164(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat165':
                    caseid_v3.caseid_giamsat165(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat166':
                    caseid_v3.caseid_giamsat166(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat167':
                    caseid_v3.caseid_giamsat167(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat168':
                    caseid_v3.caseid_giamsat168(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat169':
                    caseid_v3.caseid_giamsat169(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat170':
                    caseid_v3.caseid_giamsat170(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat171':
                    caseid_v3.caseid_giamsat171(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat172':
                    caseid_v3.caseid_giamsat172(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat173':
                    caseid_v3.caseid_giamsat173(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat174':
                    caseid_v3.caseid_giamsat174(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat175':
                    caseid_v3.caseid_giamsat175(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat176':
                    caseid_v3.caseid_giamsat176(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat177':
                    caseid_v3.caseid_giamsat177(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat178':
                    caseid_v3.caseid_giamsat178(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat179':
                    caseid_v3.caseid_giamsat179(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat180':
                    caseid_v3.caseid_giamsat180(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat181':
                    caseid_v3.caseid_giamsat181(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat182':
                    caseid_v3.caseid_giamsat182(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat183':
                    caseid_v3.caseid_giamsat183(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat184':
                    caseid_v3.caseid_giamsat184(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat185':
                    caseid_v3.caseid_giamsat185(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat186':
                    caseid_v3.caseid_giamsat186(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat187':
                    caseid_v3.caseid_giamsat187(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat188':
                    caseid_v3.caseid_giamsat188(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat189':
                    caseid_v3.caseid_giamsat189(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat190':
                    caseid_v3.caseid_giamsat190(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat191':
                    caseid_v3.caseid_giamsat191(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat192':
                    caseid_v3.caseid_giamsat192(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat193':
                    caseid_v3.caseid_giamsat193(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat194':
                    caseid_v3.caseid_giamsat194(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat195':
                    caseid_v3.caseid_giamsat195(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat196':
                    caseid_v3.caseid_giamsat196(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat197':
                    caseid_v3.caseid_giamsat197(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat198':
                    caseid_v3.caseid_giamsat198(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat199':
                    caseid_v3.caseid_giamsat199(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat200':
                    caseid_v3.caseid_giamsat200(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat201':
                    caseid_v3.caseid_giamsat201(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat202':
                    caseid_v3.caseid_giamsat202(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat203':
                    caseid_v3.caseid_giamsat203(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat204':
                    caseid_v3.caseid_giamsat204(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat205':
                    caseid_v3.caseid_giamsat205(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat206':
                    caseid_v3.caseid_giamsat206(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat207':
                    caseid_v3.caseid_giamsat207(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat208':
                    caseid_v3.caseid_giamsat208(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat209':
                    caseid_v3.caseid_giamsat209(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat210':
                    caseid_v3.caseid_giamsat210(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat211':
                    caseid_v3.caseid_giamsat211(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat212':
                    caseid_v3.caseid_giamsat212(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat213':
                    caseid_v3.caseid_giamsat213(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat214':
                    caseid_v3.caseid_giamsat214(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat215':
                    caseid_v3.caseid_giamsat215(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat216':
                    caseid_v3.caseid_giamsat216(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat217':
                    caseid_v3.caseid_giamsat217(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat218':
                    caseid_v3.caseid_giamsat218(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat219':
                    caseid_v3.caseid_giamsat219(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat220':
                    caseid_v3.caseid_giamsat220(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat221':
                    caseid_v3.caseid_giamsat221(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat222':
                    caseid_v3.caseid_giamsat222(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat223':
                    caseid_v3.caseid_giamsat223(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat224':
                    caseid_v3.caseid_giamsat224(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat225':
                    caseid_v3.caseid_giamsat225(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat226':
                    caseid_v3.caseid_giamsat226(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat227':
                    caseid_v3.caseid_giamsat227(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat228':
                    caseid_v3.caseid_giamsat228(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'GiamSat229':
                    caseid_v3.caseid_giamsat229(self)
            except:
                module_other_v3.swich_tab_0()

            try:
                if case == 'Route01':
                    caseid_v3.caseid_route01(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route02':
                    caseid_v3.caseid_route02(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route03':
                    caseid_v3.caseid_route03(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route04':
                    caseid_v3.caseid_route04(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route05':
                    caseid_v3.caseid_route05(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route06':
                    caseid_v3.caseid_route06(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route07':
                    caseid_v3.caseid_route07(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route08':
                    caseid_v3.caseid_route08(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route09':
                    caseid_v3.caseid_route09(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route10':
                    caseid_v3.caseid_route10(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route11':
                    caseid_v3.caseid_route11(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route12':
                    caseid_v3.caseid_route12(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route13':
                    caseid_v3.caseid_route13(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route14':
                    caseid_v3.caseid_route14(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route15':
                    caseid_v3.caseid_route15(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route16':
                    caseid_v3.caseid_route16(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route17':
                    caseid_v3.caseid_route17(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route18':
                    caseid_v3.caseid_route18(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route19':
                    caseid_v3.caseid_route19(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route20':
                    caseid_v3.caseid_route20(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route21':
                    caseid_v3.caseid_route21(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route22':
                    caseid_v3.caseid_route22(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route23':
                    caseid_v3.caseid_route23(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route24':
                    caseid_v3.caseid_route24(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route25':
                    caseid_v3.caseid_route25(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route26':
                    caseid_v3.caseid_route26(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route27':
                    caseid_v3.caseid_route27(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route28':
                    caseid_v3.caseid_route28(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route29':
                    caseid_v3.caseid_route29(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route30':
                    caseid_v3.caseid_route30(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route31':
                    caseid_v3.caseid_route31(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route32':
                    caseid_v3.caseid_route32(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route33':
                    caseid_v3.caseid_route33(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route34':
                    caseid_v3.caseid_route34(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route35':
                    caseid_v3.caseid_route35(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route36':
                    caseid_v3.caseid_route36(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route37':
                    caseid_v3.caseid_route37(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route38':
                    caseid_v3.caseid_route38(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route39':
                    caseid_v3.caseid_route39(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route40':
                    caseid_v3.caseid_route40(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route41':
                    caseid_v3.caseid_route41(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route42':
                    caseid_v3.caseid_route42(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route43':
                    caseid_v3.caseid_route43(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route44':
                    caseid_v3.caseid_route44(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route45':
                    caseid_v3.caseid_route45(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route46':
                    caseid_v3.caseid_route46(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route47':
                    caseid_v3.caseid_route47(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route48':
                    caseid_v3.caseid_route48(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route49':
                    caseid_v3.caseid_route49(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route50':
                    caseid_v3.caseid_route50(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route51':
                    caseid_v3.caseid_route51(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route52':
                    caseid_v3.caseid_route52(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Route53':
                    caseid_v3.caseid_route53(self)
            except:
                module_other_v3.swich_tab_0()

            try:
                if case == 'Admin01':
                    caseid_v3.caseid_admin01(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin02':
                    caseid_v3.caseid_admin02(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin03':
                    caseid_v3.caseid_admin03(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin04':
                    caseid_v3.caseid_admin04(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin05':
                    caseid_v3.caseid_admin05(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin06':
                    caseid_v3.caseid_admin06(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin07':
                    caseid_v3.caseid_admin07(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin08':
                    caseid_v3.caseid_admin08(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin09':
                    caseid_v3.caseid_admin09(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin10':
                    caseid_v3.caseid_admin10(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin11':
                    caseid_v3.caseid_admin11(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin12':
                    caseid_v3.caseid_admin12(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin13':
                    caseid_v3.caseid_admin13(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin14':
                    caseid_v3.caseid_admin14(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin15':
                    caseid_v3.caseid_admin15(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin16':
                    caseid_v3.caseid_admin16(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin17':
                    caseid_v3.caseid_admin17(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin18':
                    caseid_v3.caseid_admin18(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin19':
                    caseid_v3.caseid_admin19(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin20':
                    caseid_v3.caseid_admin20(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin21':
                    caseid_v3.caseid_admin21(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin22':
                    caseid_v3.caseid_admin22(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin23':
                    caseid_v3.caseid_admin23(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin24':
                    caseid_v3.caseid_admin24(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin25':
                    caseid_v3.caseid_admin25(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin26':
                    caseid_v3.caseid_admin26(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin27':
                    caseid_v3.caseid_admin27(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin28':
                    caseid_v3.caseid_admin28(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin29':
                    caseid_v3.caseid_admin29(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin30':
                    caseid_v3.caseid_admin30(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin31':
                    caseid_v3.caseid_admin31(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin32':
                    caseid_v3.caseid_admin32(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin33':
                    caseid_v3.caseid_admin33(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin34':
                    caseid_v3.caseid_admin34(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin35':
                    caseid_v3.caseid_admin35(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin36':
                    caseid_v3.caseid_admin36(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin37':
                    caseid_v3.caseid_admin37(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin38':
                    caseid_v3.caseid_admin38(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin39':
                    caseid_v3.caseid_admin39(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin40':
                    caseid_v3.caseid_admin40(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin41':
                    caseid_v3.caseid_admin41(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin42':
                    caseid_v3.caseid_admin42(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin43':
                    caseid_v3.caseid_admin43(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin44':
                    caseid_v3.caseid_admin44(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin45':
                    caseid_v3.caseid_admin45(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin46':
                    caseid_v3.caseid_admin46(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin47':
                    caseid_v3.caseid_admin47(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin48':
                    caseid_v3.caseid_admin48(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin49':
                    caseid_v3.caseid_admin49(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin50':
                    caseid_v3.caseid_admin50(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin51':
                    caseid_v3.caseid_admin51(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin52':
                    caseid_v3.caseid_admin52(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin53':
                    caseid_v3.caseid_admin53(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin54':
                    caseid_v3.caseid_admin54(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin55':
                    caseid_v3.caseid_admin55(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin56':
                    caseid_v3.caseid_admin56(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin57':
                    caseid_v3.caseid_admin57(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin58':
                    caseid_v3.caseid_admin58(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin59':
                    caseid_v3.caseid_admin59(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin60':
                    caseid_v3.caseid_admin60(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin61':
                    caseid_v3.caseid_admin61(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin62':
                    caseid_v3.caseid_admin62(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin63':
                    caseid_v3.caseid_admin63(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin64':
                    caseid_v3.caseid_admin64(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin65':
                    caseid_v3.caseid_admin65(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin66':
                    caseid_v3.caseid_admin66(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin67':
                    caseid_v3.caseid_admin67(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin68':
                    caseid_v3.caseid_admin68(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin69':
                    caseid_v3.caseid_admin69(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin70':
                    caseid_v3.caseid_admin70(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin71':
                    caseid_v3.caseid_admin71(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin72':
                    caseid_v3.caseid_admin72(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin73':
                    caseid_v3.caseid_admin73(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin74':
                    caseid_v3.caseid_admin74(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin75':
                    caseid_v3.caseid_admin75(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin76':
                    caseid_v3.caseid_admin76(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin77':
                    caseid_v3.caseid_admin77(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin78':
                    caseid_v3.caseid_admin78(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin79':
                    caseid_v3.caseid_admin79(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin80':
                    caseid_v3.caseid_admin80(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin81':
                    caseid_v3.caseid_admin81(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin82':
                    caseid_v3.caseid_admin82(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin83':
                    caseid_v3.caseid_admin83(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin84':
                    caseid_v3.caseid_admin84(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin85':
                    caseid_v3.caseid_admin85(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin86':
                    caseid_v3.caseid_admin86(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin87':
                    caseid_v3.caseid_admin87(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin88':
                    caseid_v3.caseid_admin88(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin89':
                    caseid_v3.caseid_admin89(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin90':
                    caseid_v3.caseid_admin90(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin91':
                    caseid_v3.caseid_admin91(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin92':
                    caseid_v3.caseid_admin92(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin93':
                    caseid_v3.caseid_admin93(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin94':
                    caseid_v3.caseid_admin94(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin95':
                    caseid_v3.caseid_admin95(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin96':
                    caseid_v3.caseid_admin96(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin97':
                    caseid_v3.caseid_admin97(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin98':
                    caseid_v3.caseid_admin98(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin99':
                    caseid_v3.caseid_admin99(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin100':
                    caseid_v3.caseid_admin100(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin101':
                    caseid_v3.caseid_admin101(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin102':
                    caseid_v3.caseid_admin102(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin103':
                    caseid_v3.caseid_admin103(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin104':
                    caseid_v3.caseid_admin104(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin105':
                    caseid_v3.caseid_admin105(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin106':
                    caseid_v3.caseid_admin106(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin107':
                    caseid_v3.caseid_admin107(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin108':
                    caseid_v3.caseid_admin108(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin109':
                    caseid_v3.caseid_admin109(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin110':
                    caseid_v3.caseid_admin110(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin111':
                    caseid_v3.caseid_admin111(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin112':
                    caseid_v3.caseid_admin112(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin113':
                    caseid_v3.caseid_admin113(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin114':
                    caseid_v3.caseid_admin114(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin115':
                    caseid_v3.caseid_admin115(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin116':
                    caseid_v3.caseid_admin116(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin117':
                    caseid_v3.caseid_admin117(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin118':
                    caseid_v3.caseid_admin118(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin119':
                    caseid_v3.caseid_admin119(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin120':
                    caseid_v3.caseid_admin120(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin121':
                    caseid_v3.caseid_admin121(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin122':
                    caseid_v3.caseid_admin122(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin123':
                    caseid_v3.caseid_admin123(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin124':
                    caseid_v3.caseid_admin124(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin125':
                    caseid_v3.caseid_admin125(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin126':
                    caseid_v3.caseid_admin126(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin127':
                    caseid_v3.caseid_admin127(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin128':
                    caseid_v3.caseid_admin128(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin129':
                    caseid_v3.caseid_admin129(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin130':
                    caseid_v3.caseid_admin130(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin131':
                    caseid_v3.caseid_admin131(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin132':
                    caseid_v3.caseid_admin132(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin133':
                    caseid_v3.caseid_admin133(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin134':
                    caseid_v3.caseid_admin134(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin135':
                    caseid_v3.caseid_admin135(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin136':
                    caseid_v3.caseid_admin136(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin137':
                    caseid_v3.caseid_admin137(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin138':
                    caseid_v3.caseid_admin138(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin139':
                    caseid_v3.caseid_admin139(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin140':
                    caseid_v3.caseid_admin140(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin141':
                    caseid_v3.caseid_admin141(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin142':
                    caseid_v3.caseid_admin142(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin143':
                    caseid_v3.caseid_admin143(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin144':
                    caseid_v3.caseid_admin144(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin145':
                    caseid_v3.caseid_admin145(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin146':
                    caseid_v3.caseid_admin146(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin147':
                    caseid_v3.caseid_admin147(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin148':
                    caseid_v3.caseid_admin148(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin149':
                    caseid_v3.caseid_admin149(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin150':
                    caseid_v3.caseid_admin150(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin151':
                    caseid_v3.caseid_admin151(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin152':
                    caseid_v3.caseid_admin152(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin153':
                    caseid_v3.caseid_admin153(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin154':
                    caseid_v3.caseid_admin154(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin155':
                    caseid_v3.caseid_admin155(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin156':
                    caseid_v3.caseid_admin156(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin157':
                    caseid_v3.caseid_admin157(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin158':
                    caseid_v3.caseid_admin158(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin159':
                    caseid_v3.caseid_admin159(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin160':
                    caseid_v3.caseid_admin160(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin161':
                    caseid_v3.caseid_admin161(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin162':
                    caseid_v3.caseid_admin162(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin163':
                    caseid_v3.caseid_admin163(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin164':
                    caseid_v3.caseid_admin164(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin165':
                    caseid_v3.caseid_admin165(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin166':
                    caseid_v3.caseid_admin166(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin167':
                    caseid_v3.caseid_admin167(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin168':
                    caseid_v3.caseid_admin168(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin169':
                    caseid_v3.caseid_admin169(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin170':
                    caseid_v3.caseid_admin170(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin171':
                    caseid_v3.caseid_admin171(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin172':
                    caseid_v3.caseid_admin172(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin173':
                    caseid_v3.caseid_admin173(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin174':
                    caseid_v3.caseid_admin174(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin175':
                    caseid_v3.caseid_admin175(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin176':
                    caseid_v3.caseid_admin176(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin177':
                    caseid_v3.caseid_admin177(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin178':
                    caseid_v3.caseid_admin178(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin179':
                    caseid_v3.caseid_admin179(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin180':
                    caseid_v3.caseid_admin180(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin181':
                    caseid_v3.caseid_admin181(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin182':
                    caseid_v3.caseid_admin182(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin183':
                    caseid_v3.caseid_admin183(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin184':
                    caseid_v3.caseid_admin184(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin185':
                    caseid_v3.caseid_admin185(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin186':
                    caseid_v3.caseid_admin186(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin187':
                    caseid_v3.caseid_admin187(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin188':
                    caseid_v3.caseid_admin188(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin189':
                    caseid_v3.caseid_admin189(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin190':
                    caseid_v3.caseid_admin190(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin191':
                    caseid_v3.caseid_admin191(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin192':
                    caseid_v3.caseid_admin192(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin193':
                    caseid_v3.caseid_admin193(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin194':
                    caseid_v3.caseid_admin194(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin195':
                    caseid_v3.caseid_admin195(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin196':
                    caseid_v3.caseid_admin196(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin197':
                    caseid_v3.caseid_admin197(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin198':
                    caseid_v3.caseid_admin198(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Admin199':
                    caseid_v3.caseid_admin199(self)
            except:
                module_other_v3.swich_tab_0()

            try:
                if case == 'Report01':
                    caseid_v3.caseid_report01(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report02':
                    caseid_v3.caseid_report02(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report03':
                    caseid_v3.caseid_report03(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report04':
                    caseid_v3.caseid_report04(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report05':
                    caseid_v3.caseid_report05(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report06':
                    caseid_v3.caseid_report06(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report07':
                    caseid_v3.caseid_report07(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report08':
                    caseid_v3.caseid_report08(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report09':
                    caseid_v3.caseid_report09(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report10':
                    caseid_v3.caseid_report10(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report11':
                    caseid_v3.caseid_report11(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report12':
                    caseid_v3.caseid_report12(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report13':
                    caseid_v3.caseid_report13(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report14':
                    caseid_v3.caseid_report14(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report15':
                    caseid_v3.caseid_report15(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report16':
                    caseid_v3.caseid_report16(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report17':
                    caseid_v3.caseid_report17(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report18':
                    caseid_v3.caseid_report18(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report19':
                    caseid_v3.caseid_report19(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report20':
                    caseid_v3.caseid_report20(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report21':
                    caseid_v3.caseid_report21(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report22':
                    caseid_v3.caseid_report22(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report23':
                    caseid_v3.caseid_report23(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report24':
                    caseid_v3.caseid_report24(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report25':
                    caseid_v3.caseid_report25(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report26':
                    caseid_v3.caseid_report26(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report27':
                    caseid_v3.caseid_report27(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report28':
                    caseid_v3.caseid_report28(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report29':
                    caseid_v3.caseid_report29(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report30':
                    caseid_v3.caseid_report30(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report31':
                    caseid_v3.caseid_report31(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report32':
                    caseid_v3.caseid_report32(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report33':
                    caseid_v3.caseid_report33(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report34':
                    caseid_v3.caseid_report34(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report35':
                    caseid_v3.caseid_report35(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report36':
                    caseid_v3.caseid_report36(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report37':
                    caseid_v3.caseid_report37(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report38':
                    caseid_v3.caseid_report38(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report39':
                    caseid_v3.caseid_report39(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report40':
                    caseid_v3.caseid_report40(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report41':
                    caseid_v3.caseid_report41(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report42':
                    caseid_v3.caseid_report42(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report43':
                    caseid_v3.caseid_report43(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report44':
                    caseid_v3.caseid_report44(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report45':
                    caseid_v3.caseid_report45(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report46':
                    caseid_v3.caseid_report46(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report47':
                    caseid_v3.caseid_report47(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report48':
                    caseid_v3.caseid_report48(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report49':
                    caseid_v3.caseid_report49(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report50':
                    caseid_v3.caseid_report50(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report51':
                    caseid_v3.caseid_report51(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report52':
                    caseid_v3.caseid_report52(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report53':
                    caseid_v3.caseid_report53(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report54':
                    caseid_v3.caseid_report54(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report55':
                    caseid_v3.caseid_report55(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report56':
                    caseid_v3.caseid_report56(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report57':
                    caseid_v3.caseid_report57(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report58':
                    caseid_v3.caseid_report58(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report59':
                    caseid_v3.caseid_report59(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report60':
                    caseid_v3.caseid_report60(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report61':
                    caseid_v3.caseid_report61(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report62':
                    caseid_v3.caseid_report62(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report63':
                    caseid_v3.caseid_report63(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report64':
                    caseid_v3.caseid_report64(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report65':
                    caseid_v3.caseid_report65(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report66':
                    caseid_v3.caseid_report66(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report67':
                    caseid_v3.caseid_report67(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report68':
                    caseid_v3.caseid_report68(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report69':
                    caseid_v3.caseid_report69(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report70':
                    caseid_v3.caseid_report70(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report71':
                    caseid_v3.caseid_report71(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report72':
                    caseid_v3.caseid_report72(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report73':
                    caseid_v3.caseid_report73(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report74':
                    caseid_v3.caseid_report74(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report75':
                    caseid_v3.caseid_report75(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report76':
                    caseid_v3.caseid_report76(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report77':
                    caseid_v3.caseid_report77(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report78':
                    caseid_v3.caseid_report78(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report79':
                    caseid_v3.caseid_report79(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report80':
                    caseid_v3.caseid_report80(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report81':
                    caseid_v3.caseid_report81(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report82':
                    caseid_v3.caseid_report82(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report83':
                    caseid_v3.caseid_report83(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report84':
                    caseid_v3.caseid_report84(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report85':
                    caseid_v3.caseid_report85(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report86':
                    caseid_v3.caseid_report86(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report87':
                    caseid_v3.caseid_report87(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report88':
                    caseid_v3.caseid_report88(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report89':
                    caseid_v3.caseid_report89(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report90':
                    caseid_v3.caseid_report90(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report91':
                    caseid_v3.caseid_report91(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report92':
                    caseid_v3.caseid_report92(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report93':
                    caseid_v3.caseid_report93(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report94':
                    caseid_v3.caseid_report94(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Report95':
                    caseid_v3.caseid_report95(self)
            except:
                module_other_v3.swich_tab_0()

            try:
                if case == 'Video01':
                    caseid_v3.caseid_video01(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video02':
                    caseid_v3.caseid_video02(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video03':
                    caseid_v3.caseid_video03(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video04':
                    caseid_v3.caseid_video04(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video05':
                    caseid_v3.caseid_video05(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video06':
                    caseid_v3.caseid_video06(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video07':
                    caseid_v3.caseid_video07(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video08':
                    caseid_v3.caseid_video08(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video09':
                    caseid_v3.caseid_video09(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video10':
                    caseid_v3.caseid_video10(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video11':
                    caseid_v3.caseid_video11(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video12':
                    caseid_v3.caseid_video12(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video13':
                    caseid_v3.caseid_video13(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video14':
                    caseid_v3.caseid_video14(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video15':
                    caseid_v3.caseid_video15(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video16':
                    caseid_v3.caseid_video16(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video17':
                    caseid_v3.caseid_video17(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video18':
                    caseid_v3.caseid_video18(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video19':
                    caseid_v3.caseid_video19(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video20':
                    caseid_v3.caseid_video20(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video21':
                    caseid_v3.caseid_video21(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video22':
                    caseid_v3.caseid_video22(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video23':
                    caseid_v3.caseid_video23(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video24':
                    caseid_v3.caseid_video24(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video25':
                    caseid_v3.caseid_video25(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video26':
                    caseid_v3.caseid_video26(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video27':
                    caseid_v3.caseid_video27(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Video28':
                    caseid_v3.caseid_video28(self)
            except:
                module_other_v3.swich_tab_0()

            try:
                if case == 'Image01':
                    caseid_v3.caseid_image01(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image02':
                    caseid_v3.caseid_image02(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image03':
                    caseid_v3.caseid_image03(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image04':
                    caseid_v3.caseid_image04(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image05':
                    caseid_v3.caseid_image05(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image06':
                    caseid_v3.caseid_image06(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image07':
                    caseid_v3.caseid_image07(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image08':
                    caseid_v3.caseid_image08(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image09':
                    caseid_v3.caseid_image09(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image10':
                    caseid_v3.caseid_image10(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image11':
                    caseid_v3.caseid_image11(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image12':
                    caseid_v3.caseid_image12(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image13':
                    caseid_v3.caseid_image13(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image14':
                    caseid_v3.caseid_image14(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image15':
                    caseid_v3.caseid_image15(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image16':
                    caseid_v3.caseid_image16(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image17':
                    caseid_v3.caseid_image17(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image18':
                    caseid_v3.caseid_image18(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image19':
                    caseid_v3.caseid_image19(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image20':
                    caseid_v3.caseid_image20(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image21':
                    caseid_v3.caseid_image21(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image22':
                    caseid_v3.caseid_image22(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image23':
                    caseid_v3.caseid_image23(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image24':
                    caseid_v3.caseid_image24(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Image25':
                    caseid_v3.caseid_image25(self)
            except:
                module_other_v3.swich_tab_0()

            try:
                if case == 'Utility01':
                    caseid_v3.caseid_utility01(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility02':
                    caseid_v3.caseid_utility02(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility03':
                    caseid_v3.caseid_utility03(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility04':
                    caseid_v3.caseid_utility04(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility05':
                    caseid_v3.caseid_utility05(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility06':
                    caseid_v3.caseid_utility06(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility07':
                    caseid_v3.caseid_utility07(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility08':
                    caseid_v3.caseid_utility08(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility09':
                    caseid_v3.caseid_utility09(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility10':
                    caseid_v3.caseid_utility10(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility11':
                    caseid_v3.caseid_utility11(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility12':
                    caseid_v3.caseid_utility12(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility13':
                    caseid_v3.caseid_utility13(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility14':
                    caseid_v3.caseid_utility14(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility15':
                    caseid_v3.caseid_utility15(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility16':
                    caseid_v3.caseid_utility16(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility17':
                    caseid_v3.caseid_utility17(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility18':
                    caseid_v3.caseid_utility18(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility19':
                    caseid_v3.caseid_utility19(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility20':
                    caseid_v3.caseid_utility20(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility21':
                    caseid_v3.caseid_utility21(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility22':
                    caseid_v3.caseid_utility22(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility23':
                    caseid_v3.caseid_utility23(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility24':
                    caseid_v3.caseid_utility24(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Utility25':
                    caseid_v3.caseid_utility25(self)
            except:
                module_other_v3.swich_tab_0()

            try:
                if case == 'Ai01':
                    caseid_v3.caseid_ai01(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai02':
                    caseid_v3.caseid_ai02(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai03':
                    caseid_v3.caseid_ai03(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai04':
                    caseid_v3.caseid_ai04(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai05':
                    caseid_v3.caseid_ai05(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai06':
                    caseid_v3.caseid_ai06(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai07':
                    caseid_v3.caseid_ai07(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai08':
                    caseid_v3.caseid_ai08(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai09':
                    caseid_v3.caseid_ai09(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai10':
                    caseid_v3.caseid_ai10(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai11':
                    caseid_v3.caseid_ai11(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai12':
                    caseid_v3.caseid_ai12(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai13':
                    caseid_v3.caseid_ai13(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai14':
                    caseid_v3.caseid_ai14(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai15':
                    caseid_v3.caseid_ai15(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai16':
                    caseid_v3.caseid_ai16(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai17':
                    caseid_v3.caseid_ai17(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai18':
                    caseid_v3.caseid_ai18(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai19':
                    caseid_v3.caseid_ai19(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai20':
                    caseid_v3.caseid_ai20(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai21':
                    caseid_v3.caseid_ai21(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai22':
                    caseid_v3.caseid_ai22(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai23':
                    caseid_v3.caseid_ai23(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai24':
                    caseid_v3.caseid_ai24(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai25':
                    caseid_v3.caseid_ai25(self)
            except:
                module_other_v3.swich_tab_0()
            try:
                if case == 'Ai26':
                    caseid_v3.caseid_ai26(self)
            except:
                module_other_v3.swich_tab_0()



def retest_casefail(self):
    list_casefail = []
    wordbook = openpyxl.load_workbook(var_v3.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    while (rownum < 1000):
        rownum += 1
        rownum = str(rownum)
        if sheet["G"+rownum].value == "Fail":
            print(sheet["A"+rownum].value)
            case_fail = sheet["A"+rownum].value
            list_casefail.append(case_fail)
        rownum = int(rownum)
    print(list_casefail)
    for case in list_casefail:
        try:
            if case == 'Login01':
                caseid_v3.caseid_login01(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login02':
                caseid_v3.caseid_login02(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login03':
                caseid_v3.caseid_login03(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login04':
                caseid_v3.caseid_login04(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login05':
                caseid_v3.caseid_login05(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login06':
                caseid_v3.caseid_login06(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login07':
                caseid_v3.caseid_login07(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login08':
                caseid_v3.caseid_login08(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login09':
                caseid_v3.caseid_login09(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login10':
                caseid_v3.caseid_login10(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login11':
                caseid_v3.caseid_login11(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login12':
                caseid_v3.caseid_login12(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login13':
                caseid_v3.caseid_login13(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login14':
                caseid_v3.caseid_login14(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login15':
                caseid_v3.caseid_login15(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login16':
                caseid_v3.caseid_login16(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login17':
                caseid_v3.caseid_login17(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login18':
                caseid_v3.caseid_login18(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login19':
                caseid_v3.caseid_login19(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login20':
                caseid_v3.caseid_login20(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login21':
                caseid_v3.caseid_login21(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login22':
                caseid_v3.caseid_login22(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login23':
                caseid_v3.caseid_login23(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login24':
                caseid_v3.caseid_login24(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login25':
                caseid_v3.caseid_login25(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login26':
                caseid_v3.caseid_login26(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat01':
                caseid_v3.caseid_giamsat01(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat02':
                caseid_v3.caseid_giamsat02(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat03':
                caseid_v3.caseid_giamsat03(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat04':
                caseid_v3.caseid_giamsat04(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat05':
                caseid_v3.caseid_giamsat05(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat06':
                caseid_v3.caseid_giamsat06(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat07':
                caseid_v3.caseid_giamsat07(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat08':
                caseid_v3.caseid_giamsat08(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat09':
                caseid_v3.caseid_giamsat09(self)
        except:
            module_other_v3.swich_tab_0()

        try:
            if case == 'GiamSat10':
                caseid_v3.caseid_giamsat10(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat11':
                caseid_v3.caseid_giamsat11(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat12':
                caseid_v3.caseid_giamsat12(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat13':
                caseid_v3.caseid_giamsat13(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat14':
                caseid_v3.caseid_giamsat14(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat15':
                caseid_v3.caseid_giamsat15(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat16':
                caseid_v3.caseid_giamsat16(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat17':
                caseid_v3.caseid_giamsat17(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat18':
                caseid_v3.caseid_giamsat18(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat19':
                caseid_v3.caseid_giamsat19(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat20':
                caseid_v3.caseid_giamsat20(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat21':
                caseid_v3.caseid_giamsat21(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat22':
                caseid_v3.caseid_giamsat22(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat23':
                caseid_v3.caseid_giamsat23(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat24':
                caseid_v3.caseid_giamsat24(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat25':
                caseid_v3.caseid_giamsat25(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat26':
                caseid_v3.caseid_giamsat26(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat27':
                caseid_v3.caseid_giamsat27(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat28':
                caseid_v3.caseid_giamsat28(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat29':
                caseid_v3.caseid_giamsat29(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat30':
                caseid_v3.caseid_giamsat30(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat31':
                caseid_v3.caseid_giamsat31(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat32':
                caseid_v3.caseid_giamsat32(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat33':
                caseid_v3.caseid_giamsat33(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat34':
                caseid_v3.caseid_giamsat34(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat35':
                caseid_v3.caseid_giamsat35(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat36':
                caseid_v3.caseid_giamsat36(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat37':
                caseid_v3.caseid_giamsat37(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat38':
                caseid_v3.caseid_giamsat38(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat39':
                caseid_v3.caseid_giamsat39(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat40':
                caseid_v3.caseid_giamsat40(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat41':
                caseid_v3.caseid_giamsat41(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat42':
                caseid_v3.caseid_giamsat42(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat43':
                caseid_v3.caseid_giamsat43(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat44':
                caseid_v3.caseid_giamsat44(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat45':
                caseid_v3.caseid_giamsat45(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat46':
                caseid_v3.caseid_giamsat46(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat47':
                caseid_v3.caseid_giamsat47(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat48':
                caseid_v3.caseid_giamsat48(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat49':
                caseid_v3.caseid_giamsat49(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat50':
                caseid_v3.caseid_giamsat50(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat51':
                caseid_v3.caseid_giamsat51(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat52':
                caseid_v3.caseid_giamsat52(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat53':
                caseid_v3.caseid_giamsat53(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat54':
                caseid_v3.caseid_giamsat54(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat55':
                caseid_v3.caseid_giamsat55(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat56':
                caseid_v3.caseid_giamsat56(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat57':
                caseid_v3.caseid_giamsat57(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat58':
                caseid_v3.caseid_giamsat58(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat59':
                caseid_v3.caseid_giamsat59(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat60':
                caseid_v3.caseid_giamsat60(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat61':
                caseid_v3.caseid_giamsat61(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat62':
                caseid_v3.caseid_giamsat62(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat63':
                caseid_v3.caseid_giamsat63(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat64':
                caseid_v3.caseid_giamsat64(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat65':
                caseid_v3.caseid_giamsat65(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat66':
                caseid_v3.caseid_giamsat66(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat67':
                caseid_v3.caseid_giamsat67(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat68':
                caseid_v3.caseid_giamsat68(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat69':
                caseid_v3.caseid_giamsat69(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat70':
                caseid_v3.caseid_giamsat70(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat71':
                caseid_v3.caseid_giamsat71(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat72':
                caseid_v3.caseid_giamsat72(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat73':
                caseid_v3.caseid_giamsat73(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat74':
                caseid_v3.caseid_giamsat74(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat75':
                caseid_v3.caseid_giamsat75(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat76':
                caseid_v3.caseid_giamsat76(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat77':
                caseid_v3.caseid_giamsat77(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat78':
                caseid_v3.caseid_giamsat78(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat79':
                caseid_v3.caseid_giamsat79(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat80':
                caseid_v3.caseid_giamsat80(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat81':
                caseid_v3.caseid_giamsat81(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat82':
                caseid_v3.caseid_giamsat82(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat83':
                caseid_v3.caseid_giamsat83(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat84':
                caseid_v3.caseid_giamsat84(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat85':
                caseid_v3.caseid_giamsat85(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat86':
                caseid_v3.caseid_giamsat86(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat87':
                caseid_v3.caseid_giamsat87(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat88':
                caseid_v3.caseid_giamsat88(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat89':
                caseid_v3.caseid_giamsat89(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat90':
                caseid_v3.caseid_giamsat90(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat91':
                caseid_v3.caseid_giamsat91(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat92':
                caseid_v3.caseid_giamsat92(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat93':
                caseid_v3.caseid_giamsat93(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat94':
                caseid_v3.caseid_giamsat94(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat95':
                caseid_v3.caseid_giamsat95(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat96':
                caseid_v3.caseid_giamsat96(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat97':
                caseid_v3.caseid_giamsat97(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat98':
                caseid_v3.caseid_giamsat98(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat99':
                caseid_v3.caseid_giamsat99(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat100':
                caseid_v3.caseid_giamsat100(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat101':
                caseid_v3.caseid_giamsat101(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat102':
                caseid_v3.caseid_giamsat102(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat103':
                caseid_v3.caseid_giamsat103(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat104':
                caseid_v3.caseid_giamsat104(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat105':
                caseid_v3.caseid_giamsat105(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat106':
                caseid_v3.caseid_giamsat106(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat107':
                caseid_v3.caseid_giamsat107(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat108':
                caseid_v3.caseid_giamsat108(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat109':
                caseid_v3.caseid_giamsat109(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat110':
                caseid_v3.caseid_giamsat110(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat111':
                caseid_v3.caseid_giamsat111(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat112':
                caseid_v3.caseid_giamsat112(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat113':
                caseid_v3.caseid_giamsat113(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat114':
                caseid_v3.caseid_giamsat114(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat115':
                caseid_v3.caseid_giamsat115(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat116':
                caseid_v3.caseid_giamsat116(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat117':
                caseid_v3.caseid_giamsat117(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat118':
                caseid_v3.caseid_giamsat118(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat119':
                caseid_v3.caseid_giamsat119(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat120':
                caseid_v3.caseid_giamsat120(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat121':
                caseid_v3.caseid_giamsat121(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat122':
                caseid_v3.caseid_giamsat122(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat123':
                caseid_v3.caseid_giamsat123(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat124':
                caseid_v3.caseid_giamsat124(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat125':
                caseid_v3.caseid_giamsat125(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat126':
                caseid_v3.caseid_giamsat126(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat127':
                caseid_v3.caseid_giamsat127(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat128':
                caseid_v3.caseid_giamsat128(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat129':
                caseid_v3.caseid_giamsat129(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat130':
                caseid_v3.caseid_giamsat130(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat131':
                caseid_v3.caseid_giamsat131(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat132':
                caseid_v3.caseid_giamsat132(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat133':
                caseid_v3.caseid_giamsat133(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat134':
                caseid_v3.caseid_giamsat134(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat135':
                caseid_v3.caseid_giamsat135(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat136':
                caseid_v3.caseid_giamsat136(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat137':
                caseid_v3.caseid_giamsat137(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat138':
                caseid_v3.caseid_giamsat138(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat139':
                caseid_v3.caseid_giamsat139(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat140':
                caseid_v3.caseid_giamsat140(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat141':
                caseid_v3.caseid_giamsat141(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat142':
                caseid_v3.caseid_giamsat142(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat143':
                caseid_v3.caseid_giamsat143(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat144':
                caseid_v3.caseid_giamsat144(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat145':
                caseid_v3.caseid_giamsat145(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat146':
                caseid_v3.caseid_giamsat146(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat147':
                caseid_v3.caseid_giamsat147(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat148':
                caseid_v3.caseid_giamsat148(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat149':
                caseid_v3.caseid_giamsat149(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat150':
                caseid_v3.caseid_giamsat150(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat151':
                caseid_v3.caseid_giamsat151(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat152':
                caseid_v3.caseid_giamsat152(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat153':
                caseid_v3.caseid_giamsat153(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat154':
                caseid_v3.caseid_giamsat154(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat155':
                caseid_v3.caseid_giamsat155(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat156':
                caseid_v3.caseid_giamsat156(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat157':
                caseid_v3.caseid_giamsat157(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat158':
                caseid_v3.caseid_giamsat158(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat159':
                caseid_v3.caseid_giamsat159(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat160':
                caseid_v3.caseid_giamsat160(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat161':
                caseid_v3.caseid_giamsat161(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat162':
                caseid_v3.caseid_giamsat162(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat163':
                caseid_v3.caseid_giamsat163(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat164':
                caseid_v3.caseid_giamsat164(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat165':
                caseid_v3.caseid_giamsat165(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat166':
                caseid_v3.caseid_giamsat166(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat167':
                caseid_v3.caseid_giamsat167(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat168':
                caseid_v3.caseid_giamsat168(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat169':
                caseid_v3.caseid_giamsat169(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat170':
                caseid_v3.caseid_giamsat170(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat171':
                caseid_v3.caseid_giamsat171(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat172':
                caseid_v3.caseid_giamsat172(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat173':
                caseid_v3.caseid_giamsat173(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat174':
                caseid_v3.caseid_giamsat174(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat175':
                caseid_v3.caseid_giamsat175(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat176':
                caseid_v3.caseid_giamsat176(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat177':
                caseid_v3.caseid_giamsat177(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat178':
                caseid_v3.caseid_giamsat178(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat179':
                caseid_v3.caseid_giamsat179(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat180':
                caseid_v3.caseid_giamsat180(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat181':
                caseid_v3.caseid_giamsat181(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat182':
                caseid_v3.caseid_giamsat182(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat183':
                caseid_v3.caseid_giamsat183(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat184':
                caseid_v3.caseid_giamsat184(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat185':
                caseid_v3.caseid_giamsat185(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat186':
                caseid_v3.caseid_giamsat186(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat187':
                caseid_v3.caseid_giamsat187(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat188':
                caseid_v3.caseid_giamsat188(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat189':
                caseid_v3.caseid_giamsat189(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat190':
                caseid_v3.caseid_giamsat190(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat191':
                caseid_v3.caseid_giamsat191(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat192':
                caseid_v3.caseid_giamsat192(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat193':
                caseid_v3.caseid_giamsat193(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat194':
                caseid_v3.caseid_giamsat194(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat195':
                caseid_v3.caseid_giamsat195(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat196':
                caseid_v3.caseid_giamsat196(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat197':
                caseid_v3.caseid_giamsat197(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat198':
                caseid_v3.caseid_giamsat198(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat199':
                caseid_v3.caseid_giamsat199(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat200':
                caseid_v3.caseid_giamsat200(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat201':
                caseid_v3.caseid_giamsat201(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat202':
                caseid_v3.caseid_giamsat202(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat203':
                caseid_v3.caseid_giamsat203(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat204':
                caseid_v3.caseid_giamsat204(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat205':
                caseid_v3.caseid_giamsat205(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat206':
                caseid_v3.caseid_giamsat206(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat207':
                caseid_v3.caseid_giamsat207(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat208':
                caseid_v3.caseid_giamsat208(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat209':
                caseid_v3.caseid_giamsat209(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat210':
                caseid_v3.caseid_giamsat210(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat211':
                caseid_v3.caseid_giamsat211(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat212':
                caseid_v3.caseid_giamsat212(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat213':
                caseid_v3.caseid_giamsat213(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat214':
                caseid_v3.caseid_giamsat214(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat215':
                caseid_v3.caseid_giamsat215(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat216':
                caseid_v3.caseid_giamsat216(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat217':
                caseid_v3.caseid_giamsat217(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat218':
                caseid_v3.caseid_giamsat218(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat219':
                caseid_v3.caseid_giamsat219(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat220':
                caseid_v3.caseid_giamsat220(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat221':
                caseid_v3.caseid_giamsat221(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat222':
                caseid_v3.caseid_giamsat222(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat223':
                caseid_v3.caseid_giamsat223(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat224':
                caseid_v3.caseid_giamsat224(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat225':
                caseid_v3.caseid_giamsat225(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat226':
                caseid_v3.caseid_giamsat226(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat227':
                caseid_v3.caseid_giamsat227(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat228':
                caseid_v3.caseid_giamsat228(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat229':
                caseid_v3.caseid_giamsat229(self)
        except:
            module_other_v3.swich_tab_0()

        try:
            if case == 'Route01':
                caseid_v3.caseid_route01(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route02':
                caseid_v3.caseid_route02(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route03':
                caseid_v3.caseid_route03(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route04':
                caseid_v3.caseid_route04(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route05':
                caseid_v3.caseid_route05(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route06':
                caseid_v3.caseid_route06(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route07':
                caseid_v3.caseid_route07(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route08':
                caseid_v3.caseid_route08(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route09':
                caseid_v3.caseid_route09(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route10':
                caseid_v3.caseid_route10(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route11':
                caseid_v3.caseid_route11(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route12':
                caseid_v3.caseid_route12(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route13':
                caseid_v3.caseid_route13(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route14':
                caseid_v3.caseid_route14(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route15':
                caseid_v3.caseid_route15(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route16':
                caseid_v3.caseid_route16(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route17':
                caseid_v3.caseid_route17(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route18':
                caseid_v3.caseid_route18(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route19':
                caseid_v3.caseid_route19(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route20':
                caseid_v3.caseid_route20(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route21':
                caseid_v3.caseid_route21(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route22':
                caseid_v3.caseid_route22(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route23':
                caseid_v3.caseid_route23(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route24':
                caseid_v3.caseid_route24(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route25':
                caseid_v3.caseid_route25(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route26':
                caseid_v3.caseid_route26(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route27':
                caseid_v3.caseid_route27(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route28':
                caseid_v3.caseid_route28(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route29':
                caseid_v3.caseid_route29(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route30':
                caseid_v3.caseid_route30(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route31':
                caseid_v3.caseid_route31(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route32':
                caseid_v3.caseid_route32(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route33':
                caseid_v3.caseid_route33(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route34':
                caseid_v3.caseid_route34(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route35':
                caseid_v3.caseid_route35(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route36':
                caseid_v3.caseid_route36(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route37':
                caseid_v3.caseid_route37(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route38':
                caseid_v3.caseid_route38(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route39':
                caseid_v3.caseid_route39(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route40':
                caseid_v3.caseid_route40(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route41':
                caseid_v3.caseid_route41(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route42':
                caseid_v3.caseid_route42(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route43':
                caseid_v3.caseid_route43(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route44':
                caseid_v3.caseid_route44(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route45':
                caseid_v3.caseid_route45(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route46':
                caseid_v3.caseid_route46(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route47':
                caseid_v3.caseid_route47(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route48':
                caseid_v3.caseid_route48(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route49':
                caseid_v3.caseid_route49(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route50':
                caseid_v3.caseid_route50(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route51':
                caseid_v3.caseid_route51(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route52':
                caseid_v3.caseid_route52(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route53':
                caseid_v3.caseid_route53(self)
        except:
            module_other_v3.swich_tab_0()

        try:
            if case == 'Admin01':
                caseid_v3.caseid_admin01(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin02':
                caseid_v3.caseid_admin02(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin03':
                caseid_v3.caseid_admin03(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin04':
                caseid_v3.caseid_admin04(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin05':
                caseid_v3.caseid_admin05(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin06':
                caseid_v3.caseid_admin06(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin07':
                caseid_v3.caseid_admin07(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin08':
                caseid_v3.caseid_admin08(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin09':
                caseid_v3.caseid_admin09(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin10':
                caseid_v3.caseid_admin10(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin11':
                caseid_v3.caseid_admin11(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin12':
                caseid_v3.caseid_admin12(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin13':
                caseid_v3.caseid_admin13(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin14':
                caseid_v3.caseid_admin14(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin15':
                caseid_v3.caseid_admin15(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin16':
                caseid_v3.caseid_admin16(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin17':
                caseid_v3.caseid_admin17(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin18':
                caseid_v3.caseid_admin18(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin19':
                caseid_v3.caseid_admin19(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin20':
                caseid_v3.caseid_admin20(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin21':
                caseid_v3.caseid_admin21(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin22':
                caseid_v3.caseid_admin22(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin23':
                caseid_v3.caseid_admin23(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin24':
                caseid_v3.caseid_admin24(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin25':
                caseid_v3.caseid_admin25(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin26':
                caseid_v3.caseid_admin26(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin27':
                caseid_v3.caseid_admin27(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin28':
                caseid_v3.caseid_admin28(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin29':
                caseid_v3.caseid_admin29(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin30':
                caseid_v3.caseid_admin30(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin31':
                caseid_v3.caseid_admin31(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin32':
                caseid_v3.caseid_admin32(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin33':
                caseid_v3.caseid_admin33(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin34':
                caseid_v3.caseid_admin34(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin35':
                caseid_v3.caseid_admin35(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin36':
                caseid_v3.caseid_admin36(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin37':
                caseid_v3.caseid_admin37(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin38':
                caseid_v3.caseid_admin38(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin39':
                caseid_v3.caseid_admin39(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin40':
                caseid_v3.caseid_admin40(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin41':
                caseid_v3.caseid_admin41(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin42':
                caseid_v3.caseid_admin42(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin43':
                caseid_v3.caseid_admin43(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin44':
                caseid_v3.caseid_admin44(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin45':
                caseid_v3.caseid_admin45(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin46':
                caseid_v3.caseid_admin46(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin47':
                caseid_v3.caseid_admin47(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin48':
                caseid_v3.caseid_admin48(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin49':
                caseid_v3.caseid_admin49(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin50':
                caseid_v3.caseid_admin50(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin51':
                caseid_v3.caseid_admin51(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin52':
                caseid_v3.caseid_admin52(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin53':
                caseid_v3.caseid_admin53(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin54':
                caseid_v3.caseid_admin54(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin55':
                caseid_v3.caseid_admin55(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin56':
                caseid_v3.caseid_admin56(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin57':
                caseid_v3.caseid_admin57(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin58':
                caseid_v3.caseid_admin58(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin59':
                caseid_v3.caseid_admin59(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin60':
                caseid_v3.caseid_admin60(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin61':
                caseid_v3.caseid_admin61(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin62':
                caseid_v3.caseid_admin62(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin63':
                caseid_v3.caseid_admin63(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin64':
                caseid_v3.caseid_admin64(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin65':
                caseid_v3.caseid_admin65(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin66':
                caseid_v3.caseid_admin66(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin67':
                caseid_v3.caseid_admin67(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin68':
                caseid_v3.caseid_admin68(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin69':
                caseid_v3.caseid_admin69(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin70':
                caseid_v3.caseid_admin70(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin71':
                caseid_v3.caseid_admin71(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin72':
                caseid_v3.caseid_admin72(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin73':
                caseid_v3.caseid_admin73(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin74':
                caseid_v3.caseid_admin74(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin75':
                caseid_v3.caseid_admin75(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin76':
                caseid_v3.caseid_admin76(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin77':
                caseid_v3.caseid_admin77(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin78':
                caseid_v3.caseid_admin78(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin79':
                caseid_v3.caseid_admin79(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin80':
                caseid_v3.caseid_admin80(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin81':
                caseid_v3.caseid_admin81(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin82':
                caseid_v3.caseid_admin82(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin83':
                caseid_v3.caseid_admin83(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin84':
                caseid_v3.caseid_admin84(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin85':
                caseid_v3.caseid_admin85(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin86':
                caseid_v3.caseid_admin86(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin87':
                caseid_v3.caseid_admin87(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin88':
                caseid_v3.caseid_admin88(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin89':
                caseid_v3.caseid_admin89(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin90':
                caseid_v3.caseid_admin90(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin91':
                caseid_v3.caseid_admin91(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin92':
                caseid_v3.caseid_admin92(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin93':
                caseid_v3.caseid_admin93(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin94':
                caseid_v3.caseid_admin94(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin95':
                caseid_v3.caseid_admin95(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin96':
                caseid_v3.caseid_admin96(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin97':
                caseid_v3.caseid_admin97(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin98':
                caseid_v3.caseid_admin98(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin99':
                caseid_v3.caseid_admin99(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin100':
                caseid_v3.caseid_admin100(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin101':
                caseid_v3.caseid_admin101(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin102':
                caseid_v3.caseid_admin102(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin103':
                caseid_v3.caseid_admin103(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin104':
                caseid_v3.caseid_admin104(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin105':
                caseid_v3.caseid_admin105(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin106':
                caseid_v3.caseid_admin106(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin107':
                caseid_v3.caseid_admin107(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin108':
                caseid_v3.caseid_admin108(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin109':
                caseid_v3.caseid_admin109(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin110':
                caseid_v3.caseid_admin110(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin111':
                caseid_v3.caseid_admin111(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin112':
                caseid_v3.caseid_admin112(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin113':
                caseid_v3.caseid_admin113(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin114':
                caseid_v3.caseid_admin114(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin115':
                caseid_v3.caseid_admin115(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin116':
                caseid_v3.caseid_admin116(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin117':
                caseid_v3.caseid_admin117(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin118':
                caseid_v3.caseid_admin118(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin119':
                caseid_v3.caseid_admin119(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin120':
                caseid_v3.caseid_admin120(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin121':
                caseid_v3.caseid_admin121(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin122':
                caseid_v3.caseid_admin122(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin123':
                caseid_v3.caseid_admin123(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin124':
                caseid_v3.caseid_admin124(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin125':
                caseid_v3.caseid_admin125(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin126':
                caseid_v3.caseid_admin126(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin127':
                caseid_v3.caseid_admin127(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin128':
                caseid_v3.caseid_admin128(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin129':
                caseid_v3.caseid_admin129(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin130':
                caseid_v3.caseid_admin130(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin131':
                caseid_v3.caseid_admin131(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin132':
                caseid_v3.caseid_admin132(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin133':
                caseid_v3.caseid_admin133(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin134':
                caseid_v3.caseid_admin134(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin135':
                caseid_v3.caseid_admin135(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin136':
                caseid_v3.caseid_admin136(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin137':
                caseid_v3.caseid_admin137(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin138':
                caseid_v3.caseid_admin138(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin139':
                caseid_v3.caseid_admin139(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin140':
                caseid_v3.caseid_admin140(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin141':
                caseid_v3.caseid_admin141(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin142':
                caseid_v3.caseid_admin142(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin143':
                caseid_v3.caseid_admin143(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin144':
                caseid_v3.caseid_admin144(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin145':
                caseid_v3.caseid_admin145(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin146':
                caseid_v3.caseid_admin146(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin147':
                caseid_v3.caseid_admin147(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin148':
                caseid_v3.caseid_admin148(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin149':
                caseid_v3.caseid_admin149(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin150':
                caseid_v3.caseid_admin150(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin151':
                caseid_v3.caseid_admin151(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin152':
                caseid_v3.caseid_admin152(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin153':
                caseid_v3.caseid_admin153(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin154':
                caseid_v3.caseid_admin154(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin155':
                caseid_v3.caseid_admin155(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin156':
                caseid_v3.caseid_admin156(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin157':
                caseid_v3.caseid_admin157(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin158':
                caseid_v3.caseid_admin158(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin159':
                caseid_v3.caseid_admin159(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin160':
                caseid_v3.caseid_admin160(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin161':
                caseid_v3.caseid_admin161(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin162':
                caseid_v3.caseid_admin162(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin163':
                caseid_v3.caseid_admin163(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin164':
                caseid_v3.caseid_admin164(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin165':
                caseid_v3.caseid_admin165(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin166':
                caseid_v3.caseid_admin166(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin167':
                caseid_v3.caseid_admin167(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin168':
                caseid_v3.caseid_admin168(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin169':
                caseid_v3.caseid_admin169(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin170':
                caseid_v3.caseid_admin170(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin171':
                caseid_v3.caseid_admin171(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin172':
                caseid_v3.caseid_admin172(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin173':
                caseid_v3.caseid_admin173(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin174':
                caseid_v3.caseid_admin174(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin175':
                caseid_v3.caseid_admin175(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin176':
                caseid_v3.caseid_admin176(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin177':
                caseid_v3.caseid_admin177(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin178':
                caseid_v3.caseid_admin178(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin179':
                caseid_v3.caseid_admin179(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin180':
                caseid_v3.caseid_admin180(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin181':
                caseid_v3.caseid_admin181(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin182':
                caseid_v3.caseid_admin182(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin183':
                caseid_v3.caseid_admin183(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin184':
                caseid_v3.caseid_admin184(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin185':
                caseid_v3.caseid_admin185(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin186':
                caseid_v3.caseid_admin186(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin187':
                caseid_v3.caseid_admin187(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin188':
                caseid_v3.caseid_admin188(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin189':
                caseid_v3.caseid_admin189(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin190':
                caseid_v3.caseid_admin190(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin191':
                caseid_v3.caseid_admin191(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin192':
                caseid_v3.caseid_admin192(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin193':
                caseid_v3.caseid_admin193(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin194':
                caseid_v3.caseid_admin194(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin195':
                caseid_v3.caseid_admin195(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin196':
                caseid_v3.caseid_admin196(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin197':
                caseid_v3.caseid_admin197(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin198':
                caseid_v3.caseid_admin198(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin199':
                caseid_v3.caseid_admin199(self)
        except:
            module_other_v3.swich_tab_0()

        try:
            if case == 'Report01':
                caseid_v3.caseid_report01(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report02':
                caseid_v3.caseid_report02(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report03':
                caseid_v3.caseid_report03(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report04':
                caseid_v3.caseid_report04(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report05':
                caseid_v3.caseid_report05(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report06':
                caseid_v3.caseid_report06(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report07':
                caseid_v3.caseid_report07(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report08':
                caseid_v3.caseid_report08(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report09':
                caseid_v3.caseid_report09(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report10':
                caseid_v3.caseid_report10(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report11':
                caseid_v3.caseid_report11(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report12':
                caseid_v3.caseid_report12(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report13':
                caseid_v3.caseid_report13(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report14':
                caseid_v3.caseid_report14(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report15':
                caseid_v3.caseid_report15(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report16':
                caseid_v3.caseid_report16(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report17':
                caseid_v3.caseid_report17(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report18':
                caseid_v3.caseid_report18(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report19':
                caseid_v3.caseid_report19(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report20':
                caseid_v3.caseid_report20(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report21':
                caseid_v3.caseid_report21(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report22':
                caseid_v3.caseid_report22(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report23':
                caseid_v3.caseid_report23(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report24':
                caseid_v3.caseid_report24(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report25':
                caseid_v3.caseid_report25(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report26':
                caseid_v3.caseid_report26(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report27':
                caseid_v3.caseid_report27(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report28':
                caseid_v3.caseid_report28(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report29':
                caseid_v3.caseid_report29(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report30':
                caseid_v3.caseid_report30(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report31':
                caseid_v3.caseid_report31(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report32':
                caseid_v3.caseid_report32(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report33':
                caseid_v3.caseid_report33(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report34':
                caseid_v3.caseid_report34(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report35':
                caseid_v3.caseid_report35(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report36':
                caseid_v3.caseid_report36(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report37':
                caseid_v3.caseid_report37(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report38':
                caseid_v3.caseid_report38(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report39':
                caseid_v3.caseid_report39(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report40':
                caseid_v3.caseid_report40(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report41':
                caseid_v3.caseid_report41(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report42':
                caseid_v3.caseid_report42(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report43':
                caseid_v3.caseid_report43(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report44':
                caseid_v3.caseid_report44(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report45':
                caseid_v3.caseid_report45(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report46':
                caseid_v3.caseid_report46(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report47':
                caseid_v3.caseid_report47(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report48':
                caseid_v3.caseid_report48(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report49':
                caseid_v3.caseid_report49(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report50':
                caseid_v3.caseid_report50(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report51':
                caseid_v3.caseid_report51(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report52':
                caseid_v3.caseid_report52(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report53':
                caseid_v3.caseid_report53(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report54':
                caseid_v3.caseid_report54(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report55':
                caseid_v3.caseid_report55(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report56':
                caseid_v3.caseid_report56(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report57':
                caseid_v3.caseid_report57(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report58':
                caseid_v3.caseid_report58(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report59':
                caseid_v3.caseid_report59(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report60':
                caseid_v3.caseid_report60(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report61':
                caseid_v3.caseid_report61(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report62':
                caseid_v3.caseid_report62(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report63':
                caseid_v3.caseid_report63(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report64':
                caseid_v3.caseid_report64(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report65':
                caseid_v3.caseid_report65(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report66':
                caseid_v3.caseid_report66(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report67':
                caseid_v3.caseid_report67(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report68':
                caseid_v3.caseid_report68(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report69':
                caseid_v3.caseid_report69(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report70':
                caseid_v3.caseid_report70(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report71':
                caseid_v3.caseid_report71(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report72':
                caseid_v3.caseid_report72(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report73':
                caseid_v3.caseid_report73(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report74':
                caseid_v3.caseid_report74(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report75':
                caseid_v3.caseid_report75(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report76':
                caseid_v3.caseid_report76(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report77':
                caseid_v3.caseid_report77(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report78':
                caseid_v3.caseid_report78(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report79':
                caseid_v3.caseid_report79(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report80':
                caseid_v3.caseid_report80(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report81':
                caseid_v3.caseid_report81(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report82':
                caseid_v3.caseid_report82(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report83':
                caseid_v3.caseid_report83(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report84':
                caseid_v3.caseid_report84(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report85':
                caseid_v3.caseid_report85(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report86':
                caseid_v3.caseid_report86(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report87':
                caseid_v3.caseid_report87(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report88':
                caseid_v3.caseid_report88(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report89':
                caseid_v3.caseid_report89(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report90':
                caseid_v3.caseid_report90(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report91':
                caseid_v3.caseid_report91(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report92':
                caseid_v3.caseid_report92(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report93':
                caseid_v3.caseid_report93(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report94':
                caseid_v3.caseid_report94(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report95':
                caseid_v3.caseid_report95(self)
        except:
            module_other_v3.swich_tab_0()

        try:
            if case == 'Video01':
                caseid_v3.caseid_video01(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video02':
                caseid_v3.caseid_video02(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video03':
                caseid_v3.caseid_video03(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video04':
                caseid_v3.caseid_video04(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video05':
                caseid_v3.caseid_video05(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video06':
                caseid_v3.caseid_video06(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video07':
                caseid_v3.caseid_video07(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video08':
                caseid_v3.caseid_video08(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video09':
                caseid_v3.caseid_video09(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video10':
                caseid_v3.caseid_video10(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video11':
                caseid_v3.caseid_video11(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video12':
                caseid_v3.caseid_video12(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video13':
                caseid_v3.caseid_video13(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video14':
                caseid_v3.caseid_video14(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video15':
                caseid_v3.caseid_video15(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video16':
                caseid_v3.caseid_video16(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video17':
                caseid_v3.caseid_video17(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video18':
                caseid_v3.caseid_video18(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video19':
                caseid_v3.caseid_video19(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video20':
                caseid_v3.caseid_video20(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video21':
                caseid_v3.caseid_video21(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video22':
                caseid_v3.caseid_video22(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video23':
                caseid_v3.caseid_video23(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video24':
                caseid_v3.caseid_video24(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video25':
                caseid_v3.caseid_video25(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video26':
                caseid_v3.caseid_video26(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video27':
                caseid_v3.caseid_video27(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video28':
                caseid_v3.caseid_video28(self)
        except:
            module_other_v3.swich_tab_0()

        try:
            if case == 'Image01':
                caseid_v3.caseid_image01(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image02':
                caseid_v3.caseid_image02(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image03':
                caseid_v3.caseid_image03(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image04':
                caseid_v3.caseid_image04(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image05':
                caseid_v3.caseid_image05(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image06':
                caseid_v3.caseid_image06(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image07':
                caseid_v3.caseid_image07(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image08':
                caseid_v3.caseid_image08(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image09':
                caseid_v3.caseid_image09(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image10':
                caseid_v3.caseid_image10(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image11':
                caseid_v3.caseid_image11(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image12':
                caseid_v3.caseid_image12(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image13':
                caseid_v3.caseid_image13(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image14':
                caseid_v3.caseid_image14(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image15':
                caseid_v3.caseid_image15(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image16':
                caseid_v3.caseid_image16(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image17':
                caseid_v3.caseid_image17(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image18':
                caseid_v3.caseid_image18(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image19':
                caseid_v3.caseid_image19(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image20':
                caseid_v3.caseid_image20(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image21':
                caseid_v3.caseid_image21(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image22':
                caseid_v3.caseid_image22(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image23':
                caseid_v3.caseid_image23(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image24':
                caseid_v3.caseid_image24(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image25':
                caseid_v3.caseid_image25(self)
        except:
            module_other_v3.swich_tab_0()

        try:
            if case == 'Utility01':
                caseid_v3.caseid_utility01(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility02':
                caseid_v3.caseid_utility02(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility03':
                caseid_v3.caseid_utility03(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility04':
                caseid_v3.caseid_utility04(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility05':
                caseid_v3.caseid_utility05(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility06':
                caseid_v3.caseid_utility06(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility07':
                caseid_v3.caseid_utility07(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility08':
                caseid_v3.caseid_utility08(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility09':
                caseid_v3.caseid_utility09(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility10':
                caseid_v3.caseid_utility10(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility11':
                caseid_v3.caseid_utility11(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility12':
                caseid_v3.caseid_utility12(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility13':
                caseid_v3.caseid_utility13(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility14':
                caseid_v3.caseid_utility14(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility15':
                caseid_v3.caseid_utility15(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility16':
                caseid_v3.caseid_utility16(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility17':
                caseid_v3.caseid_utility17(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility18':
                caseid_v3.caseid_utility18(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility19':
                caseid_v3.caseid_utility19(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility20':
                caseid_v3.caseid_utility20(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility21':
                caseid_v3.caseid_utility21(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility22':
                caseid_v3.caseid_utility22(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility23':
                caseid_v3.caseid_utility23(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility24':
                caseid_v3.caseid_utility24(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility25':
                caseid_v3.caseid_utility25(self)
        except:
            module_other_v3.swich_tab_0()

        try:
            if case == 'Ai01':
                caseid_v3.caseid_ai01(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai02':
                caseid_v3.caseid_ai02(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai03':
                caseid_v3.caseid_ai03(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai04':
                caseid_v3.caseid_ai04(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai05':
                caseid_v3.caseid_ai05(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai06':
                caseid_v3.caseid_ai06(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai07':
                caseid_v3.caseid_ai07(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai08':
                caseid_v3.caseid_ai08(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai09':
                caseid_v3.caseid_ai09(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai10':
                caseid_v3.caseid_ai10(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai11':
                caseid_v3.caseid_ai11(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai12':
                caseid_v3.caseid_ai12(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai13':
                caseid_v3.caseid_ai13(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai14':
                caseid_v3.caseid_ai14(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai15':
                caseid_v3.caseid_ai15(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai16':
                caseid_v3.caseid_ai16(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai17':
                caseid_v3.caseid_ai17(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai18':
                caseid_v3.caseid_ai18(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai19':
                caseid_v3.caseid_ai19(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai20':
                caseid_v3.caseid_ai20(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai21':
                caseid_v3.caseid_ai21(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai22':
                caseid_v3.caseid_ai22(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai23':
                caseid_v3.caseid_ai23(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai24':
                caseid_v3.caseid_ai24(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai25':
                caseid_v3.caseid_ai25(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai26':
                caseid_v3.caseid_ai26(self)
        except:
            module_other_v3.swich_tab_0()



def run_all(self):
    list_0 = []
    wordbook = openpyxl.load_workbook(var_v3.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    while (rownum < 1000):
        rownum += 1
        rownum = str(rownum)
        print(sheet["G" + rownum].value)
        print(sheet["H" + rownum].value)
        if (sheet["H" + rownum].value == "x" or sheet["I" + rownum].value == "x" or sheet[ "J" + rownum].value == "x"
            or sheet["K" + rownum].value == "x") and sheet["A" + rownum].value != None:
            print(sheet["A" + rownum].value)
            case_0 = sheet["A" + rownum].value
            list_0.append(case_0)
        rownum = int(rownum)
    print(list_0)
    for case in list_0:
        try:
            if case == 'Login01':
                caseid_v3.caseid_login01(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login02':
                caseid_v3.caseid_login02(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login03':
                caseid_v3.caseid_login03(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login04':
                caseid_v3.caseid_login04(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login05':
                caseid_v3.caseid_login05(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login06':
                caseid_v3.caseid_login06(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login07':
                caseid_v3.caseid_login07(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login08':
                caseid_v3.caseid_login08(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login09':
                caseid_v3.caseid_login09(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login10':
                caseid_v3.caseid_login10(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login11':
                caseid_v3.caseid_login11(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login12':
                caseid_v3.caseid_login12(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login13':
                caseid_v3.caseid_login13(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login14':
                caseid_v3.caseid_login14(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login15':
                caseid_v3.caseid_login15(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login16':
                caseid_v3.caseid_login16(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login17':
                caseid_v3.caseid_login17(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login18':
                caseid_v3.caseid_login18(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login19':
                caseid_v3.caseid_login19(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login20':
                caseid_v3.caseid_login20(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login21':
                caseid_v3.caseid_login21(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login22':
                caseid_v3.caseid_login22(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login23':
                caseid_v3.caseid_login23(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login24':
                caseid_v3.caseid_login24(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login25':
                caseid_v3.caseid_login25(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Login26':
                caseid_v3.caseid_login26(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat01':
                caseid_v3.caseid_giamsat01(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat02':
                caseid_v3.caseid_giamsat02(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat03':
                caseid_v3.caseid_giamsat03(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat04':
                caseid_v3.caseid_giamsat04(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat05':
                caseid_v3.caseid_giamsat05(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat06':
                caseid_v3.caseid_giamsat06(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat07':
                caseid_v3.caseid_giamsat07(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat08':
                caseid_v3.caseid_giamsat08(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat09':
                caseid_v3.caseid_giamsat09(self)
        except:
            module_other_v3.swich_tab_0()

        try:
            if case == 'GiamSat10':
                caseid_v3.caseid_giamsat10(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat11':
                caseid_v3.caseid_giamsat11(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat12':
                caseid_v3.caseid_giamsat12(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat13':
                caseid_v3.caseid_giamsat13(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat14':
                caseid_v3.caseid_giamsat14(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat15':
                caseid_v3.caseid_giamsat15(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat16':
                caseid_v3.caseid_giamsat16(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat17':
                caseid_v3.caseid_giamsat17(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat18':
                caseid_v3.caseid_giamsat18(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat19':
                caseid_v3.caseid_giamsat19(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat20':
                caseid_v3.caseid_giamsat20(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat21':
                caseid_v3.caseid_giamsat21(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat22':
                caseid_v3.caseid_giamsat22(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat23':
                caseid_v3.caseid_giamsat23(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat24':
                caseid_v3.caseid_giamsat24(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat25':
                caseid_v3.caseid_giamsat25(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat26':
                caseid_v3.caseid_giamsat26(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat27':
                caseid_v3.caseid_giamsat27(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat28':
                caseid_v3.caseid_giamsat28(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat29':
                caseid_v3.caseid_giamsat29(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat30':
                caseid_v3.caseid_giamsat30(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat31':
                caseid_v3.caseid_giamsat31(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat32':
                caseid_v3.caseid_giamsat32(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat33':
                caseid_v3.caseid_giamsat33(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat34':
                caseid_v3.caseid_giamsat34(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat35':
                caseid_v3.caseid_giamsat35(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat36':
                caseid_v3.caseid_giamsat36(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat37':
                caseid_v3.caseid_giamsat37(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat38':
                caseid_v3.caseid_giamsat38(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat39':
                caseid_v3.caseid_giamsat39(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat40':
                caseid_v3.caseid_giamsat40(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat41':
                caseid_v3.caseid_giamsat41(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat42':
                caseid_v3.caseid_giamsat42(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat43':
                caseid_v3.caseid_giamsat43(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat44':
                caseid_v3.caseid_giamsat44(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat45':
                caseid_v3.caseid_giamsat45(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat46':
                caseid_v3.caseid_giamsat46(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat47':
                caseid_v3.caseid_giamsat47(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat48':
                caseid_v3.caseid_giamsat48(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat49':
                caseid_v3.caseid_giamsat49(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat50':
                caseid_v3.caseid_giamsat50(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat51':
                caseid_v3.caseid_giamsat51(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat52':
                caseid_v3.caseid_giamsat52(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat53':
                caseid_v3.caseid_giamsat53(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat54':
                caseid_v3.caseid_giamsat54(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat55':
                caseid_v3.caseid_giamsat55(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat56':
                caseid_v3.caseid_giamsat56(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat57':
                caseid_v3.caseid_giamsat57(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat58':
                caseid_v3.caseid_giamsat58(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat59':
                caseid_v3.caseid_giamsat59(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat60':
                caseid_v3.caseid_giamsat60(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat61':
                caseid_v3.caseid_giamsat61(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat62':
                caseid_v3.caseid_giamsat62(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat63':
                caseid_v3.caseid_giamsat63(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat64':
                caseid_v3.caseid_giamsat64(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat65':
                caseid_v3.caseid_giamsat65(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat66':
                caseid_v3.caseid_giamsat66(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat67':
                caseid_v3.caseid_giamsat67(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat68':
                caseid_v3.caseid_giamsat68(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat69':
                caseid_v3.caseid_giamsat69(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat70':
                caseid_v3.caseid_giamsat70(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat71':
                caseid_v3.caseid_giamsat71(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat72':
                caseid_v3.caseid_giamsat72(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat73':
                caseid_v3.caseid_giamsat73(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat74':
                caseid_v3.caseid_giamsat74(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat75':
                caseid_v3.caseid_giamsat75(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat76':
                caseid_v3.caseid_giamsat76(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat77':
                caseid_v3.caseid_giamsat77(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat78':
                caseid_v3.caseid_giamsat78(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat79':
                caseid_v3.caseid_giamsat79(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat80':
                caseid_v3.caseid_giamsat80(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat81':
                caseid_v3.caseid_giamsat81(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat82':
                caseid_v3.caseid_giamsat82(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat83':
                caseid_v3.caseid_giamsat83(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat84':
                caseid_v3.caseid_giamsat84(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat85':
                caseid_v3.caseid_giamsat85(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat86':
                caseid_v3.caseid_giamsat86(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat87':
                caseid_v3.caseid_giamsat87(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat88':
                caseid_v3.caseid_giamsat88(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat89':
                caseid_v3.caseid_giamsat89(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat90':
                caseid_v3.caseid_giamsat90(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat91':
                caseid_v3.caseid_giamsat91(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat92':
                caseid_v3.caseid_giamsat92(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat93':
                caseid_v3.caseid_giamsat93(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat94':
                caseid_v3.caseid_giamsat94(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat95':
                caseid_v3.caseid_giamsat95(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat96':
                caseid_v3.caseid_giamsat96(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat97':
                caseid_v3.caseid_giamsat97(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat98':
                caseid_v3.caseid_giamsat98(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat99':
                caseid_v3.caseid_giamsat99(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat100':
                caseid_v3.caseid_giamsat100(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat101':
                caseid_v3.caseid_giamsat101(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat102':
                caseid_v3.caseid_giamsat102(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat103':
                caseid_v3.caseid_giamsat103(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat104':
                caseid_v3.caseid_giamsat104(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat105':
                caseid_v3.caseid_giamsat105(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat106':
                caseid_v3.caseid_giamsat106(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat107':
                caseid_v3.caseid_giamsat107(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat108':
                caseid_v3.caseid_giamsat108(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat109':
                caseid_v3.caseid_giamsat109(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat110':
                caseid_v3.caseid_giamsat110(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat111':
                caseid_v3.caseid_giamsat111(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat112':
                caseid_v3.caseid_giamsat112(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat113':
                caseid_v3.caseid_giamsat113(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat114':
                caseid_v3.caseid_giamsat114(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat115':
                caseid_v3.caseid_giamsat115(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat116':
                caseid_v3.caseid_giamsat116(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat117':
                caseid_v3.caseid_giamsat117(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat118':
                caseid_v3.caseid_giamsat118(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat119':
                caseid_v3.caseid_giamsat119(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat120':
                caseid_v3.caseid_giamsat120(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat121':
                caseid_v3.caseid_giamsat121(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat122':
                caseid_v3.caseid_giamsat122(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat123':
                caseid_v3.caseid_giamsat123(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat124':
                caseid_v3.caseid_giamsat124(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat125':
                caseid_v3.caseid_giamsat125(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat126':
                caseid_v3.caseid_giamsat126(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat127':
                caseid_v3.caseid_giamsat127(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat128':
                caseid_v3.caseid_giamsat128(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat129':
                caseid_v3.caseid_giamsat129(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat130':
                caseid_v3.caseid_giamsat130(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat131':
                caseid_v3.caseid_giamsat131(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat132':
                caseid_v3.caseid_giamsat132(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat133':
                caseid_v3.caseid_giamsat133(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat134':
                caseid_v3.caseid_giamsat134(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat135':
                caseid_v3.caseid_giamsat135(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat136':
                caseid_v3.caseid_giamsat136(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat137':
                caseid_v3.caseid_giamsat137(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat138':
                caseid_v3.caseid_giamsat138(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat139':
                caseid_v3.caseid_giamsat139(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat140':
                caseid_v3.caseid_giamsat140(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat141':
                caseid_v3.caseid_giamsat141(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat142':
                caseid_v3.caseid_giamsat142(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat143':
                caseid_v3.caseid_giamsat143(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat144':
                caseid_v3.caseid_giamsat144(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat145':
                caseid_v3.caseid_giamsat145(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat146':
                caseid_v3.caseid_giamsat146(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat147':
                caseid_v3.caseid_giamsat147(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat148':
                caseid_v3.caseid_giamsat148(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat149':
                caseid_v3.caseid_giamsat149(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat150':
                caseid_v3.caseid_giamsat150(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat151':
                caseid_v3.caseid_giamsat151(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat152':
                caseid_v3.caseid_giamsat152(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat153':
                caseid_v3.caseid_giamsat153(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat154':
                caseid_v3.caseid_giamsat154(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat155':
                caseid_v3.caseid_giamsat155(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat156':
                caseid_v3.caseid_giamsat156(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat157':
                caseid_v3.caseid_giamsat157(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat158':
                caseid_v3.caseid_giamsat158(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat159':
                caseid_v3.caseid_giamsat159(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat160':
                caseid_v3.caseid_giamsat160(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat161':
                caseid_v3.caseid_giamsat161(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat162':
                caseid_v3.caseid_giamsat162(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat163':
                caseid_v3.caseid_giamsat163(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat164':
                caseid_v3.caseid_giamsat164(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat165':
                caseid_v3.caseid_giamsat165(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat166':
                caseid_v3.caseid_giamsat166(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat167':
                caseid_v3.caseid_giamsat167(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat168':
                caseid_v3.caseid_giamsat168(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat169':
                caseid_v3.caseid_giamsat169(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat170':
                caseid_v3.caseid_giamsat170(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat171':
                caseid_v3.caseid_giamsat171(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat172':
                caseid_v3.caseid_giamsat172(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat173':
                caseid_v3.caseid_giamsat173(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat174':
                caseid_v3.caseid_giamsat174(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat175':
                caseid_v3.caseid_giamsat175(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat176':
                caseid_v3.caseid_giamsat176(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat177':
                caseid_v3.caseid_giamsat177(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat178':
                caseid_v3.caseid_giamsat178(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat179':
                caseid_v3.caseid_giamsat179(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat180':
                caseid_v3.caseid_giamsat180(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat181':
                caseid_v3.caseid_giamsat181(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat182':
                caseid_v3.caseid_giamsat182(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat183':
                caseid_v3.caseid_giamsat183(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat184':
                caseid_v3.caseid_giamsat184(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat185':
                caseid_v3.caseid_giamsat185(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat186':
                caseid_v3.caseid_giamsat186(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat187':
                caseid_v3.caseid_giamsat187(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat188':
                caseid_v3.caseid_giamsat188(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat189':
                caseid_v3.caseid_giamsat189(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat190':
                caseid_v3.caseid_giamsat190(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat191':
                caseid_v3.caseid_giamsat191(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat192':
                caseid_v3.caseid_giamsat192(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat193':
                caseid_v3.caseid_giamsat193(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat194':
                caseid_v3.caseid_giamsat194(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat195':
                caseid_v3.caseid_giamsat195(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat196':
                caseid_v3.caseid_giamsat196(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat197':
                caseid_v3.caseid_giamsat197(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat198':
                caseid_v3.caseid_giamsat198(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat199':
                caseid_v3.caseid_giamsat199(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat200':
                caseid_v3.caseid_giamsat200(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat201':
                caseid_v3.caseid_giamsat201(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat202':
                caseid_v3.caseid_giamsat202(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat203':
                caseid_v3.caseid_giamsat203(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat204':
                caseid_v3.caseid_giamsat204(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat205':
                caseid_v3.caseid_giamsat205(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat206':
                caseid_v3.caseid_giamsat206(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat207':
                caseid_v3.caseid_giamsat207(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat208':
                caseid_v3.caseid_giamsat208(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat209':
                caseid_v3.caseid_giamsat209(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat210':
                caseid_v3.caseid_giamsat210(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat211':
                caseid_v3.caseid_giamsat211(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat212':
                caseid_v3.caseid_giamsat212(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat213':
                caseid_v3.caseid_giamsat213(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat214':
                caseid_v3.caseid_giamsat214(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat215':
                caseid_v3.caseid_giamsat215(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat216':
                caseid_v3.caseid_giamsat216(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat217':
                caseid_v3.caseid_giamsat217(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat218':
                caseid_v3.caseid_giamsat218(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat219':
                caseid_v3.caseid_giamsat219(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat220':
                caseid_v3.caseid_giamsat220(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat221':
                caseid_v3.caseid_giamsat221(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat222':
                caseid_v3.caseid_giamsat222(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat223':
                caseid_v3.caseid_giamsat223(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat224':
                caseid_v3.caseid_giamsat224(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat225':
                caseid_v3.caseid_giamsat225(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat226':
                caseid_v3.caseid_giamsat226(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat227':
                caseid_v3.caseid_giamsat227(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat228':
                caseid_v3.caseid_giamsat228(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'GiamSat229':
                caseid_v3.caseid_giamsat229(self)
        except:
            module_other_v3.swich_tab_0()

        try:
            if case == 'Route01':
                caseid_v3.caseid_route01(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route02':
                caseid_v3.caseid_route02(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route03':
                caseid_v3.caseid_route03(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route04':
                caseid_v3.caseid_route04(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route05':
                caseid_v3.caseid_route05(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route06':
                caseid_v3.caseid_route06(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route07':
                caseid_v3.caseid_route07(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route08':
                caseid_v3.caseid_route08(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route09':
                caseid_v3.caseid_route09(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route10':
                caseid_v3.caseid_route10(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route11':
                caseid_v3.caseid_route11(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route12':
                caseid_v3.caseid_route12(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route13':
                caseid_v3.caseid_route13(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route14':
                caseid_v3.caseid_route14(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route15':
                caseid_v3.caseid_route15(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route16':
                caseid_v3.caseid_route16(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route17':
                caseid_v3.caseid_route17(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route18':
                caseid_v3.caseid_route18(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route19':
                caseid_v3.caseid_route19(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route20':
                caseid_v3.caseid_route20(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route21':
                caseid_v3.caseid_route21(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route22':
                caseid_v3.caseid_route22(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route23':
                caseid_v3.caseid_route23(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route24':
                caseid_v3.caseid_route24(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route25':
                caseid_v3.caseid_route25(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route26':
                caseid_v3.caseid_route26(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route27':
                caseid_v3.caseid_route27(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route28':
                caseid_v3.caseid_route28(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route29':
                caseid_v3.caseid_route29(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route30':
                caseid_v3.caseid_route30(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route31':
                caseid_v3.caseid_route31(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route32':
                caseid_v3.caseid_route32(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route33':
                caseid_v3.caseid_route33(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route34':
                caseid_v3.caseid_route34(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route35':
                caseid_v3.caseid_route35(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route36':
                caseid_v3.caseid_route36(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route37':
                caseid_v3.caseid_route37(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route38':
                caseid_v3.caseid_route38(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route39':
                caseid_v3.caseid_route39(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route40':
                caseid_v3.caseid_route40(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route41':
                caseid_v3.caseid_route41(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route42':
                caseid_v3.caseid_route42(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route43':
                caseid_v3.caseid_route43(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route44':
                caseid_v3.caseid_route44(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route45':
                caseid_v3.caseid_route45(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route46':
                caseid_v3.caseid_route46(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route47':
                caseid_v3.caseid_route47(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route48':
                caseid_v3.caseid_route48(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route49':
                caseid_v3.caseid_route49(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route50':
                caseid_v3.caseid_route50(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route51':
                caseid_v3.caseid_route51(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route52':
                caseid_v3.caseid_route52(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Route53':
                caseid_v3.caseid_route53(self)
        except:
            module_other_v3.swich_tab_0()

        try:
            if case == 'Admin01':
                caseid_v3.caseid_admin01(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin02':
                caseid_v3.caseid_admin02(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin03':
                caseid_v3.caseid_admin03(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin04':
                caseid_v3.caseid_admin04(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin05':
                caseid_v3.caseid_admin05(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin06':
                caseid_v3.caseid_admin06(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin07':
                caseid_v3.caseid_admin07(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin08':
                caseid_v3.caseid_admin08(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin09':
                caseid_v3.caseid_admin09(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin10':
                caseid_v3.caseid_admin10(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin11':
                caseid_v3.caseid_admin11(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin12':
                caseid_v3.caseid_admin12(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin13':
                caseid_v3.caseid_admin13(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin14':
                caseid_v3.caseid_admin14(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin15':
                caseid_v3.caseid_admin15(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin16':
                caseid_v3.caseid_admin16(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin17':
                caseid_v3.caseid_admin17(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin18':
                caseid_v3.caseid_admin18(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin19':
                caseid_v3.caseid_admin19(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin20':
                caseid_v3.caseid_admin20(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin21':
                caseid_v3.caseid_admin21(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin22':
                caseid_v3.caseid_admin22(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin23':
                caseid_v3.caseid_admin23(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin24':
                caseid_v3.caseid_admin24(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin25':
                caseid_v3.caseid_admin25(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin26':
                caseid_v3.caseid_admin26(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin27':
                caseid_v3.caseid_admin27(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin28':
                caseid_v3.caseid_admin28(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin29':
                caseid_v3.caseid_admin29(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin30':
                caseid_v3.caseid_admin30(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin31':
                caseid_v3.caseid_admin31(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin32':
                caseid_v3.caseid_admin32(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin33':
                caseid_v3.caseid_admin33(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin34':
                caseid_v3.caseid_admin34(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin35':
                caseid_v3.caseid_admin35(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin36':
                caseid_v3.caseid_admin36(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin37':
                caseid_v3.caseid_admin37(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin38':
                caseid_v3.caseid_admin38(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin39':
                caseid_v3.caseid_admin39(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin40':
                caseid_v3.caseid_admin40(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin41':
                caseid_v3.caseid_admin41(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin42':
                caseid_v3.caseid_admin42(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin43':
                caseid_v3.caseid_admin43(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin44':
                caseid_v3.caseid_admin44(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin45':
                caseid_v3.caseid_admin45(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin46':
                caseid_v3.caseid_admin46(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin47':
                caseid_v3.caseid_admin47(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin48':
                caseid_v3.caseid_admin48(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin49':
                caseid_v3.caseid_admin49(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin50':
                caseid_v3.caseid_admin50(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin51':
                caseid_v3.caseid_admin51(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin52':
                caseid_v3.caseid_admin52(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin53':
                caseid_v3.caseid_admin53(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin54':
                caseid_v3.caseid_admin54(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin55':
                caseid_v3.caseid_admin55(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin56':
                caseid_v3.caseid_admin56(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin57':
                caseid_v3.caseid_admin57(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin58':
                caseid_v3.caseid_admin58(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin59':
                caseid_v3.caseid_admin59(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin60':
                caseid_v3.caseid_admin60(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin61':
                caseid_v3.caseid_admin61(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin62':
                caseid_v3.caseid_admin62(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin63':
                caseid_v3.caseid_admin63(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin64':
                caseid_v3.caseid_admin64(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin65':
                caseid_v3.caseid_admin65(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin66':
                caseid_v3.caseid_admin66(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin67':
                caseid_v3.caseid_admin67(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin68':
                caseid_v3.caseid_admin68(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin69':
                caseid_v3.caseid_admin69(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin70':
                caseid_v3.caseid_admin70(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin71':
                caseid_v3.caseid_admin71(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin72':
                caseid_v3.caseid_admin72(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin73':
                caseid_v3.caseid_admin73(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin74':
                caseid_v3.caseid_admin74(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin75':
                caseid_v3.caseid_admin75(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin76':
                caseid_v3.caseid_admin76(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin77':
                caseid_v3.caseid_admin77(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin78':
                caseid_v3.caseid_admin78(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin79':
                caseid_v3.caseid_admin79(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin80':
                caseid_v3.caseid_admin80(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin81':
                caseid_v3.caseid_admin81(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin82':
                caseid_v3.caseid_admin82(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin83':
                caseid_v3.caseid_admin83(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin84':
                caseid_v3.caseid_admin84(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin85':
                caseid_v3.caseid_admin85(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin86':
                caseid_v3.caseid_admin86(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin87':
                caseid_v3.caseid_admin87(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin88':
                caseid_v3.caseid_admin88(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin89':
                caseid_v3.caseid_admin89(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin90':
                caseid_v3.caseid_admin90(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin91':
                caseid_v3.caseid_admin91(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin92':
                caseid_v3.caseid_admin92(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin93':
                caseid_v3.caseid_admin93(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin94':
                caseid_v3.caseid_admin94(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin95':
                caseid_v3.caseid_admin95(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin96':
                caseid_v3.caseid_admin96(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin97':
                caseid_v3.caseid_admin97(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin98':
                caseid_v3.caseid_admin98(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin99':
                caseid_v3.caseid_admin99(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin100':
                caseid_v3.caseid_admin100(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin101':
                caseid_v3.caseid_admin101(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin102':
                caseid_v3.caseid_admin102(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin103':
                caseid_v3.caseid_admin103(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin104':
                caseid_v3.caseid_admin104(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin105':
                caseid_v3.caseid_admin105(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin106':
                caseid_v3.caseid_admin106(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin107':
                caseid_v3.caseid_admin107(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin108':
                caseid_v3.caseid_admin108(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin109':
                caseid_v3.caseid_admin109(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin110':
                caseid_v3.caseid_admin110(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin111':
                caseid_v3.caseid_admin111(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin112':
                caseid_v3.caseid_admin112(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin113':
                caseid_v3.caseid_admin113(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin114':
                caseid_v3.caseid_admin114(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin115':
                caseid_v3.caseid_admin115(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin116':
                caseid_v3.caseid_admin116(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin117':
                caseid_v3.caseid_admin117(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin118':
                caseid_v3.caseid_admin118(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin119':
                caseid_v3.caseid_admin119(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin120':
                caseid_v3.caseid_admin120(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin121':
                caseid_v3.caseid_admin121(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin122':
                caseid_v3.caseid_admin122(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin123':
                caseid_v3.caseid_admin123(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin124':
                caseid_v3.caseid_admin124(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin125':
                caseid_v3.caseid_admin125(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin126':
                caseid_v3.caseid_admin126(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin127':
                caseid_v3.caseid_admin127(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin128':
                caseid_v3.caseid_admin128(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin129':
                caseid_v3.caseid_admin129(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin130':
                caseid_v3.caseid_admin130(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin131':
                caseid_v3.caseid_admin131(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin132':
                caseid_v3.caseid_admin132(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin133':
                caseid_v3.caseid_admin133(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin134':
                caseid_v3.caseid_admin134(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin135':
                caseid_v3.caseid_admin135(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin136':
                caseid_v3.caseid_admin136(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin137':
                caseid_v3.caseid_admin137(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin138':
                caseid_v3.caseid_admin138(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin139':
                caseid_v3.caseid_admin139(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin140':
                caseid_v3.caseid_admin140(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin141':
                caseid_v3.caseid_admin141(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin142':
                caseid_v3.caseid_admin142(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin143':
                caseid_v3.caseid_admin143(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin144':
                caseid_v3.caseid_admin144(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin145':
                caseid_v3.caseid_admin145(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin146':
                caseid_v3.caseid_admin146(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin147':
                caseid_v3.caseid_admin147(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin148':
                caseid_v3.caseid_admin148(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin149':
                caseid_v3.caseid_admin149(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin150':
                caseid_v3.caseid_admin150(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin151':
                caseid_v3.caseid_admin151(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin152':
                caseid_v3.caseid_admin152(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin153':
                caseid_v3.caseid_admin153(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin154':
                caseid_v3.caseid_admin154(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin155':
                caseid_v3.caseid_admin155(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin156':
                caseid_v3.caseid_admin156(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin157':
                caseid_v3.caseid_admin157(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin158':
                caseid_v3.caseid_admin158(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin159':
                caseid_v3.caseid_admin159(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin160':
                caseid_v3.caseid_admin160(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin161':
                caseid_v3.caseid_admin161(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin162':
                caseid_v3.caseid_admin162(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin163':
                caseid_v3.caseid_admin163(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin164':
                caseid_v3.caseid_admin164(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin165':
                caseid_v3.caseid_admin165(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin166':
                caseid_v3.caseid_admin166(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin167':
                caseid_v3.caseid_admin167(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin168':
                caseid_v3.caseid_admin168(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin169':
                caseid_v3.caseid_admin169(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin170':
                caseid_v3.caseid_admin170(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin171':
                caseid_v3.caseid_admin171(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin172':
                caseid_v3.caseid_admin172(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin173':
                caseid_v3.caseid_admin173(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin174':
                caseid_v3.caseid_admin174(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin175':
                caseid_v3.caseid_admin175(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin176':
                caseid_v3.caseid_admin176(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin177':
                caseid_v3.caseid_admin177(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin178':
                caseid_v3.caseid_admin178(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin179':
                caseid_v3.caseid_admin179(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin180':
                caseid_v3.caseid_admin180(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin181':
                caseid_v3.caseid_admin181(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin182':
                caseid_v3.caseid_admin182(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin183':
                caseid_v3.caseid_admin183(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin184':
                caseid_v3.caseid_admin184(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin185':
                caseid_v3.caseid_admin185(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin186':
                caseid_v3.caseid_admin186(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin187':
                caseid_v3.caseid_admin187(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin188':
                caseid_v3.caseid_admin188(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin189':
                caseid_v3.caseid_admin189(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin190':
                caseid_v3.caseid_admin190(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin191':
                caseid_v3.caseid_admin191(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin192':
                caseid_v3.caseid_admin192(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin193':
                caseid_v3.caseid_admin193(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin194':
                caseid_v3.caseid_admin194(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin195':
                caseid_v3.caseid_admin195(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin196':
                caseid_v3.caseid_admin196(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin197':
                caseid_v3.caseid_admin197(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin198':
                caseid_v3.caseid_admin198(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Admin199':
                caseid_v3.caseid_admin199(self)
        except:
            module_other_v3.swich_tab_0()

        try:
            if case == 'Report01':
                caseid_v3.caseid_report01(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report02':
                caseid_v3.caseid_report02(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report03':
                caseid_v3.caseid_report03(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report04':
                caseid_v3.caseid_report04(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report05':
                caseid_v3.caseid_report05(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report06':
                caseid_v3.caseid_report06(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report07':
                caseid_v3.caseid_report07(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report08':
                caseid_v3.caseid_report08(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report09':
                caseid_v3.caseid_report09(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report10':
                caseid_v3.caseid_report10(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report11':
                caseid_v3.caseid_report11(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report12':
                caseid_v3.caseid_report12(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report13':
                caseid_v3.caseid_report13(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report14':
                caseid_v3.caseid_report14(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report15':
                caseid_v3.caseid_report15(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report16':
                caseid_v3.caseid_report16(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report17':
                caseid_v3.caseid_report17(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report18':
                caseid_v3.caseid_report18(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report19':
                caseid_v3.caseid_report19(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report20':
                caseid_v3.caseid_report20(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report21':
                caseid_v3.caseid_report21(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report22':
                caseid_v3.caseid_report22(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report23':
                caseid_v3.caseid_report23(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report24':
                caseid_v3.caseid_report24(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report25':
                caseid_v3.caseid_report25(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report26':
                caseid_v3.caseid_report26(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report27':
                caseid_v3.caseid_report27(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report28':
                caseid_v3.caseid_report28(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report29':
                caseid_v3.caseid_report29(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report30':
                caseid_v3.caseid_report30(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report31':
                caseid_v3.caseid_report31(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report32':
                caseid_v3.caseid_report32(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report33':
                caseid_v3.caseid_report33(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report34':
                caseid_v3.caseid_report34(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report35':
                caseid_v3.caseid_report35(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report36':
                caseid_v3.caseid_report36(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report37':
                caseid_v3.caseid_report37(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report38':
                caseid_v3.caseid_report38(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report39':
                caseid_v3.caseid_report39(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report40':
                caseid_v3.caseid_report40(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report41':
                caseid_v3.caseid_report41(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report42':
                caseid_v3.caseid_report42(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report43':
                caseid_v3.caseid_report43(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report44':
                caseid_v3.caseid_report44(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report45':
                caseid_v3.caseid_report45(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report46':
                caseid_v3.caseid_report46(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report47':
                caseid_v3.caseid_report47(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report48':
                caseid_v3.caseid_report48(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report49':
                caseid_v3.caseid_report49(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report50':
                caseid_v3.caseid_report50(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report51':
                caseid_v3.caseid_report51(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report52':
                caseid_v3.caseid_report52(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report53':
                caseid_v3.caseid_report53(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report54':
                caseid_v3.caseid_report54(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report55':
                caseid_v3.caseid_report55(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report56':
                caseid_v3.caseid_report56(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report57':
                caseid_v3.caseid_report57(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report58':
                caseid_v3.caseid_report58(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report59':
                caseid_v3.caseid_report59(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report60':
                caseid_v3.caseid_report60(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report61':
                caseid_v3.caseid_report61(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report62':
                caseid_v3.caseid_report62(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report63':
                caseid_v3.caseid_report63(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report64':
                caseid_v3.caseid_report64(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report65':
                caseid_v3.caseid_report65(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report66':
                caseid_v3.caseid_report66(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report67':
                caseid_v3.caseid_report67(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report68':
                caseid_v3.caseid_report68(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report69':
                caseid_v3.caseid_report69(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report70':
                caseid_v3.caseid_report70(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report71':
                caseid_v3.caseid_report71(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report72':
                caseid_v3.caseid_report72(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report73':
                caseid_v3.caseid_report73(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report74':
                caseid_v3.caseid_report74(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report75':
                caseid_v3.caseid_report75(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report76':
                caseid_v3.caseid_report76(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report77':
                caseid_v3.caseid_report77(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report78':
                caseid_v3.caseid_report78(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report79':
                caseid_v3.caseid_report79(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report80':
                caseid_v3.caseid_report80(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report81':
                caseid_v3.caseid_report81(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report82':
                caseid_v3.caseid_report82(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report83':
                caseid_v3.caseid_report83(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report84':
                caseid_v3.caseid_report84(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report85':
                caseid_v3.caseid_report85(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report86':
                caseid_v3.caseid_report86(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report87':
                caseid_v3.caseid_report87(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report88':
                caseid_v3.caseid_report88(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report89':
                caseid_v3.caseid_report89(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report90':
                caseid_v3.caseid_report90(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report91':
                caseid_v3.caseid_report91(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report92':
                caseid_v3.caseid_report92(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report93':
                caseid_v3.caseid_report93(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report94':
                caseid_v3.caseid_report94(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Report95':
                caseid_v3.caseid_report95(self)
        except:
            module_other_v3.swich_tab_0()

        try:
            if case == 'Video01':
                caseid_v3.caseid_video01(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video02':
                caseid_v3.caseid_video02(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video03':
                caseid_v3.caseid_video03(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video04':
                caseid_v3.caseid_video04(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video05':
                caseid_v3.caseid_video05(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video06':
                caseid_v3.caseid_video06(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video07':
                caseid_v3.caseid_video07(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video08':
                caseid_v3.caseid_video08(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video09':
                caseid_v3.caseid_video09(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video10':
                caseid_v3.caseid_video10(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video11':
                caseid_v3.caseid_video11(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video12':
                caseid_v3.caseid_video12(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video13':
                caseid_v3.caseid_video13(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video14':
                caseid_v3.caseid_video14(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video15':
                caseid_v3.caseid_video15(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video16':
                caseid_v3.caseid_video16(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video17':
                caseid_v3.caseid_video17(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video18':
                caseid_v3.caseid_video18(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video19':
                caseid_v3.caseid_video19(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video20':
                caseid_v3.caseid_video20(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video21':
                caseid_v3.caseid_video21(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video22':
                caseid_v3.caseid_video22(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video23':
                caseid_v3.caseid_video23(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video24':
                caseid_v3.caseid_video24(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video25':
                caseid_v3.caseid_video25(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video26':
                caseid_v3.caseid_video26(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video27':
                caseid_v3.caseid_video27(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Video28':
                caseid_v3.caseid_video28(self)
        except:
            module_other_v3.swich_tab_0()

        try:
            if case == 'Image01':
                caseid_v3.caseid_image01(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image02':
                caseid_v3.caseid_image02(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image03':
                caseid_v3.caseid_image03(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image04':
                caseid_v3.caseid_image04(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image05':
                caseid_v3.caseid_image05(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image06':
                caseid_v3.caseid_image06(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image07':
                caseid_v3.caseid_image07(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image08':
                caseid_v3.caseid_image08(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image09':
                caseid_v3.caseid_image09(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image10':
                caseid_v3.caseid_image10(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image11':
                caseid_v3.caseid_image11(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image12':
                caseid_v3.caseid_image12(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image13':
                caseid_v3.caseid_image13(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image14':
                caseid_v3.caseid_image14(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image15':
                caseid_v3.caseid_image15(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image16':
                caseid_v3.caseid_image16(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image17':
                caseid_v3.caseid_image17(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image18':
                caseid_v3.caseid_image18(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image19':
                caseid_v3.caseid_image19(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image20':
                caseid_v3.caseid_image20(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image21':
                caseid_v3.caseid_image21(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image22':
                caseid_v3.caseid_image22(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image23':
                caseid_v3.caseid_image23(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image24':
                caseid_v3.caseid_image24(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Image25':
                caseid_v3.caseid_image25(self)
        except:
            module_other_v3.swich_tab_0()

        try:
            if case == 'Utility01':
                caseid_v3.caseid_utility01(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility02':
                caseid_v3.caseid_utility02(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility03':
                caseid_v3.caseid_utility03(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility04':
                caseid_v3.caseid_utility04(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility05':
                caseid_v3.caseid_utility05(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility06':
                caseid_v3.caseid_utility06(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility07':
                caseid_v3.caseid_utility07(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility08':
                caseid_v3.caseid_utility08(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility09':
                caseid_v3.caseid_utility09(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility10':
                caseid_v3.caseid_utility10(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility11':
                caseid_v3.caseid_utility11(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility12':
                caseid_v3.caseid_utility12(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility13':
                caseid_v3.caseid_utility13(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility14':
                caseid_v3.caseid_utility14(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility15':
                caseid_v3.caseid_utility15(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility16':
                caseid_v3.caseid_utility16(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility17':
                caseid_v3.caseid_utility17(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility18':
                caseid_v3.caseid_utility18(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility19':
                caseid_v3.caseid_utility19(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility20':
                caseid_v3.caseid_utility20(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility21':
                caseid_v3.caseid_utility21(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility22':
                caseid_v3.caseid_utility22(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility23':
                caseid_v3.caseid_utility23(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility24':
                caseid_v3.caseid_utility24(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Utility25':
                caseid_v3.caseid_utility25(self)
        except:
            module_other_v3.swich_tab_0()

        try:
            if case == 'Ai01':
                caseid_v3.caseid_ai01(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai02':
                caseid_v3.caseid_ai02(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai03':
                caseid_v3.caseid_ai03(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai04':
                caseid_v3.caseid_ai04(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai05':
                caseid_v3.caseid_ai05(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai06':
                caseid_v3.caseid_ai06(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai07':
                caseid_v3.caseid_ai07(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai08':
                caseid_v3.caseid_ai08(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai09':
                caseid_v3.caseid_ai09(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai10':
                caseid_v3.caseid_ai10(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai11':
                caseid_v3.caseid_ai11(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai12':
                caseid_v3.caseid_ai12(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai13':
                caseid_v3.caseid_ai13(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai14':
                caseid_v3.caseid_ai14(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai15':
                caseid_v3.caseid_ai15(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai16':
                caseid_v3.caseid_ai16(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai17':
                caseid_v3.caseid_ai17(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai18':
                caseid_v3.caseid_ai18(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai19':
                caseid_v3.caseid_ai19(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai20':
                caseid_v3.caseid_ai20(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai21':
                caseid_v3.caseid_ai21(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai22':
                caseid_v3.caseid_ai22(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai23':
                caseid_v3.caseid_ai23(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai24':
                caseid_v3.caseid_ai24(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai25':
                caseid_v3.caseid_ai25(self)
        except:
            module_other_v3.swich_tab_0()
        try:
            if case == 'Ai26':
                caseid_v3.caseid_ai26(self)
        except:
            module_other_v3.swich_tab_0()









def login(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_v3.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 10
    while (rownum < 40):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print("muc do 1", list_mucdo1)
    print("muc do 2", list_mucdo2)
    print("muc do 3", list_mucdo3)
    print("muc do 4", list_mucdo4)
    modetest = ''.join(re.findall(r'\d+', var_v3.modetest))
    for i in modetest:
        print("i", i)
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Login01':
                        caseid_v3.caseid_login01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login02':
                        caseid_v3.caseid_login02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login03':
                        caseid_v3.caseid_login03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login04':
                        caseid_v3.caseid_login04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login05':
                        caseid_v3.caseid_login05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login06':
                        caseid_v3.caseid_login06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login07':
                        caseid_v3.caseid_login07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login08':
                        caseid_v3.caseid_login08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login09':
                        caseid_v3.caseid_login09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login10':
                        caseid_v3.caseid_login10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login11':
                        caseid_v3.caseid_login11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login12':
                        caseid_v3.caseid_login12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login13':
                        caseid_v3.caseid_login13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login14':
                        caseid_v3.caseid_login14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login15':
                        caseid_v3.caseid_login15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login16':
                        caseid_v3.caseid_login16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login17':
                        caseid_v3.caseid_login17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login18':
                        caseid_v3.caseid_login18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login19':
                        caseid_v3.caseid_login19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login20':
                        caseid_v3.caseid_login20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login21':
                        caseid_v3.caseid_login21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login22':
                        caseid_v3.caseid_login22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login23':
                        caseid_v3.caseid_login23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login24':
                        caseid_v3.caseid_login24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login25':
                        caseid_v3.caseid_login25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login26':
                        caseid_v3.caseid_login26(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Login01':
                        caseid_v3.caseid_login01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login02':
                        caseid_v3.caseid_login02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login03':
                        caseid_v3.caseid_login03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login04':
                        caseid_v3.caseid_login04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login05':
                        caseid_v3.caseid_login05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login06':
                        caseid_v3.caseid_login06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login07':
                        caseid_v3.caseid_login07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login08':
                        caseid_v3.caseid_login08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login09':
                        caseid_v3.caseid_login09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login10':
                        caseid_v3.caseid_login10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login11':
                        caseid_v3.caseid_login11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login12':
                        caseid_v3.caseid_login12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login13':
                        caseid_v3.caseid_login13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login14':
                        caseid_v3.caseid_login14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login15':
                        caseid_v3.caseid_login15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login16':
                        caseid_v3.caseid_login16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login17':
                        caseid_v3.caseid_login17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login18':
                        caseid_v3.caseid_login18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login19':
                        caseid_v3.caseid_login19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login20':
                        caseid_v3.caseid_login20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login21':
                        caseid_v3.caseid_login21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login22':
                        caseid_v3.caseid_login22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login23':
                        caseid_v3.caseid_login23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login24':
                        caseid_v3.caseid_login24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login25':
                        caseid_v3.caseid_login25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login26':
                        caseid_v3.caseid_login26(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Login01':
                        caseid_v3.caseid_login01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login02':
                        caseid_v3.caseid_login02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login03':
                        caseid_v3.caseid_login03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login04':
                        caseid_v3.caseid_login04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login05':
                        caseid_v3.caseid_login05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login06':
                        caseid_v3.caseid_login06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login07':
                        caseid_v3.caseid_login07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login08':
                        caseid_v3.caseid_login08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login09':
                        caseid_v3.caseid_login09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login10':
                        caseid_v3.caseid_login10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login11':
                        caseid_v3.caseid_login11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login12':
                        caseid_v3.caseid_login12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login13':
                        caseid_v3.caseid_login13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login14':
                        caseid_v3.caseid_login14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login15':
                        caseid_v3.caseid_login15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login16':
                        caseid_v3.caseid_login16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login17':
                        caseid_v3.caseid_login17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login18':
                        caseid_v3.caseid_login18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login19':
                        caseid_v3.caseid_login19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login20':
                        caseid_v3.caseid_login20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login21':
                        caseid_v3.caseid_login21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login22':
                        caseid_v3.caseid_login22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login23':
                        caseid_v3.caseid_login23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login24':
                        caseid_v3.caseid_login24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login25':
                        caseid_v3.caseid_login25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login26':
                        caseid_v3.caseid_login26(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Login01':
                        caseid_v3.caseid_login01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login02':
                        caseid_v3.caseid_login02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login03':
                        caseid_v3.caseid_login03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login04':
                        caseid_v3.caseid_login04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login05':
                        caseid_v3.caseid_login05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login06':
                        caseid_v3.caseid_login06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login07':
                        caseid_v3.caseid_login07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login08':
                        caseid_v3.caseid_login08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login09':
                        caseid_v3.caseid_login09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login10':
                        caseid_v3.caseid_login10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login11':
                        caseid_v3.caseid_login11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login12':
                        caseid_v3.caseid_login12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login13':
                        caseid_v3.caseid_login13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login14':
                        caseid_v3.caseid_login14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login15':
                        caseid_v3.caseid_login15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login16':
                        caseid_v3.caseid_login16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login17':
                        caseid_v3.caseid_login17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login18':
                        caseid_v3.caseid_login18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login19':
                        caseid_v3.caseid_login19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login20':
                        caseid_v3.caseid_login20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login21':
                        caseid_v3.caseid_login21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login22':
                        caseid_v3.caseid_login22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login23':
                        caseid_v3.caseid_login23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login24':
                        caseid_v3.caseid_login24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login25':
                        caseid_v3.caseid_login25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Login26':
                        caseid_v3.caseid_login26(self)
                except:
                    module_other_v3.swich_tab_0()


#2 giám sát
def monitor(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_v3.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 41
    while (rownum < 343):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)

    modetest = ''.join(re.findall(r'\d+', var_v3.modetest))
    for i in modetest:
        if i == "1":
            print(list_mucdo1)
            for case in list_mucdo1:
                try:
                    if case == 'GiamSat01':
                        caseid_v3.caseid_giamsat01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat02':
                        caseid_v3.caseid_giamsat02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat03':
                        caseid_v3.caseid_giamsat03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat04':
                        caseid_v3.caseid_giamsat04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat05':
                        caseid_v3.caseid_giamsat05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat06':
                        caseid_v3.caseid_giamsat06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat07':
                        caseid_v3.caseid_giamsat07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat08':
                        caseid_v3.caseid_giamsat08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat09':
                        caseid_v3.caseid_giamsat09(self)
                except:
                    module_other_v3.swich_tab_0()

                try:
                    if case == 'GiamSat10':
                        caseid_v3.caseid_giamsat10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat11':
                        caseid_v3.caseid_giamsat11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat12':
                        caseid_v3.caseid_giamsat12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat13':
                        caseid_v3.caseid_giamsat13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat14':
                        caseid_v3.caseid_giamsat14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat15':
                        caseid_v3.caseid_giamsat15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat16':
                        caseid_v3.caseid_giamsat16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat17':
                        caseid_v3.caseid_giamsat17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat18':
                        caseid_v3.caseid_giamsat18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat19':
                        caseid_v3.caseid_giamsat19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat20':
                        caseid_v3.caseid_giamsat20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat21':
                        caseid_v3.caseid_giamsat21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat22':
                        caseid_v3.caseid_giamsat22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat23':
                        caseid_v3.caseid_giamsat23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat24':
                        caseid_v3.caseid_giamsat24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat25':
                        caseid_v3.caseid_giamsat25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat26':
                        caseid_v3.caseid_giamsat26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat27':
                        caseid_v3.caseid_giamsat27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat28':
                        caseid_v3.caseid_giamsat28(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat29':
                        caseid_v3.caseid_giamsat29(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat30':
                        caseid_v3.caseid_giamsat30(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat31':
                        caseid_v3.caseid_giamsat31(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat32':
                        caseid_v3.caseid_giamsat32(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat33':
                        caseid_v3.caseid_giamsat33(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat34':
                        caseid_v3.caseid_giamsat34(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat35':
                        caseid_v3.caseid_giamsat35(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat36':
                        caseid_v3.caseid_giamsat36(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat37':
                        caseid_v3.caseid_giamsat37(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat38':
                        caseid_v3.caseid_giamsat38(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat39':
                        caseid_v3.caseid_giamsat39(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat40':
                        caseid_v3.caseid_giamsat40(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat41':
                        caseid_v3.caseid_giamsat41(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat42':
                        caseid_v3.caseid_giamsat42(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat43':
                        caseid_v3.caseid_giamsat43(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat44':
                        caseid_v3.caseid_giamsat44(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat45':
                        caseid_v3.caseid_giamsat45(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat46':
                        caseid_v3.caseid_giamsat46(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat47':
                        caseid_v3.caseid_giamsat47(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat48':
                        caseid_v3.caseid_giamsat48(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat49':
                        caseid_v3.caseid_giamsat49(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat50':
                        caseid_v3.caseid_giamsat50(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat51':
                        caseid_v3.caseid_giamsat51(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat52':
                        caseid_v3.caseid_giamsat52(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat53':
                        caseid_v3.caseid_giamsat53(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat54':
                        caseid_v3.caseid_giamsat54(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat55':
                        caseid_v3.caseid_giamsat55(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat56':
                        caseid_v3.caseid_giamsat56(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat57':
                        caseid_v3.caseid_giamsat57(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat58':
                        caseid_v3.caseid_giamsat58(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat59':
                        caseid_v3.caseid_giamsat59(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat60':
                        caseid_v3.caseid_giamsat60(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat61':
                        caseid_v3.caseid_giamsat61(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat62':
                        caseid_v3.caseid_giamsat62(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat63':
                        caseid_v3.caseid_giamsat63(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat64':
                        caseid_v3.caseid_giamsat64(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat65':
                        caseid_v3.caseid_giamsat65(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat66':
                        caseid_v3.caseid_giamsat66(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat67':
                        caseid_v3.caseid_giamsat67(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat68':
                        caseid_v3.caseid_giamsat68(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat69':
                        caseid_v3.caseid_giamsat69(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat70':
                        caseid_v3.caseid_giamsat70(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat71':
                        caseid_v3.caseid_giamsat71(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat72':
                        caseid_v3.caseid_giamsat72(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat73':
                        caseid_v3.caseid_giamsat73(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat74':
                        caseid_v3.caseid_giamsat74(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat75':
                        caseid_v3.caseid_giamsat75(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat76':
                        caseid_v3.caseid_giamsat76(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat77':
                        caseid_v3.caseid_giamsat77(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat78':
                        caseid_v3.caseid_giamsat78(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat79':
                        caseid_v3.caseid_giamsat79(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat80':
                        caseid_v3.caseid_giamsat80(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat81':
                        caseid_v3.caseid_giamsat81(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat82':
                        caseid_v3.caseid_giamsat82(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat83':
                        caseid_v3.caseid_giamsat83(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat84':
                        caseid_v3.caseid_giamsat84(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat85':
                        caseid_v3.caseid_giamsat85(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat86':
                        caseid_v3.caseid_giamsat86(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat87':
                        caseid_v3.caseid_giamsat87(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat88':
                        caseid_v3.caseid_giamsat88(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat89':
                        caseid_v3.caseid_giamsat89(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat90':
                        caseid_v3.caseid_giamsat90(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat91':
                        caseid_v3.caseid_giamsat91(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat92':
                        caseid_v3.caseid_giamsat92(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat93':
                        caseid_v3.caseid_giamsat93(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat94':
                        caseid_v3.caseid_giamsat94(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat95':
                        caseid_v3.caseid_giamsat95(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat96':
                        caseid_v3.caseid_giamsat96(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat97':
                        caseid_v3.caseid_giamsat97(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat98':
                        caseid_v3.caseid_giamsat98(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat99':
                        caseid_v3.caseid_giamsat99(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat100':
                        caseid_v3.caseid_giamsat100(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat101':
                        caseid_v3.caseid_giamsat101(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat102':
                        caseid_v3.caseid_giamsat102(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat103':
                        caseid_v3.caseid_giamsat103(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat104':
                        caseid_v3.caseid_giamsat104(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat105':
                        caseid_v3.caseid_giamsat105(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat106':
                        caseid_v3.caseid_giamsat106(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat107':
                        caseid_v3.caseid_giamsat107(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat108':
                        caseid_v3.caseid_giamsat108(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat109':
                        caseid_v3.caseid_giamsat109(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat110':
                        caseid_v3.caseid_giamsat110(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat111':
                        caseid_v3.caseid_giamsat111(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat112':
                        caseid_v3.caseid_giamsat112(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat113':
                        caseid_v3.caseid_giamsat113(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat114':
                        caseid_v3.caseid_giamsat114(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat115':
                        caseid_v3.caseid_giamsat115(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat116':
                        caseid_v3.caseid_giamsat116(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat117':
                        caseid_v3.caseid_giamsat117(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat118':
                        caseid_v3.caseid_giamsat118(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat119':
                        caseid_v3.caseid_giamsat119(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat120':
                        caseid_v3.caseid_giamsat120(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat121':
                        caseid_v3.caseid_giamsat121(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat122':
                        caseid_v3.caseid_giamsat122(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat123':
                        caseid_v3.caseid_giamsat123(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat124':
                        caseid_v3.caseid_giamsat124(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat125':
                        caseid_v3.caseid_giamsat125(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat126':
                        caseid_v3.caseid_giamsat126(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat127':
                        caseid_v3.caseid_giamsat127(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat128':
                        caseid_v3.caseid_giamsat128(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat129':
                        caseid_v3.caseid_giamsat129(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat130':
                        caseid_v3.caseid_giamsat130(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat131':
                        caseid_v3.caseid_giamsat131(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat132':
                        caseid_v3.caseid_giamsat132(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat133':
                        caseid_v3.caseid_giamsat133(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat134':
                        caseid_v3.caseid_giamsat134(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat135':
                        caseid_v3.caseid_giamsat135(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat136':
                        caseid_v3.caseid_giamsat136(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat137':
                        caseid_v3.caseid_giamsat137(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat138':
                        caseid_v3.caseid_giamsat138(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat139':
                        caseid_v3.caseid_giamsat139(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat140':
                        caseid_v3.caseid_giamsat140(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat141':
                        caseid_v3.caseid_giamsat141(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat142':
                        caseid_v3.caseid_giamsat142(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat143':
                        caseid_v3.caseid_giamsat143(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat144':
                        caseid_v3.caseid_giamsat144(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat145':
                        caseid_v3.caseid_giamsat145(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat146':
                        caseid_v3.caseid_giamsat146(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat147':
                        caseid_v3.caseid_giamsat147(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat148':
                        caseid_v3.caseid_giamsat148(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat149':
                        caseid_v3.caseid_giamsat149(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat150':
                        caseid_v3.caseid_giamsat150(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat151':
                        caseid_v3.caseid_giamsat151(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat152':
                        caseid_v3.caseid_giamsat152(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat153':
                        caseid_v3.caseid_giamsat153(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat154':
                        caseid_v3.caseid_giamsat154(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat155':
                        caseid_v3.caseid_giamsat155(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat156':
                        caseid_v3.caseid_giamsat156(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat157':
                        caseid_v3.caseid_giamsat157(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat158':
                        caseid_v3.caseid_giamsat158(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat159':
                        caseid_v3.caseid_giamsat159(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat160':
                        caseid_v3.caseid_giamsat160(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat161':
                        caseid_v3.caseid_giamsat161(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat162':
                        caseid_v3.caseid_giamsat162(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat163':
                        caseid_v3.caseid_giamsat163(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat164':
                        caseid_v3.caseid_giamsat164(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat165':
                        caseid_v3.caseid_giamsat165(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat166':
                        caseid_v3.caseid_giamsat166(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat167':
                        caseid_v3.caseid_giamsat167(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat168':
                        caseid_v3.caseid_giamsat168(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat169':
                        caseid_v3.caseid_giamsat169(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat170':
                        caseid_v3.caseid_giamsat170(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat171':
                        caseid_v3.caseid_giamsat171(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat172':
                        caseid_v3.caseid_giamsat172(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat173':
                        caseid_v3.caseid_giamsat173(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat174':
                        caseid_v3.caseid_giamsat174(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat175':
                        caseid_v3.caseid_giamsat175(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat176':
                        caseid_v3.caseid_giamsat176(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat177':
                        caseid_v3.caseid_giamsat177(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat178':
                        caseid_v3.caseid_giamsat178(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat179':
                        caseid_v3.caseid_giamsat179(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat180':
                        caseid_v3.caseid_giamsat180(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat181':
                        caseid_v3.caseid_giamsat181(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat182':
                        caseid_v3.caseid_giamsat182(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat183':
                        caseid_v3.caseid_giamsat183(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat184':
                        caseid_v3.caseid_giamsat184(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat185':
                        caseid_v3.caseid_giamsat185(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat186':
                        caseid_v3.caseid_giamsat186(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat187':
                        caseid_v3.caseid_giamsat187(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat188':
                        caseid_v3.caseid_giamsat188(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat189':
                        caseid_v3.caseid_giamsat189(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat190':
                        caseid_v3.caseid_giamsat190(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat191':
                        caseid_v3.caseid_giamsat191(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat192':
                        caseid_v3.caseid_giamsat192(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat193':
                        caseid_v3.caseid_giamsat193(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat194':
                        caseid_v3.caseid_giamsat194(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat195':
                        caseid_v3.caseid_giamsat195(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat196':
                        caseid_v3.caseid_giamsat196(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat197':
                        caseid_v3.caseid_giamsat197(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat198':
                        caseid_v3.caseid_giamsat198(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat199':
                        caseid_v3.caseid_giamsat199(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat200':
                        caseid_v3.caseid_giamsat200(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat201':
                        caseid_v3.caseid_giamsat201(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat202':
                        caseid_v3.caseid_giamsat202(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat203':
                        caseid_v3.caseid_giamsat203(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat204':
                        caseid_v3.caseid_giamsat204(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat205':
                        caseid_v3.caseid_giamsat205(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat206':
                        caseid_v3.caseid_giamsat206(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat207':
                        caseid_v3.caseid_giamsat207(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat208':
                        caseid_v3.caseid_giamsat208(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat209':
                        caseid_v3.caseid_giamsat209(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat210':
                        caseid_v3.caseid_giamsat210(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat211':
                        caseid_v3.caseid_giamsat211(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat212':
                        caseid_v3.caseid_giamsat212(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat213':
                        caseid_v3.caseid_giamsat213(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat214':
                        caseid_v3.caseid_giamsat214(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat215':
                        caseid_v3.caseid_giamsat215(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat216':
                        caseid_v3.caseid_giamsat216(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat217':
                        caseid_v3.caseid_giamsat217(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat218':
                        caseid_v3.caseid_giamsat218(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat219':
                        caseid_v3.caseid_giamsat219(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat220':
                        caseid_v3.caseid_giamsat220(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat221':
                        caseid_v3.caseid_giamsat221(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat222':
                        caseid_v3.caseid_giamsat222(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat223':
                        caseid_v3.caseid_giamsat223(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat224':
                        caseid_v3.caseid_giamsat224(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat225':
                        caseid_v3.caseid_giamsat225(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat226':
                        caseid_v3.caseid_giamsat226(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat227':
                        caseid_v3.caseid_giamsat227(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat228':
                        caseid_v3.caseid_giamsat228(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat229':
                        caseid_v3.caseid_giamsat229(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "2":
            print(list_mucdo2)
            for case in list_mucdo2:
                try:
                    if case == 'GiamSat01':
                        caseid_v3.caseid_giamsat01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat02':
                        caseid_v3.caseid_giamsat02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat03':
                        caseid_v3.caseid_giamsat03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat04':
                        caseid_v3.caseid_giamsat04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat05':
                        caseid_v3.caseid_giamsat05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat06':
                        caseid_v3.caseid_giamsat06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat07':
                        caseid_v3.caseid_giamsat07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat08':
                        caseid_v3.caseid_giamsat08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat09':
                        caseid_v3.caseid_giamsat09(self)
                except:
                    module_other_v3.swich_tab_0()

                try:
                    if case == 'GiamSat10':
                        caseid_v3.caseid_giamsat10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat11':
                        caseid_v3.caseid_giamsat11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat12':
                        caseid_v3.caseid_giamsat12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat13':
                        caseid_v3.caseid_giamsat13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat14':
                        caseid_v3.caseid_giamsat14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat15':
                        caseid_v3.caseid_giamsat15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat16':
                        caseid_v3.caseid_giamsat16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat17':
                        caseid_v3.caseid_giamsat17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat18':
                        caseid_v3.caseid_giamsat18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat19':
                        caseid_v3.caseid_giamsat19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat20':
                        caseid_v3.caseid_giamsat20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat21':
                        caseid_v3.caseid_giamsat21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat22':
                        caseid_v3.caseid_giamsat22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat23':
                        caseid_v3.caseid_giamsat23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat24':
                        caseid_v3.caseid_giamsat24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat25':
                        caseid_v3.caseid_giamsat25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat26':
                        caseid_v3.caseid_giamsat26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat27':
                        caseid_v3.caseid_giamsat27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat28':
                        caseid_v3.caseid_giamsat28(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat29':
                        caseid_v3.caseid_giamsat29(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat30':
                        caseid_v3.caseid_giamsat30(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat31':
                        caseid_v3.caseid_giamsat31(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat32':
                        caseid_v3.caseid_giamsat32(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat33':
                        caseid_v3.caseid_giamsat33(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat34':
                        caseid_v3.caseid_giamsat34(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat35':
                        caseid_v3.caseid_giamsat35(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat36':
                        caseid_v3.caseid_giamsat36(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat37':
                        caseid_v3.caseid_giamsat37(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat38':
                        caseid_v3.caseid_giamsat38(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat39':
                        caseid_v3.caseid_giamsat39(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat40':
                        caseid_v3.caseid_giamsat40(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat41':
                        caseid_v3.caseid_giamsat41(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat42':
                        caseid_v3.caseid_giamsat42(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat43':
                        caseid_v3.caseid_giamsat43(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat44':
                        caseid_v3.caseid_giamsat44(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat45':
                        caseid_v3.caseid_giamsat45(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat46':
                        caseid_v3.caseid_giamsat46(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat47':
                        caseid_v3.caseid_giamsat47(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat48':
                        caseid_v3.caseid_giamsat48(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat49':
                        caseid_v3.caseid_giamsat49(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat50':
                        caseid_v3.caseid_giamsat50(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat51':
                        caseid_v3.caseid_giamsat51(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat52':
                        caseid_v3.caseid_giamsat52(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat53':
                        caseid_v3.caseid_giamsat53(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat54':
                        caseid_v3.caseid_giamsat54(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat55':
                        caseid_v3.caseid_giamsat55(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat56':
                        caseid_v3.caseid_giamsat56(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat57':
                        caseid_v3.caseid_giamsat57(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat58':
                        caseid_v3.caseid_giamsat58(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat59':
                        caseid_v3.caseid_giamsat59(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat60':
                        caseid_v3.caseid_giamsat60(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat61':
                        caseid_v3.caseid_giamsat61(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat62':
                        caseid_v3.caseid_giamsat62(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat63':
                        caseid_v3.caseid_giamsat63(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat64':
                        caseid_v3.caseid_giamsat64(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat65':
                        caseid_v3.caseid_giamsat65(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat66':
                        caseid_v3.caseid_giamsat66(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat67':
                        caseid_v3.caseid_giamsat67(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat68':
                        caseid_v3.caseid_giamsat68(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat69':
                        caseid_v3.caseid_giamsat69(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat70':
                        caseid_v3.caseid_giamsat70(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat71':
                        caseid_v3.caseid_giamsat71(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat72':
                        caseid_v3.caseid_giamsat72(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat73':
                        caseid_v3.caseid_giamsat73(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat74':
                        caseid_v3.caseid_giamsat74(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat75':
                        caseid_v3.caseid_giamsat75(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat76':
                        caseid_v3.caseid_giamsat76(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat77':
                        caseid_v3.caseid_giamsat77(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat78':
                        caseid_v3.caseid_giamsat78(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat79':
                        caseid_v3.caseid_giamsat79(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat80':
                        caseid_v3.caseid_giamsat80(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat81':
                        caseid_v3.caseid_giamsat81(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat82':
                        caseid_v3.caseid_giamsat82(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat83':
                        caseid_v3.caseid_giamsat83(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat84':
                        caseid_v3.caseid_giamsat84(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat85':
                        caseid_v3.caseid_giamsat85(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat86':
                        caseid_v3.caseid_giamsat86(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat87':
                        caseid_v3.caseid_giamsat87(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat88':
                        caseid_v3.caseid_giamsat88(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat89':
                        caseid_v3.caseid_giamsat89(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat90':
                        caseid_v3.caseid_giamsat90(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat91':
                        caseid_v3.caseid_giamsat91(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat92':
                        caseid_v3.caseid_giamsat92(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat93':
                        caseid_v3.caseid_giamsat93(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat94':
                        caseid_v3.caseid_giamsat94(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat95':
                        caseid_v3.caseid_giamsat95(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat96':
                        caseid_v3.caseid_giamsat96(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat97':
                        caseid_v3.caseid_giamsat97(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat98':
                        caseid_v3.caseid_giamsat98(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat99':
                        caseid_v3.caseid_giamsat99(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat100':
                        caseid_v3.caseid_giamsat100(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat101':
                        caseid_v3.caseid_giamsat101(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat102':
                        caseid_v3.caseid_giamsat102(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat103':
                        caseid_v3.caseid_giamsat103(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat104':
                        caseid_v3.caseid_giamsat104(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat105':
                        caseid_v3.caseid_giamsat105(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat106':
                        caseid_v3.caseid_giamsat106(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat107':
                        caseid_v3.caseid_giamsat107(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat108':
                        caseid_v3.caseid_giamsat108(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat109':
                        caseid_v3.caseid_giamsat109(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat110':
                        caseid_v3.caseid_giamsat110(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat111':
                        caseid_v3.caseid_giamsat111(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat112':
                        caseid_v3.caseid_giamsat112(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat113':
                        caseid_v3.caseid_giamsat113(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat114':
                        caseid_v3.caseid_giamsat114(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat115':
                        caseid_v3.caseid_giamsat115(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat116':
                        caseid_v3.caseid_giamsat116(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat117':
                        caseid_v3.caseid_giamsat117(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat118':
                        caseid_v3.caseid_giamsat118(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat119':
                        caseid_v3.caseid_giamsat119(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat120':
                        caseid_v3.caseid_giamsat120(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat121':
                        caseid_v3.caseid_giamsat121(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat122':
                        caseid_v3.caseid_giamsat122(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat123':
                        caseid_v3.caseid_giamsat123(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat124':
                        caseid_v3.caseid_giamsat124(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat125':
                        caseid_v3.caseid_giamsat125(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat126':
                        caseid_v3.caseid_giamsat126(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat127':
                        caseid_v3.caseid_giamsat127(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat128':
                        caseid_v3.caseid_giamsat128(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat129':
                        caseid_v3.caseid_giamsat129(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat130':
                        caseid_v3.caseid_giamsat130(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat131':
                        caseid_v3.caseid_giamsat131(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat132':
                        caseid_v3.caseid_giamsat132(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat133':
                        caseid_v3.caseid_giamsat133(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat134':
                        caseid_v3.caseid_giamsat134(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat135':
                        caseid_v3.caseid_giamsat135(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat136':
                        caseid_v3.caseid_giamsat136(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat137':
                        caseid_v3.caseid_giamsat137(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat138':
                        caseid_v3.caseid_giamsat138(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat139':
                        caseid_v3.caseid_giamsat139(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat140':
                        caseid_v3.caseid_giamsat140(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat141':
                        caseid_v3.caseid_giamsat141(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat142':
                        caseid_v3.caseid_giamsat142(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat143':
                        caseid_v3.caseid_giamsat143(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat144':
                        caseid_v3.caseid_giamsat144(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat145':
                        caseid_v3.caseid_giamsat145(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat146':
                        caseid_v3.caseid_giamsat146(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat147':
                        caseid_v3.caseid_giamsat147(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat148':
                        caseid_v3.caseid_giamsat148(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat149':
                        caseid_v3.caseid_giamsat149(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat150':
                        caseid_v3.caseid_giamsat150(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat151':
                        caseid_v3.caseid_giamsat151(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat152':
                        caseid_v3.caseid_giamsat152(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat153':
                        caseid_v3.caseid_giamsat153(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat154':
                        caseid_v3.caseid_giamsat154(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat155':
                        caseid_v3.caseid_giamsat155(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat156':
                        caseid_v3.caseid_giamsat156(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat157':
                        caseid_v3.caseid_giamsat157(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat158':
                        caseid_v3.caseid_giamsat158(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat159':
                        caseid_v3.caseid_giamsat159(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat160':
                        caseid_v3.caseid_giamsat160(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat161':
                        caseid_v3.caseid_giamsat161(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat162':
                        caseid_v3.caseid_giamsat162(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat163':
                        caseid_v3.caseid_giamsat163(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat164':
                        caseid_v3.caseid_giamsat164(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat165':
                        caseid_v3.caseid_giamsat165(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat166':
                        caseid_v3.caseid_giamsat166(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat167':
                        caseid_v3.caseid_giamsat167(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat168':
                        caseid_v3.caseid_giamsat168(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat169':
                        caseid_v3.caseid_giamsat169(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat170':
                        caseid_v3.caseid_giamsat170(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat171':
                        caseid_v3.caseid_giamsat171(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat172':
                        caseid_v3.caseid_giamsat172(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat173':
                        caseid_v3.caseid_giamsat173(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat174':
                        caseid_v3.caseid_giamsat174(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat175':
                        caseid_v3.caseid_giamsat175(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat176':
                        caseid_v3.caseid_giamsat176(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat177':
                        caseid_v3.caseid_giamsat177(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat178':
                        caseid_v3.caseid_giamsat178(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat179':
                        caseid_v3.caseid_giamsat179(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat180':
                        caseid_v3.caseid_giamsat180(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat181':
                        caseid_v3.caseid_giamsat181(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat182':
                        caseid_v3.caseid_giamsat182(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat183':
                        caseid_v3.caseid_giamsat183(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat184':
                        caseid_v3.caseid_giamsat184(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat185':
                        caseid_v3.caseid_giamsat185(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat186':
                        caseid_v3.caseid_giamsat186(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat187':
                        caseid_v3.caseid_giamsat187(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat188':
                        caseid_v3.caseid_giamsat188(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat189':
                        caseid_v3.caseid_giamsat189(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat190':
                        caseid_v3.caseid_giamsat190(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat191':
                        caseid_v3.caseid_giamsat191(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat192':
                        caseid_v3.caseid_giamsat192(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat193':
                        caseid_v3.caseid_giamsat193(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat194':
                        caseid_v3.caseid_giamsat194(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat195':
                        caseid_v3.caseid_giamsat195(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat196':
                        caseid_v3.caseid_giamsat196(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat197':
                        caseid_v3.caseid_giamsat197(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat198':
                        caseid_v3.caseid_giamsat198(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat199':
                        caseid_v3.caseid_giamsat199(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat200':
                        caseid_v3.caseid_giamsat200(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat201':
                        caseid_v3.caseid_giamsat201(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat202':
                        caseid_v3.caseid_giamsat202(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat203':
                        caseid_v3.caseid_giamsat203(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat204':
                        caseid_v3.caseid_giamsat204(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat205':
                        caseid_v3.caseid_giamsat205(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat206':
                        caseid_v3.caseid_giamsat206(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat207':
                        caseid_v3.caseid_giamsat207(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat208':
                        caseid_v3.caseid_giamsat208(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat209':
                        caseid_v3.caseid_giamsat209(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat210':
                        caseid_v3.caseid_giamsat210(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat211':
                        caseid_v3.caseid_giamsat211(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat212':
                        caseid_v3.caseid_giamsat212(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat213':
                        caseid_v3.caseid_giamsat213(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat214':
                        caseid_v3.caseid_giamsat214(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat215':
                        caseid_v3.caseid_giamsat215(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat216':
                        caseid_v3.caseid_giamsat216(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat217':
                        caseid_v3.caseid_giamsat217(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat218':
                        caseid_v3.caseid_giamsat218(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat219':
                        caseid_v3.caseid_giamsat219(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat220':
                        caseid_v3.caseid_giamsat220(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat221':
                        caseid_v3.caseid_giamsat221(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat222':
                        caseid_v3.caseid_giamsat222(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat223':
                        caseid_v3.caseid_giamsat223(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat224':
                        caseid_v3.caseid_giamsat224(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat225':
                        caseid_v3.caseid_giamsat225(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat226':
                        caseid_v3.caseid_giamsat226(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat227':
                        caseid_v3.caseid_giamsat227(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat228':
                        caseid_v3.caseid_giamsat228(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat229':
                        caseid_v3.caseid_giamsat229(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "3":
            print(list_mucdo3)
            for case in list_mucdo3:
                try:
                    if case == 'GiamSat01':
                        caseid_v3.caseid_giamsat01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat02':
                        caseid_v3.caseid_giamsat02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat03':
                        caseid_v3.caseid_giamsat03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat04':
                        caseid_v3.caseid_giamsat04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat05':
                        caseid_v3.caseid_giamsat05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat06':
                        caseid_v3.caseid_giamsat06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat07':
                        caseid_v3.caseid_giamsat07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat08':
                        caseid_v3.caseid_giamsat08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat09':
                        caseid_v3.caseid_giamsat09(self)
                except:
                    module_other_v3.swich_tab_0()

                try:
                    if case == 'GiamSat10':
                        caseid_v3.caseid_giamsat10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat11':
                        caseid_v3.caseid_giamsat11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat12':
                        caseid_v3.caseid_giamsat12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat13':
                        caseid_v3.caseid_giamsat13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat14':
                        caseid_v3.caseid_giamsat14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat15':
                        caseid_v3.caseid_giamsat15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat16':
                        caseid_v3.caseid_giamsat16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat17':
                        caseid_v3.caseid_giamsat17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat18':
                        caseid_v3.caseid_giamsat18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat19':
                        caseid_v3.caseid_giamsat19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat20':
                        caseid_v3.caseid_giamsat20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat21':
                        caseid_v3.caseid_giamsat21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat22':
                        caseid_v3.caseid_giamsat22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat23':
                        caseid_v3.caseid_giamsat23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat24':
                        caseid_v3.caseid_giamsat24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat25':
                        caseid_v3.caseid_giamsat25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat26':
                        caseid_v3.caseid_giamsat26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat27':
                        caseid_v3.caseid_giamsat27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat28':
                        caseid_v3.caseid_giamsat28(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat29':
                        caseid_v3.caseid_giamsat29(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat30':
                        caseid_v3.caseid_giamsat30(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat31':
                        caseid_v3.caseid_giamsat31(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat32':
                        caseid_v3.caseid_giamsat32(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat33':
                        caseid_v3.caseid_giamsat33(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat34':
                        caseid_v3.caseid_giamsat34(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat35':
                        caseid_v3.caseid_giamsat35(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat36':
                        caseid_v3.caseid_giamsat36(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat37':
                        caseid_v3.caseid_giamsat37(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat38':
                        caseid_v3.caseid_giamsat38(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat39':
                        caseid_v3.caseid_giamsat39(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat40':
                        caseid_v3.caseid_giamsat40(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat41':
                        caseid_v3.caseid_giamsat41(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat42':
                        caseid_v3.caseid_giamsat42(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat43':
                        caseid_v3.caseid_giamsat43(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat44':
                        caseid_v3.caseid_giamsat44(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat45':
                        caseid_v3.caseid_giamsat45(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat46':
                        caseid_v3.caseid_giamsat46(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat47':
                        caseid_v3.caseid_giamsat47(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat48':
                        caseid_v3.caseid_giamsat48(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat49':
                        caseid_v3.caseid_giamsat49(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat50':
                        caseid_v3.caseid_giamsat50(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat51':
                        caseid_v3.caseid_giamsat51(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat52':
                        caseid_v3.caseid_giamsat52(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat53':
                        caseid_v3.caseid_giamsat53(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat54':
                        caseid_v3.caseid_giamsat54(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat55':
                        caseid_v3.caseid_giamsat55(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat56':
                        caseid_v3.caseid_giamsat56(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat57':
                        caseid_v3.caseid_giamsat57(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat58':
                        caseid_v3.caseid_giamsat58(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat59':
                        caseid_v3.caseid_giamsat59(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat60':
                        caseid_v3.caseid_giamsat60(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat61':
                        caseid_v3.caseid_giamsat61(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat62':
                        caseid_v3.caseid_giamsat62(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat63':
                        caseid_v3.caseid_giamsat63(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat64':
                        caseid_v3.caseid_giamsat64(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat65':
                        caseid_v3.caseid_giamsat65(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat66':
                        caseid_v3.caseid_giamsat66(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat67':
                        caseid_v3.caseid_giamsat67(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat68':
                        caseid_v3.caseid_giamsat68(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat69':
                        caseid_v3.caseid_giamsat69(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat70':
                        caseid_v3.caseid_giamsat70(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat71':
                        caseid_v3.caseid_giamsat71(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat72':
                        caseid_v3.caseid_giamsat72(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat73':
                        caseid_v3.caseid_giamsat73(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat74':
                        caseid_v3.caseid_giamsat74(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat75':
                        caseid_v3.caseid_giamsat75(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat76':
                        caseid_v3.caseid_giamsat76(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat77':
                        caseid_v3.caseid_giamsat77(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat78':
                        caseid_v3.caseid_giamsat78(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat79':
                        caseid_v3.caseid_giamsat79(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat80':
                        caseid_v3.caseid_giamsat80(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat81':
                        caseid_v3.caseid_giamsat81(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat82':
                        caseid_v3.caseid_giamsat82(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat83':
                        caseid_v3.caseid_giamsat83(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat84':
                        caseid_v3.caseid_giamsat84(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat85':
                        caseid_v3.caseid_giamsat85(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat86':
                        caseid_v3.caseid_giamsat86(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat87':
                        caseid_v3.caseid_giamsat87(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat88':
                        caseid_v3.caseid_giamsat88(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat89':
                        caseid_v3.caseid_giamsat89(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat90':
                        caseid_v3.caseid_giamsat90(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat91':
                        caseid_v3.caseid_giamsat91(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat92':
                        caseid_v3.caseid_giamsat92(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat93':
                        caseid_v3.caseid_giamsat93(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat94':
                        caseid_v3.caseid_giamsat94(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat95':
                        caseid_v3.caseid_giamsat95(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat96':
                        caseid_v3.caseid_giamsat96(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat97':
                        caseid_v3.caseid_giamsat97(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat98':
                        caseid_v3.caseid_giamsat98(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat99':
                        caseid_v3.caseid_giamsat99(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat100':
                        caseid_v3.caseid_giamsat100(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat101':
                        caseid_v3.caseid_giamsat101(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat102':
                        caseid_v3.caseid_giamsat102(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat103':
                        caseid_v3.caseid_giamsat103(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat104':
                        caseid_v3.caseid_giamsat104(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat105':
                        caseid_v3.caseid_giamsat105(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat106':
                        caseid_v3.caseid_giamsat106(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat107':
                        caseid_v3.caseid_giamsat107(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat108':
                        caseid_v3.caseid_giamsat108(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat109':
                        caseid_v3.caseid_giamsat109(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat110':
                        caseid_v3.caseid_giamsat110(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat111':
                        caseid_v3.caseid_giamsat111(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat112':
                        caseid_v3.caseid_giamsat112(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat113':
                        caseid_v3.caseid_giamsat113(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat114':
                        caseid_v3.caseid_giamsat114(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat115':
                        caseid_v3.caseid_giamsat115(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat116':
                        caseid_v3.caseid_giamsat116(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat117':
                        caseid_v3.caseid_giamsat117(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat118':
                        caseid_v3.caseid_giamsat118(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat119':
                        caseid_v3.caseid_giamsat119(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat120':
                        caseid_v3.caseid_giamsat120(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat121':
                        caseid_v3.caseid_giamsat121(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat122':
                        caseid_v3.caseid_giamsat122(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat123':
                        caseid_v3.caseid_giamsat123(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat124':
                        caseid_v3.caseid_giamsat124(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat125':
                        caseid_v3.caseid_giamsat125(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat126':
                        caseid_v3.caseid_giamsat126(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat127':
                        caseid_v3.caseid_giamsat127(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat128':
                        caseid_v3.caseid_giamsat128(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat129':
                        caseid_v3.caseid_giamsat129(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat130':
                        caseid_v3.caseid_giamsat130(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat131':
                        caseid_v3.caseid_giamsat131(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat132':
                        caseid_v3.caseid_giamsat132(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat133':
                        caseid_v3.caseid_giamsat133(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat134':
                        caseid_v3.caseid_giamsat134(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat135':
                        caseid_v3.caseid_giamsat135(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat136':
                        caseid_v3.caseid_giamsat136(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat137':
                        caseid_v3.caseid_giamsat137(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat138':
                        caseid_v3.caseid_giamsat138(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat139':
                        caseid_v3.caseid_giamsat139(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat140':
                        caseid_v3.caseid_giamsat140(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat141':
                        caseid_v3.caseid_giamsat141(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat142':
                        caseid_v3.caseid_giamsat142(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat143':
                        caseid_v3.caseid_giamsat143(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat144':
                        caseid_v3.caseid_giamsat144(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat145':
                        caseid_v3.caseid_giamsat145(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat146':
                        caseid_v3.caseid_giamsat146(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat147':
                        caseid_v3.caseid_giamsat147(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat148':
                        caseid_v3.caseid_giamsat148(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat149':
                        caseid_v3.caseid_giamsat149(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat150':
                        caseid_v3.caseid_giamsat150(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat151':
                        caseid_v3.caseid_giamsat151(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat152':
                        caseid_v3.caseid_giamsat152(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat153':
                        caseid_v3.caseid_giamsat153(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat154':
                        caseid_v3.caseid_giamsat154(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat155':
                        caseid_v3.caseid_giamsat155(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat156':
                        caseid_v3.caseid_giamsat156(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat157':
                        caseid_v3.caseid_giamsat157(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat158':
                        caseid_v3.caseid_giamsat158(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat159':
                        caseid_v3.caseid_giamsat159(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat160':
                        caseid_v3.caseid_giamsat160(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat161':
                        caseid_v3.caseid_giamsat161(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat162':
                        caseid_v3.caseid_giamsat162(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat163':
                        caseid_v3.caseid_giamsat163(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat164':
                        caseid_v3.caseid_giamsat164(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat165':
                        caseid_v3.caseid_giamsat165(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat166':
                        caseid_v3.caseid_giamsat166(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat167':
                        caseid_v3.caseid_giamsat167(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat168':
                        caseid_v3.caseid_giamsat168(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat169':
                        caseid_v3.caseid_giamsat169(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat170':
                        caseid_v3.caseid_giamsat170(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat171':
                        caseid_v3.caseid_giamsat171(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat172':
                        caseid_v3.caseid_giamsat172(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat173':
                        caseid_v3.caseid_giamsat173(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat174':
                        caseid_v3.caseid_giamsat174(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat175':
                        caseid_v3.caseid_giamsat175(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat176':
                        caseid_v3.caseid_giamsat176(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat177':
                        caseid_v3.caseid_giamsat177(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat178':
                        caseid_v3.caseid_giamsat178(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat179':
                        caseid_v3.caseid_giamsat179(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat180':
                        caseid_v3.caseid_giamsat180(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat181':
                        caseid_v3.caseid_giamsat181(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat182':
                        caseid_v3.caseid_giamsat182(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat183':
                        caseid_v3.caseid_giamsat183(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat184':
                        caseid_v3.caseid_giamsat184(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat185':
                        caseid_v3.caseid_giamsat185(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat186':
                        caseid_v3.caseid_giamsat186(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat187':
                        caseid_v3.caseid_giamsat187(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat188':
                        caseid_v3.caseid_giamsat188(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat189':
                        caseid_v3.caseid_giamsat189(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat190':
                        caseid_v3.caseid_giamsat190(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat191':
                        caseid_v3.caseid_giamsat191(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat192':
                        caseid_v3.caseid_giamsat192(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat193':
                        caseid_v3.caseid_giamsat193(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat194':
                        caseid_v3.caseid_giamsat194(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat195':
                        caseid_v3.caseid_giamsat195(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat196':
                        caseid_v3.caseid_giamsat196(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat197':
                        caseid_v3.caseid_giamsat197(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat198':
                        caseid_v3.caseid_giamsat198(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat199':
                        caseid_v3.caseid_giamsat199(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat200':
                        caseid_v3.caseid_giamsat200(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat201':
                        caseid_v3.caseid_giamsat201(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat202':
                        caseid_v3.caseid_giamsat202(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat203':
                        caseid_v3.caseid_giamsat203(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat204':
                        caseid_v3.caseid_giamsat204(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat205':
                        caseid_v3.caseid_giamsat205(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat206':
                        caseid_v3.caseid_giamsat206(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat207':
                        caseid_v3.caseid_giamsat207(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat208':
                        caseid_v3.caseid_giamsat208(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat209':
                        caseid_v3.caseid_giamsat209(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat210':
                        caseid_v3.caseid_giamsat210(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat211':
                        caseid_v3.caseid_giamsat211(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat212':
                        caseid_v3.caseid_giamsat212(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat213':
                        caseid_v3.caseid_giamsat213(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat214':
                        caseid_v3.caseid_giamsat214(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat215':
                        caseid_v3.caseid_giamsat215(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat216':
                        caseid_v3.caseid_giamsat216(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat217':
                        caseid_v3.caseid_giamsat217(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat218':
                        caseid_v3.caseid_giamsat218(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat219':
                        caseid_v3.caseid_giamsat219(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat220':
                        caseid_v3.caseid_giamsat220(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat221':
                        caseid_v3.caseid_giamsat221(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat222':
                        caseid_v3.caseid_giamsat222(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat223':
                        caseid_v3.caseid_giamsat223(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat224':
                        caseid_v3.caseid_giamsat224(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat225':
                        caseid_v3.caseid_giamsat225(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat226':
                        caseid_v3.caseid_giamsat226(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat227':
                        caseid_v3.caseid_giamsat227(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat228':
                        caseid_v3.caseid_giamsat228(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat229':
                        caseid_v3.caseid_giamsat229(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "4":
            print(list_mucdo4)
            for case in list_mucdo4:
                try:
                    if case == 'GiamSat01':
                        caseid_v3.caseid_giamsat01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat02':
                        caseid_v3.caseid_giamsat02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat03':
                        caseid_v3.caseid_giamsat03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat04':
                        caseid_v3.caseid_giamsat04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat05':
                        caseid_v3.caseid_giamsat05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat06':
                        caseid_v3.caseid_giamsat06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat07':
                        caseid_v3.caseid_giamsat07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat08':
                        caseid_v3.caseid_giamsat08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat09':
                        caseid_v3.caseid_giamsat09(self)
                except:
                    module_other_v3.swich_tab_0()

                try:
                    if case == 'GiamSat10':
                        caseid_v3.caseid_giamsat10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat11':
                        caseid_v3.caseid_giamsat11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat12':
                        caseid_v3.caseid_giamsat12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat13':
                        caseid_v3.caseid_giamsat13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat14':
                        caseid_v3.caseid_giamsat14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat15':
                        caseid_v3.caseid_giamsat15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat16':
                        caseid_v3.caseid_giamsat16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat17':
                        caseid_v3.caseid_giamsat17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat18':
                        caseid_v3.caseid_giamsat18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat19':
                        caseid_v3.caseid_giamsat19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat20':
                        caseid_v3.caseid_giamsat20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat21':
                        caseid_v3.caseid_giamsat21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat22':
                        caseid_v3.caseid_giamsat22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat23':
                        caseid_v3.caseid_giamsat23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat24':
                        caseid_v3.caseid_giamsat24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat25':
                        caseid_v3.caseid_giamsat25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat26':
                        caseid_v3.caseid_giamsat26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat27':
                        caseid_v3.caseid_giamsat27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat28':
                        caseid_v3.caseid_giamsat28(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat29':
                        caseid_v3.caseid_giamsat29(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat30':
                        caseid_v3.caseid_giamsat30(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat31':
                        caseid_v3.caseid_giamsat31(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat32':
                        caseid_v3.caseid_giamsat32(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat33':
                        caseid_v3.caseid_giamsat33(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat34':
                        caseid_v3.caseid_giamsat34(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat35':
                        caseid_v3.caseid_giamsat35(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat36':
                        caseid_v3.caseid_giamsat36(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat37':
                        caseid_v3.caseid_giamsat37(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat38':
                        caseid_v3.caseid_giamsat38(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat39':
                        caseid_v3.caseid_giamsat39(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat40':
                        caseid_v3.caseid_giamsat40(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat41':
                        caseid_v3.caseid_giamsat41(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat42':
                        caseid_v3.caseid_giamsat42(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat43':
                        caseid_v3.caseid_giamsat43(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat44':
                        caseid_v3.caseid_giamsat44(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat45':
                        caseid_v3.caseid_giamsat45(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat46':
                        caseid_v3.caseid_giamsat46(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat47':
                        caseid_v3.caseid_giamsat47(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat48':
                        caseid_v3.caseid_giamsat48(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat49':
                        caseid_v3.caseid_giamsat49(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat50':
                        caseid_v3.caseid_giamsat50(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat51':
                        caseid_v3.caseid_giamsat51(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat52':
                        caseid_v3.caseid_giamsat52(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat53':
                        caseid_v3.caseid_giamsat53(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat54':
                        caseid_v3.caseid_giamsat54(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat55':
                        caseid_v3.caseid_giamsat55(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat56':
                        caseid_v3.caseid_giamsat56(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat57':
                        caseid_v3.caseid_giamsat57(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat58':
                        caseid_v3.caseid_giamsat58(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat59':
                        caseid_v3.caseid_giamsat59(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat60':
                        caseid_v3.caseid_giamsat60(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat61':
                        caseid_v3.caseid_giamsat61(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat62':
                        caseid_v3.caseid_giamsat62(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat63':
                        caseid_v3.caseid_giamsat63(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat64':
                        caseid_v3.caseid_giamsat64(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat65':
                        caseid_v3.caseid_giamsat65(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat66':
                        caseid_v3.caseid_giamsat66(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat67':
                        caseid_v3.caseid_giamsat67(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat68':
                        caseid_v3.caseid_giamsat68(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat69':
                        caseid_v3.caseid_giamsat69(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat70':
                        caseid_v3.caseid_giamsat70(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat71':
                        caseid_v3.caseid_giamsat71(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat72':
                        caseid_v3.caseid_giamsat72(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat73':
                        caseid_v3.caseid_giamsat73(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat74':
                        caseid_v3.caseid_giamsat74(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat75':
                        caseid_v3.caseid_giamsat75(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat76':
                        caseid_v3.caseid_giamsat76(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat77':
                        caseid_v3.caseid_giamsat77(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat78':
                        caseid_v3.caseid_giamsat78(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat79':
                        caseid_v3.caseid_giamsat79(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat80':
                        caseid_v3.caseid_giamsat80(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat81':
                        caseid_v3.caseid_giamsat81(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat82':
                        caseid_v3.caseid_giamsat82(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat83':
                        caseid_v3.caseid_giamsat83(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat84':
                        caseid_v3.caseid_giamsat84(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat85':
                        caseid_v3.caseid_giamsat85(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat86':
                        caseid_v3.caseid_giamsat86(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat87':
                        caseid_v3.caseid_giamsat87(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat88':
                        caseid_v3.caseid_giamsat88(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat89':
                        caseid_v3.caseid_giamsat89(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat90':
                        caseid_v3.caseid_giamsat90(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat91':
                        caseid_v3.caseid_giamsat91(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat92':
                        caseid_v3.caseid_giamsat92(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat93':
                        caseid_v3.caseid_giamsat93(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat94':
                        caseid_v3.caseid_giamsat94(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat95':
                        caseid_v3.caseid_giamsat95(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat96':
                        caseid_v3.caseid_giamsat96(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat97':
                        caseid_v3.caseid_giamsat97(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat98':
                        caseid_v3.caseid_giamsat98(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat99':
                        caseid_v3.caseid_giamsat99(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat100':
                        caseid_v3.caseid_giamsat100(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat101':
                        caseid_v3.caseid_giamsat101(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat102':
                        caseid_v3.caseid_giamsat102(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat103':
                        caseid_v3.caseid_giamsat103(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat104':
                        caseid_v3.caseid_giamsat104(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat105':
                        caseid_v3.caseid_giamsat105(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat106':
                        caseid_v3.caseid_giamsat106(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat107':
                        caseid_v3.caseid_giamsat107(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat108':
                        caseid_v3.caseid_giamsat108(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat109':
                        caseid_v3.caseid_giamsat109(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat110':
                        caseid_v3.caseid_giamsat110(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat111':
                        caseid_v3.caseid_giamsat111(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat112':
                        caseid_v3.caseid_giamsat112(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat113':
                        caseid_v3.caseid_giamsat113(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat114':
                        caseid_v3.caseid_giamsat114(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat115':
                        caseid_v3.caseid_giamsat115(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat116':
                        caseid_v3.caseid_giamsat116(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat117':
                        caseid_v3.caseid_giamsat117(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat118':
                        caseid_v3.caseid_giamsat118(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat119':
                        caseid_v3.caseid_giamsat119(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat120':
                        caseid_v3.caseid_giamsat120(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat121':
                        caseid_v3.caseid_giamsat121(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat122':
                        caseid_v3.caseid_giamsat122(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat123':
                        caseid_v3.caseid_giamsat123(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat124':
                        caseid_v3.caseid_giamsat124(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat125':
                        caseid_v3.caseid_giamsat125(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat126':
                        caseid_v3.caseid_giamsat126(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat127':
                        caseid_v3.caseid_giamsat127(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat128':
                        caseid_v3.caseid_giamsat128(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat129':
                        caseid_v3.caseid_giamsat129(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat130':
                        caseid_v3.caseid_giamsat130(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat131':
                        caseid_v3.caseid_giamsat131(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat132':
                        caseid_v3.caseid_giamsat132(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat133':
                        caseid_v3.caseid_giamsat133(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat134':
                        caseid_v3.caseid_giamsat134(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat135':
                        caseid_v3.caseid_giamsat135(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat136':
                        caseid_v3.caseid_giamsat136(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat137':
                        caseid_v3.caseid_giamsat137(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat138':
                        caseid_v3.caseid_giamsat138(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat139':
                        caseid_v3.caseid_giamsat139(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat140':
                        caseid_v3.caseid_giamsat140(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat141':
                        caseid_v3.caseid_giamsat141(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat142':
                        caseid_v3.caseid_giamsat142(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat143':
                        caseid_v3.caseid_giamsat143(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat144':
                        caseid_v3.caseid_giamsat144(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat145':
                        caseid_v3.caseid_giamsat145(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat146':
                        caseid_v3.caseid_giamsat146(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat147':
                        caseid_v3.caseid_giamsat147(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat148':
                        caseid_v3.caseid_giamsat148(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat149':
                        caseid_v3.caseid_giamsat149(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat150':
                        caseid_v3.caseid_giamsat150(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat151':
                        caseid_v3.caseid_giamsat151(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat152':
                        caseid_v3.caseid_giamsat152(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat153':
                        caseid_v3.caseid_giamsat153(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat154':
                        caseid_v3.caseid_giamsat154(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat155':
                        caseid_v3.caseid_giamsat155(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat156':
                        caseid_v3.caseid_giamsat156(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat157':
                        caseid_v3.caseid_giamsat157(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat158':
                        caseid_v3.caseid_giamsat158(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat159':
                        caseid_v3.caseid_giamsat159(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat160':
                        caseid_v3.caseid_giamsat160(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat161':
                        caseid_v3.caseid_giamsat161(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat162':
                        caseid_v3.caseid_giamsat162(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat163':
                        caseid_v3.caseid_giamsat163(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat164':
                        caseid_v3.caseid_giamsat164(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat165':
                        caseid_v3.caseid_giamsat165(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat166':
                        caseid_v3.caseid_giamsat166(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat167':
                        caseid_v3.caseid_giamsat167(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat168':
                        caseid_v3.caseid_giamsat168(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat169':
                        caseid_v3.caseid_giamsat169(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat170':
                        caseid_v3.caseid_giamsat170(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat171':
                        caseid_v3.caseid_giamsat171(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat172':
                        caseid_v3.caseid_giamsat172(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat173':
                        caseid_v3.caseid_giamsat173(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat174':
                        caseid_v3.caseid_giamsat174(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat175':
                        caseid_v3.caseid_giamsat175(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat176':
                        caseid_v3.caseid_giamsat176(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat177':
                        caseid_v3.caseid_giamsat177(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat178':
                        caseid_v3.caseid_giamsat178(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat179':
                        caseid_v3.caseid_giamsat179(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat180':
                        caseid_v3.caseid_giamsat180(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat181':
                        caseid_v3.caseid_giamsat181(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat182':
                        caseid_v3.caseid_giamsat182(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat183':
                        caseid_v3.caseid_giamsat183(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat184':
                        caseid_v3.caseid_giamsat184(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat185':
                        caseid_v3.caseid_giamsat185(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat186':
                        caseid_v3.caseid_giamsat186(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat187':
                        caseid_v3.caseid_giamsat187(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat188':
                        caseid_v3.caseid_giamsat188(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat189':
                        caseid_v3.caseid_giamsat189(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat190':
                        caseid_v3.caseid_giamsat190(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat191':
                        caseid_v3.caseid_giamsat191(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat192':
                        caseid_v3.caseid_giamsat192(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat193':
                        caseid_v3.caseid_giamsat193(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat194':
                        caseid_v3.caseid_giamsat194(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat195':
                        caseid_v3.caseid_giamsat195(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat196':
                        caseid_v3.caseid_giamsat196(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat197':
                        caseid_v3.caseid_giamsat197(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat198':
                        caseid_v3.caseid_giamsat198(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat199':
                        caseid_v3.caseid_giamsat199(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat200':
                        caseid_v3.caseid_giamsat200(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat201':
                        caseid_v3.caseid_giamsat201(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat202':
                        caseid_v3.caseid_giamsat202(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat203':
                        caseid_v3.caseid_giamsat203(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat204':
                        caseid_v3.caseid_giamsat204(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat205':
                        caseid_v3.caseid_giamsat205(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat206':
                        caseid_v3.caseid_giamsat206(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat207':
                        caseid_v3.caseid_giamsat207(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat208':
                        caseid_v3.caseid_giamsat208(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat209':
                        caseid_v3.caseid_giamsat209(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat210':
                        caseid_v3.caseid_giamsat210(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat211':
                        caseid_v3.caseid_giamsat211(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat212':
                        caseid_v3.caseid_giamsat212(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat213':
                        caseid_v3.caseid_giamsat213(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat214':
                        caseid_v3.caseid_giamsat214(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat215':
                        caseid_v3.caseid_giamsat215(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat216':
                        caseid_v3.caseid_giamsat216(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat217':
                        caseid_v3.caseid_giamsat217(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat218':
                        caseid_v3.caseid_giamsat218(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat219':
                        caseid_v3.caseid_giamsat219(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat220':
                        caseid_v3.caseid_giamsat220(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat221':
                        caseid_v3.caseid_giamsat221(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat222':
                        caseid_v3.caseid_giamsat222(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat223':
                        caseid_v3.caseid_giamsat223(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat224':
                        caseid_v3.caseid_giamsat224(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat225':
                        caseid_v3.caseid_giamsat225(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat226':
                        caseid_v3.caseid_giamsat226(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat227':
                        caseid_v3.caseid_giamsat227(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat228':
                        caseid_v3.caseid_giamsat228(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'GiamSat229':
                        caseid_v3.caseid_giamsat229(self)
                except:
                    module_other_v3.swich_tab_0()



#3 lộ trình
def route(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_v3.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 344
    while (rownum < 420):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print(list_mucdo1)

    modetest = ''.join(re.findall(r'\d+', var_v3.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Route01':
                        caseid_v3.caseid_route01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route02':
                        caseid_v3.caseid_route02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route03':
                        caseid_v3.caseid_route03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route04':
                        caseid_v3.caseid_route04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route05':
                        caseid_v3.caseid_route05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route06':
                        caseid_v3.caseid_route06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route07':
                        caseid_v3.caseid_route07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route08':
                        caseid_v3.caseid_route08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route09':
                        caseid_v3.caseid_route09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route10':
                        caseid_v3.caseid_route10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route11':
                        caseid_v3.caseid_route11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route12':
                        caseid_v3.caseid_route12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route13':
                        caseid_v3.caseid_route13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route14':
                        caseid_v3.caseid_route14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route15':
                        caseid_v3.caseid_route15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route16':
                        caseid_v3.caseid_route16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route17':
                        caseid_v3.caseid_route17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route18':
                        caseid_v3.caseid_route18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route19':
                        caseid_v3.caseid_route19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route20':
                        caseid_v3.caseid_route20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route21':
                        caseid_v3.caseid_route21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route22':
                        caseid_v3.caseid_route22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route23':
                        caseid_v3.caseid_route23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route24':
                        caseid_v3.caseid_route24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route25':
                        caseid_v3.caseid_route25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route26':
                        caseid_v3.caseid_route26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route27':
                        caseid_v3.caseid_route27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route28':
                        caseid_v3.caseid_route28(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route29':
                        caseid_v3.caseid_route29(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route30':
                        caseid_v3.caseid_route30(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route31':
                        caseid_v3.caseid_route31(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route32':
                        caseid_v3.caseid_route32(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route33':
                        caseid_v3.caseid_route33(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route34':
                        caseid_v3.caseid_route34(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route35':
                        caseid_v3.caseid_route35(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route36':
                        caseid_v3.caseid_route36(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route37':
                        caseid_v3.caseid_route37(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route38':
                        caseid_v3.caseid_route38(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route39':
                        caseid_v3.caseid_route39(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route40':
                        caseid_v3.caseid_route40(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route41':
                        caseid_v3.caseid_route41(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route42':
                        caseid_v3.caseid_route42(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route43':
                        caseid_v3.caseid_route43(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route44':
                        caseid_v3.caseid_route44(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route45':
                        caseid_v3.caseid_route45(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route46':
                        caseid_v3.caseid_route46(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route47':
                        caseid_v3.caseid_route47(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route48':
                        caseid_v3.caseid_route48(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route49':
                        caseid_v3.caseid_route49(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route50':
                        caseid_v3.caseid_route50(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route51':
                        caseid_v3.caseid_route51(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route52':
                        caseid_v3.caseid_route52(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route53':
                        caseid_v3.caseid_route53(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Route01':
                        caseid_v3.caseid_route01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route02':
                        caseid_v3.caseid_route02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route03':
                        caseid_v3.caseid_route03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route04':
                        caseid_v3.caseid_route04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route05':
                        caseid_v3.caseid_route05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route06':
                        caseid_v3.caseid_route06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route07':
                        caseid_v3.caseid_route07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route08':
                        caseid_v3.caseid_route08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route09':
                        caseid_v3.caseid_route09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route10':
                        caseid_v3.caseid_route10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route11':
                        caseid_v3.caseid_route11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route12':
                        caseid_v3.caseid_route12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route13':
                        caseid_v3.caseid_route13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route14':
                        caseid_v3.caseid_route14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route15':
                        caseid_v3.caseid_route15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route16':
                        caseid_v3.caseid_route16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route17':
                        caseid_v3.caseid_route17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route18':
                        caseid_v3.caseid_route18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route19':
                        caseid_v3.caseid_route19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route20':
                        caseid_v3.caseid_route20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route21':
                        caseid_v3.caseid_route21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route22':
                        caseid_v3.caseid_route22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route23':
                        caseid_v3.caseid_route23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route24':
                        caseid_v3.caseid_route24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route25':
                        caseid_v3.caseid_route25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route26':
                        caseid_v3.caseid_route26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route27':
                        caseid_v3.caseid_route27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route28':
                        caseid_v3.caseid_route28(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route29':
                        caseid_v3.caseid_route29(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route30':
                        caseid_v3.caseid_route30(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route31':
                        caseid_v3.caseid_route31(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route32':
                        caseid_v3.caseid_route32(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route33':
                        caseid_v3.caseid_route33(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route34':
                        caseid_v3.caseid_route34(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route35':
                        caseid_v3.caseid_route35(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route36':
                        caseid_v3.caseid_route36(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route37':
                        caseid_v3.caseid_route37(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route38':
                        caseid_v3.caseid_route38(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route39':
                        caseid_v3.caseid_route39(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route40':
                        caseid_v3.caseid_route40(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route41':
                        caseid_v3.caseid_route41(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route42':
                        caseid_v3.caseid_route42(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route43':
                        caseid_v3.caseid_route43(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route44':
                        caseid_v3.caseid_route44(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route45':
                        caseid_v3.caseid_route45(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route46':
                        caseid_v3.caseid_route46(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route47':
                        caseid_v3.caseid_route47(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route48':
                        caseid_v3.caseid_route48(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route49':
                        caseid_v3.caseid_route49(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route50':
                        caseid_v3.caseid_route50(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route51':
                        caseid_v3.caseid_route51(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route52':
                        caseid_v3.caseid_route52(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route53':
                        caseid_v3.caseid_route53(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Route01':
                        caseid_v3.caseid_route01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route02':
                        caseid_v3.caseid_route02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route03':
                        caseid_v3.caseid_route03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route04':
                        caseid_v3.caseid_route04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route05':
                        caseid_v3.caseid_route05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route06':
                        caseid_v3.caseid_route06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route07':
                        caseid_v3.caseid_route07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route08':
                        caseid_v3.caseid_route08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route09':
                        caseid_v3.caseid_route09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route10':
                        caseid_v3.caseid_route10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route11':
                        caseid_v3.caseid_route11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route12':
                        caseid_v3.caseid_route12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route13':
                        caseid_v3.caseid_route13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route14':
                        caseid_v3.caseid_route14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route15':
                        caseid_v3.caseid_route15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route16':
                        caseid_v3.caseid_route16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route17':
                        caseid_v3.caseid_route17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route18':
                        caseid_v3.caseid_route18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route19':
                        caseid_v3.caseid_route19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route20':
                        caseid_v3.caseid_route20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route21':
                        caseid_v3.caseid_route21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route22':
                        caseid_v3.caseid_route22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route23':
                        caseid_v3.caseid_route23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route24':
                        caseid_v3.caseid_route24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route25':
                        caseid_v3.caseid_route25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route26':
                        caseid_v3.caseid_route26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route27':
                        caseid_v3.caseid_route27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route28':
                        caseid_v3.caseid_route28(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route29':
                        caseid_v3.caseid_route29(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route30':
                        caseid_v3.caseid_route30(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route31':
                        caseid_v3.caseid_route31(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route32':
                        caseid_v3.caseid_route32(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route33':
                        caseid_v3.caseid_route33(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route34':
                        caseid_v3.caseid_route34(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route35':
                        caseid_v3.caseid_route35(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route36':
                        caseid_v3.caseid_route36(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route37':
                        caseid_v3.caseid_route37(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route38':
                        caseid_v3.caseid_route38(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route39':
                        caseid_v3.caseid_route39(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route40':
                        caseid_v3.caseid_route40(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route41':
                        caseid_v3.caseid_route41(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route42':
                        caseid_v3.caseid_route42(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route43':
                        caseid_v3.caseid_route43(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route44':
                        caseid_v3.caseid_route44(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route45':
                        caseid_v3.caseid_route45(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route46':
                        caseid_v3.caseid_route46(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route47':
                        caseid_v3.caseid_route47(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route48':
                        caseid_v3.caseid_route48(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route49':
                        caseid_v3.caseid_route49(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route50':
                        caseid_v3.caseid_route50(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route51':
                        caseid_v3.caseid_route51(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route52':
                        caseid_v3.caseid_route52(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route53':
                        caseid_v3.caseid_route53(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Route01':
                        caseid_v3.caseid_route01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route02':
                        caseid_v3.caseid_route02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route03':
                        caseid_v3.caseid_route03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route04':
                        caseid_v3.caseid_route04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route05':
                        caseid_v3.caseid_route05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route06':
                        caseid_v3.caseid_route06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route07':
                        caseid_v3.caseid_route07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route08':
                        caseid_v3.caseid_route08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route09':
                        caseid_v3.caseid_route09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route10':
                        caseid_v3.caseid_route10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route11':
                        caseid_v3.caseid_route11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route12':
                        caseid_v3.caseid_route12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route13':
                        caseid_v3.caseid_route13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route14':
                        caseid_v3.caseid_route14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route15':
                        caseid_v3.caseid_route15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route16':
                        caseid_v3.caseid_route16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route17':
                        caseid_v3.caseid_route17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route18':
                        caseid_v3.caseid_route18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route19':
                        caseid_v3.caseid_route19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route20':
                        caseid_v3.caseid_route20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route21':
                        caseid_v3.caseid_route21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route22':
                        caseid_v3.caseid_route22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route23':
                        caseid_v3.caseid_route23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route24':
                        caseid_v3.caseid_route24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route25':
                        caseid_v3.caseid_route25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route26':
                        caseid_v3.caseid_route26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route27':
                        caseid_v3.caseid_route27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route28':
                        caseid_v3.caseid_route28(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route29':
                        caseid_v3.caseid_route29(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route30':
                        caseid_v3.caseid_route30(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route31':
                        caseid_v3.caseid_route31(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route32':
                        caseid_v3.caseid_route32(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route33':
                        caseid_v3.caseid_route33(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route34':
                        caseid_v3.caseid_route34(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route35':
                        caseid_v3.caseid_route35(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route36':
                        caseid_v3.caseid_route36(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route37':
                        caseid_v3.caseid_route37(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route38':
                        caseid_v3.caseid_route38(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route39':
                        caseid_v3.caseid_route39(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route40':
                        caseid_v3.caseid_route40(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route41':
                        caseid_v3.caseid_route41(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route42':
                        caseid_v3.caseid_route42(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route43':
                        caseid_v3.caseid_route43(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route44':
                        caseid_v3.caseid_route44(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route45':
                        caseid_v3.caseid_route45(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route46':
                        caseid_v3.caseid_route46(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route47':
                        caseid_v3.caseid_route47(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route48':
                        caseid_v3.caseid_route48(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route49':
                        caseid_v3.caseid_route49(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route50':
                        caseid_v3.caseid_route50(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route51':
                        caseid_v3.caseid_route51(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route52':
                        caseid_v3.caseid_route52(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Route53':
                        caseid_v3.caseid_route53(self)
                except:
                    module_other_v3.swich_tab_0()



#4 quản trị
def administration(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_v3.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 421
    while (rownum < 661):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print(list_mucdo1)

    modetest = ''.join(re.findall(r'\d+', var_v3.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Admin01':
                        caseid_v3.caseid_admin01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin02':
                        caseid_v3.caseid_admin02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin03':
                        caseid_v3.caseid_admin03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin04':
                        caseid_v3.caseid_admin04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin05':
                        caseid_v3.caseid_admin05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin06':
                        caseid_v3.caseid_admin06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin07':
                        caseid_v3.caseid_admin07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin08':
                        caseid_v3.caseid_admin08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin09':
                        caseid_v3.caseid_admin09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin10':
                        caseid_v3.caseid_admin10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin11':
                        caseid_v3.caseid_admin11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin12':
                        caseid_v3.caseid_admin12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin13':
                        caseid_v3.caseid_admin13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin14':
                        caseid_v3.caseid_admin14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin15':
                        caseid_v3.caseid_admin15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin16':
                        caseid_v3.caseid_admin16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin17':
                        caseid_v3.caseid_admin17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin18':
                        caseid_v3.caseid_admin18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin19':
                        caseid_v3.caseid_admin19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin20':
                        caseid_v3.caseid_admin20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin21':
                        caseid_v3.caseid_admin21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin22':
                        caseid_v3.caseid_admin22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin23':
                        caseid_v3.caseid_admin23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin24':
                        caseid_v3.caseid_admin24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin25':
                        caseid_v3.caseid_admin25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin26':
                        caseid_v3.caseid_admin26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin27':
                        caseid_v3.caseid_admin27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin28':
                        caseid_v3.caseid_admin28(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin29':
                        caseid_v3.caseid_admin29(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin30':
                        caseid_v3.caseid_admin30(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin31':
                        caseid_v3.caseid_admin31(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin32':
                        caseid_v3.caseid_admin32(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin33':
                        caseid_v3.caseid_admin33(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin34':
                        caseid_v3.caseid_admin34(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin35':
                        caseid_v3.caseid_admin35(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin36':
                        caseid_v3.caseid_admin36(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin37':
                        caseid_v3.caseid_admin37(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin38':
                        caseid_v3.caseid_admin38(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin39':
                        caseid_v3.caseid_admin39(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin40':
                        caseid_v3.caseid_admin40(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin41':
                        caseid_v3.caseid_admin41(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin42':
                        caseid_v3.caseid_admin42(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin43':
                        caseid_v3.caseid_admin43(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin44':
                        caseid_v3.caseid_admin44(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin45':
                        caseid_v3.caseid_admin45(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin46':
                        caseid_v3.caseid_admin46(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin47':
                        caseid_v3.caseid_admin47(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin48':
                        caseid_v3.caseid_admin48(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin49':
                        caseid_v3.caseid_admin49(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin50':
                        caseid_v3.caseid_admin50(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin51':
                        caseid_v3.caseid_admin51(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin52':
                        caseid_v3.caseid_admin52(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin53':
                        caseid_v3.caseid_admin53(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin54':
                        caseid_v3.caseid_admin54(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin55':
                        caseid_v3.caseid_admin55(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin56':
                        caseid_v3.caseid_admin56(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin57':
                        caseid_v3.caseid_admin57(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin58':
                        caseid_v3.caseid_admin58(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin59':
                        caseid_v3.caseid_admin59(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin60':
                        caseid_v3.caseid_admin60(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin61':
                        caseid_v3.caseid_admin61(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin62':
                        caseid_v3.caseid_admin62(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin63':
                        caseid_v3.caseid_admin63(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin64':
                        caseid_v3.caseid_admin64(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin65':
                        caseid_v3.caseid_admin65(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin66':
                        caseid_v3.caseid_admin66(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin67':
                        caseid_v3.caseid_admin67(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin68':
                        caseid_v3.caseid_admin68(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin69':
                        caseid_v3.caseid_admin69(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin70':
                        caseid_v3.caseid_admin70(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin71':
                        caseid_v3.caseid_admin71(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin72':
                        caseid_v3.caseid_admin72(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin73':
                        caseid_v3.caseid_admin73(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin74':
                        caseid_v3.caseid_admin74(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin75':
                        caseid_v3.caseid_admin75(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin76':
                        caseid_v3.caseid_admin76(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin77':
                        caseid_v3.caseid_admin77(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin78':
                        caseid_v3.caseid_admin78(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin79':
                        caseid_v3.caseid_admin79(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin80':
                        caseid_v3.caseid_admin80(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin81':
                        caseid_v3.caseid_admin81(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin82':
                        caseid_v3.caseid_admin82(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin83':
                        caseid_v3.caseid_admin83(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin84':
                        caseid_v3.caseid_admin84(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin85':
                        caseid_v3.caseid_admin85(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin86':
                        caseid_v3.caseid_admin86(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin87':
                        caseid_v3.caseid_admin87(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin88':
                        caseid_v3.caseid_admin88(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin89':
                        caseid_v3.caseid_admin89(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin90':
                        caseid_v3.caseid_admin90(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin91':
                        caseid_v3.caseid_admin91(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin92':
                        caseid_v3.caseid_admin92(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin93':
                        caseid_v3.caseid_admin93(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin94':
                        caseid_v3.caseid_admin94(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin95':
                        caseid_v3.caseid_admin95(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin96':
                        caseid_v3.caseid_admin96(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin97':
                        caseid_v3.caseid_admin97(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin98':
                        caseid_v3.caseid_admin98(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin99':
                        caseid_v3.caseid_admin99(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin100':
                        caseid_v3.caseid_admin100(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin101':
                        caseid_v3.caseid_admin101(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin102':
                        caseid_v3.caseid_admin102(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin103':
                        caseid_v3.caseid_admin103(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin104':
                        caseid_v3.caseid_admin104(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin105':
                        caseid_v3.caseid_admin105(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin106':
                        caseid_v3.caseid_admin106(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin107':
                        caseid_v3.caseid_admin107(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin108':
                        caseid_v3.caseid_admin108(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin109':
                        caseid_v3.caseid_admin109(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin110':
                        caseid_v3.caseid_admin110(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin111':
                        caseid_v3.caseid_admin111(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin112':
                        caseid_v3.caseid_admin112(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin113':
                        caseid_v3.caseid_admin113(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin114':
                        caseid_v3.caseid_admin114(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin115':
                        caseid_v3.caseid_admin115(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin116':
                        caseid_v3.caseid_admin116(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin117':
                        caseid_v3.caseid_admin117(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin118':
                        caseid_v3.caseid_admin118(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin119':
                        caseid_v3.caseid_admin119(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin120':
                        caseid_v3.caseid_admin120(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin121':
                        caseid_v3.caseid_admin121(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin122':
                        caseid_v3.caseid_admin122(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin123':
                        caseid_v3.caseid_admin123(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin124':
                        caseid_v3.caseid_admin124(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin125':
                        caseid_v3.caseid_admin125(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin126':
                        caseid_v3.caseid_admin126(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin127':
                        caseid_v3.caseid_admin127(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin128':
                        caseid_v3.caseid_admin128(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin129':
                        caseid_v3.caseid_admin129(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin130':
                        caseid_v3.caseid_admin130(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin131':
                        caseid_v3.caseid_admin131(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin132':
                        caseid_v3.caseid_admin132(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin133':
                        caseid_v3.caseid_admin133(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin134':
                        caseid_v3.caseid_admin134(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin135':
                        caseid_v3.caseid_admin135(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin136':
                        caseid_v3.caseid_admin136(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin137':
                        caseid_v3.caseid_admin137(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin138':
                        caseid_v3.caseid_admin138(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin139':
                        caseid_v3.caseid_admin139(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin140':
                        caseid_v3.caseid_admin140(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin141':
                        caseid_v3.caseid_admin141(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin142':
                        caseid_v3.caseid_admin142(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin143':
                        caseid_v3.caseid_admin143(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin144':
                        caseid_v3.caseid_admin144(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin145':
                        caseid_v3.caseid_admin145(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin146':
                        caseid_v3.caseid_admin146(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin147':
                        caseid_v3.caseid_admin147(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin148':
                        caseid_v3.caseid_admin148(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin149':
                        caseid_v3.caseid_admin149(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin150':
                        caseid_v3.caseid_admin150(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin151':
                        caseid_v3.caseid_admin151(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin152':
                        caseid_v3.caseid_admin152(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin153':
                        caseid_v3.caseid_admin153(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin154':
                        caseid_v3.caseid_admin154(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin155':
                        caseid_v3.caseid_admin155(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin156':
                        caseid_v3.caseid_admin156(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin157':
                        caseid_v3.caseid_admin157(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin158':
                        caseid_v3.caseid_admin158(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin159':
                        caseid_v3.caseid_admin159(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin160':
                        caseid_v3.caseid_admin160(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin161':
                        caseid_v3.caseid_admin161(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin162':
                        caseid_v3.caseid_admin162(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin163':
                        caseid_v3.caseid_admin163(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin164':
                        caseid_v3.caseid_admin164(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin165':
                        caseid_v3.caseid_admin165(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin166':
                        caseid_v3.caseid_admin166(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin167':
                        caseid_v3.caseid_admin167(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin168':
                        caseid_v3.caseid_admin168(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin169':
                        caseid_v3.caseid_admin169(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin170':
                        caseid_v3.caseid_admin170(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin171':
                        caseid_v3.caseid_admin171(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin172':
                        caseid_v3.caseid_admin172(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin173':
                        caseid_v3.caseid_admin173(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin174':
                        caseid_v3.caseid_admin174(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin175':
                        caseid_v3.caseid_admin175(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin176':
                        caseid_v3.caseid_admin176(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin177':
                        caseid_v3.caseid_admin177(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin178':
                        caseid_v3.caseid_admin178(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin179':
                        caseid_v3.caseid_admin179(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin180':
                        caseid_v3.caseid_admin180(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin181':
                        caseid_v3.caseid_admin181(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin182':
                        caseid_v3.caseid_admin182(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin183':
                        caseid_v3.caseid_admin183(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin184':
                        caseid_v3.caseid_admin184(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin185':
                        caseid_v3.caseid_admin185(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin186':
                        caseid_v3.caseid_admin186(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin187':
                        caseid_v3.caseid_admin187(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin188':
                        caseid_v3.caseid_admin188(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin189':
                        caseid_v3.caseid_admin189(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin190':
                        caseid_v3.caseid_admin190(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin191':
                        caseid_v3.caseid_admin191(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin192':
                        caseid_v3.caseid_admin192(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin193':
                        caseid_v3.caseid_admin193(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin194':
                        caseid_v3.caseid_admin194(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin195':
                        caseid_v3.caseid_admin195(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin196':
                        caseid_v3.caseid_admin196(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin197':
                        caseid_v3.caseid_admin197(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin198':
                        caseid_v3.caseid_admin198(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin199':
                        caseid_v3.caseid_admin199(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Admin01':
                        caseid_v3.caseid_admin01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin02':
                        caseid_v3.caseid_admin02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin03':
                        caseid_v3.caseid_admin03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin04':
                        caseid_v3.caseid_admin04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin05':
                        caseid_v3.caseid_admin05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin06':
                        caseid_v3.caseid_admin06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin07':
                        caseid_v3.caseid_admin07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin08':
                        caseid_v3.caseid_admin08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin09':
                        caseid_v3.caseid_admin09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin10':
                        caseid_v3.caseid_admin10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin11':
                        caseid_v3.caseid_admin11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin12':
                        caseid_v3.caseid_admin12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin13':
                        caseid_v3.caseid_admin13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin14':
                        caseid_v3.caseid_admin14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin15':
                        caseid_v3.caseid_admin15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin16':
                        caseid_v3.caseid_admin16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin17':
                        caseid_v3.caseid_admin17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin18':
                        caseid_v3.caseid_admin18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin19':
                        caseid_v3.caseid_admin19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin20':
                        caseid_v3.caseid_admin20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin21':
                        caseid_v3.caseid_admin21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin22':
                        caseid_v3.caseid_admin22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin23':
                        caseid_v3.caseid_admin23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin24':
                        caseid_v3.caseid_admin24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin25':
                        caseid_v3.caseid_admin25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin26':
                        caseid_v3.caseid_admin26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin27':
                        caseid_v3.caseid_admin27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin28':
                        caseid_v3.caseid_admin28(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin29':
                        caseid_v3.caseid_admin29(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin30':
                        caseid_v3.caseid_admin30(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin31':
                        caseid_v3.caseid_admin31(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin32':
                        caseid_v3.caseid_admin32(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin33':
                        caseid_v3.caseid_admin33(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin34':
                        caseid_v3.caseid_admin34(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin35':
                        caseid_v3.caseid_admin35(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin36':
                        caseid_v3.caseid_admin36(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin37':
                        caseid_v3.caseid_admin37(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin38':
                        caseid_v3.caseid_admin38(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin39':
                        caseid_v3.caseid_admin39(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin40':
                        caseid_v3.caseid_admin40(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin41':
                        caseid_v3.caseid_admin41(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin42':
                        caseid_v3.caseid_admin42(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin43':
                        caseid_v3.caseid_admin43(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin44':
                        caseid_v3.caseid_admin44(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin45':
                        caseid_v3.caseid_admin45(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin46':
                        caseid_v3.caseid_admin46(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin47':
                        caseid_v3.caseid_admin47(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin48':
                        caseid_v3.caseid_admin48(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin49':
                        caseid_v3.caseid_admin49(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin50':
                        caseid_v3.caseid_admin50(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin51':
                        caseid_v3.caseid_admin51(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin52':
                        caseid_v3.caseid_admin52(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin53':
                        caseid_v3.caseid_admin53(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin54':
                        caseid_v3.caseid_admin54(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin55':
                        caseid_v3.caseid_admin55(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin56':
                        caseid_v3.caseid_admin56(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin57':
                        caseid_v3.caseid_admin57(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin58':
                        caseid_v3.caseid_admin58(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin59':
                        caseid_v3.caseid_admin59(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin60':
                        caseid_v3.caseid_admin60(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin61':
                        caseid_v3.caseid_admin61(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin62':
                        caseid_v3.caseid_admin62(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin63':
                        caseid_v3.caseid_admin63(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin64':
                        caseid_v3.caseid_admin64(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin65':
                        caseid_v3.caseid_admin65(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin66':
                        caseid_v3.caseid_admin66(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin67':
                        caseid_v3.caseid_admin67(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin68':
                        caseid_v3.caseid_admin68(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin69':
                        caseid_v3.caseid_admin69(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin70':
                        caseid_v3.caseid_admin70(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin71':
                        caseid_v3.caseid_admin71(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin72':
                        caseid_v3.caseid_admin72(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin73':
                        caseid_v3.caseid_admin73(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin74':
                        caseid_v3.caseid_admin74(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin75':
                        caseid_v3.caseid_admin75(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin76':
                        caseid_v3.caseid_admin76(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin77':
                        caseid_v3.caseid_admin77(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin78':
                        caseid_v3.caseid_admin78(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin79':
                        caseid_v3.caseid_admin79(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin80':
                        caseid_v3.caseid_admin80(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin81':
                        caseid_v3.caseid_admin81(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin82':
                        caseid_v3.caseid_admin82(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin83':
                        caseid_v3.caseid_admin83(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin84':
                        caseid_v3.caseid_admin84(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin85':
                        caseid_v3.caseid_admin85(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin86':
                        caseid_v3.caseid_admin86(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin87':
                        caseid_v3.caseid_admin87(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin88':
                        caseid_v3.caseid_admin88(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin89':
                        caseid_v3.caseid_admin89(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin90':
                        caseid_v3.caseid_admin90(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin91':
                        caseid_v3.caseid_admin91(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin92':
                        caseid_v3.caseid_admin92(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin93':
                        caseid_v3.caseid_admin93(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin94':
                        caseid_v3.caseid_admin94(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin95':
                        caseid_v3.caseid_admin95(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin96':
                        caseid_v3.caseid_admin96(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin97':
                        caseid_v3.caseid_admin97(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin98':
                        caseid_v3.caseid_admin98(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin99':
                        caseid_v3.caseid_admin99(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin100':
                        caseid_v3.caseid_admin100(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin101':
                        caseid_v3.caseid_admin101(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin102':
                        caseid_v3.caseid_admin102(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin103':
                        caseid_v3.caseid_admin103(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin104':
                        caseid_v3.caseid_admin104(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin105':
                        caseid_v3.caseid_admin105(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin106':
                        caseid_v3.caseid_admin106(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin107':
                        caseid_v3.caseid_admin107(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin108':
                        caseid_v3.caseid_admin108(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin109':
                        caseid_v3.caseid_admin109(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin110':
                        caseid_v3.caseid_admin110(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin111':
                        caseid_v3.caseid_admin111(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin112':
                        caseid_v3.caseid_admin112(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin113':
                        caseid_v3.caseid_admin113(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin114':
                        caseid_v3.caseid_admin114(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin115':
                        caseid_v3.caseid_admin115(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin116':
                        caseid_v3.caseid_admin116(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin117':
                        caseid_v3.caseid_admin117(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin118':
                        caseid_v3.caseid_admin118(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin119':
                        caseid_v3.caseid_admin119(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin120':
                        caseid_v3.caseid_admin120(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin121':
                        caseid_v3.caseid_admin121(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin122':
                        caseid_v3.caseid_admin122(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin123':
                        caseid_v3.caseid_admin123(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin124':
                        caseid_v3.caseid_admin124(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin125':
                        caseid_v3.caseid_admin125(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin126':
                        caseid_v3.caseid_admin126(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin127':
                        caseid_v3.caseid_admin127(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin128':
                        caseid_v3.caseid_admin128(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin129':
                        caseid_v3.caseid_admin129(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin130':
                        caseid_v3.caseid_admin130(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin131':
                        caseid_v3.caseid_admin131(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin132':
                        caseid_v3.caseid_admin132(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin133':
                        caseid_v3.caseid_admin133(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin134':
                        caseid_v3.caseid_admin134(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin135':
                        caseid_v3.caseid_admin135(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin136':
                        caseid_v3.caseid_admin136(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin137':
                        caseid_v3.caseid_admin137(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin138':
                        caseid_v3.caseid_admin138(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin139':
                        caseid_v3.caseid_admin139(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin140':
                        caseid_v3.caseid_admin140(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin141':
                        caseid_v3.caseid_admin141(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin142':
                        caseid_v3.caseid_admin142(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin143':
                        caseid_v3.caseid_admin143(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin144':
                        caseid_v3.caseid_admin144(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin145':
                        caseid_v3.caseid_admin145(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin146':
                        caseid_v3.caseid_admin146(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin147':
                        caseid_v3.caseid_admin147(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin148':
                        caseid_v3.caseid_admin148(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin149':
                        caseid_v3.caseid_admin149(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin150':
                        caseid_v3.caseid_admin150(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin151':
                        caseid_v3.caseid_admin151(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin152':
                        caseid_v3.caseid_admin152(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin153':
                        caseid_v3.caseid_admin153(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin154':
                        caseid_v3.caseid_admin154(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin155':
                        caseid_v3.caseid_admin155(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin156':
                        caseid_v3.caseid_admin156(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin157':
                        caseid_v3.caseid_admin157(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin158':
                        caseid_v3.caseid_admin158(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin159':
                        caseid_v3.caseid_admin159(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin160':
                        caseid_v3.caseid_admin160(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin161':
                        caseid_v3.caseid_admin161(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin162':
                        caseid_v3.caseid_admin162(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin163':
                        caseid_v3.caseid_admin163(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin164':
                        caseid_v3.caseid_admin164(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin165':
                        caseid_v3.caseid_admin165(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin166':
                        caseid_v3.caseid_admin166(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin167':
                        caseid_v3.caseid_admin167(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin168':
                        caseid_v3.caseid_admin168(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin169':
                        caseid_v3.caseid_admin169(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin170':
                        caseid_v3.caseid_admin170(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin171':
                        caseid_v3.caseid_admin171(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin172':
                        caseid_v3.caseid_admin172(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin173':
                        caseid_v3.caseid_admin173(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin174':
                        caseid_v3.caseid_admin174(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin175':
                        caseid_v3.caseid_admin175(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin176':
                        caseid_v3.caseid_admin176(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin177':
                        caseid_v3.caseid_admin177(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin178':
                        caseid_v3.caseid_admin178(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin179':
                        caseid_v3.caseid_admin179(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin180':
                        caseid_v3.caseid_admin180(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin181':
                        caseid_v3.caseid_admin181(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin182':
                        caseid_v3.caseid_admin182(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin183':
                        caseid_v3.caseid_admin183(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin184':
                        caseid_v3.caseid_admin184(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin185':
                        caseid_v3.caseid_admin185(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin186':
                        caseid_v3.caseid_admin186(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin187':
                        caseid_v3.caseid_admin187(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin188':
                        caseid_v3.caseid_admin188(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin189':
                        caseid_v3.caseid_admin189(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin190':
                        caseid_v3.caseid_admin190(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin191':
                        caseid_v3.caseid_admin191(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin192':
                        caseid_v3.caseid_admin192(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin193':
                        caseid_v3.caseid_admin193(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin194':
                        caseid_v3.caseid_admin194(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin195':
                        caseid_v3.caseid_admin195(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin196':
                        caseid_v3.caseid_admin196(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin197':
                        caseid_v3.caseid_admin197(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin198':
                        caseid_v3.caseid_admin198(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin199':
                        caseid_v3.caseid_admin199(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Admin01':
                        caseid_v3.caseid_admin01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin02':
                        caseid_v3.caseid_admin02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin03':
                        caseid_v3.caseid_admin03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin04':
                        caseid_v3.caseid_admin04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin05':
                        caseid_v3.caseid_admin05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin06':
                        caseid_v3.caseid_admin06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin07':
                        caseid_v3.caseid_admin07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin08':
                        caseid_v3.caseid_admin08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin09':
                        caseid_v3.caseid_admin09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin10':
                        caseid_v3.caseid_admin10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin11':
                        caseid_v3.caseid_admin11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin12':
                        caseid_v3.caseid_admin12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin13':
                        caseid_v3.caseid_admin13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin14':
                        caseid_v3.caseid_admin14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin15':
                        caseid_v3.caseid_admin15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin16':
                        caseid_v3.caseid_admin16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin17':
                        caseid_v3.caseid_admin17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin18':
                        caseid_v3.caseid_admin18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin19':
                        caseid_v3.caseid_admin19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin20':
                        caseid_v3.caseid_admin20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin21':
                        caseid_v3.caseid_admin21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin22':
                        caseid_v3.caseid_admin22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin23':
                        caseid_v3.caseid_admin23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin24':
                        caseid_v3.caseid_admin24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin25':
                        caseid_v3.caseid_admin25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin26':
                        caseid_v3.caseid_admin26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin27':
                        caseid_v3.caseid_admin27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin28':
                        caseid_v3.caseid_admin28(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin29':
                        caseid_v3.caseid_admin29(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin30':
                        caseid_v3.caseid_admin30(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin31':
                        caseid_v3.caseid_admin31(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin32':
                        caseid_v3.caseid_admin32(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin33':
                        caseid_v3.caseid_admin33(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin34':
                        caseid_v3.caseid_admin34(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin35':
                        caseid_v3.caseid_admin35(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin36':
                        caseid_v3.caseid_admin36(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin37':
                        caseid_v3.caseid_admin37(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin38':
                        caseid_v3.caseid_admin38(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin39':
                        caseid_v3.caseid_admin39(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin40':
                        caseid_v3.caseid_admin40(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin41':
                        caseid_v3.caseid_admin41(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin42':
                        caseid_v3.caseid_admin42(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin43':
                        caseid_v3.caseid_admin43(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin44':
                        caseid_v3.caseid_admin44(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin45':
                        caseid_v3.caseid_admin45(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin46':
                        caseid_v3.caseid_admin46(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin47':
                        caseid_v3.caseid_admin47(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin48':
                        caseid_v3.caseid_admin48(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin49':
                        caseid_v3.caseid_admin49(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin50':
                        caseid_v3.caseid_admin50(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin51':
                        caseid_v3.caseid_admin51(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin52':
                        caseid_v3.caseid_admin52(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin53':
                        caseid_v3.caseid_admin53(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin54':
                        caseid_v3.caseid_admin54(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin55':
                        caseid_v3.caseid_admin55(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin56':
                        caseid_v3.caseid_admin56(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin57':
                        caseid_v3.caseid_admin57(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin58':
                        caseid_v3.caseid_admin58(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin59':
                        caseid_v3.caseid_admin59(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin60':
                        caseid_v3.caseid_admin60(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin61':
                        caseid_v3.caseid_admin61(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin62':
                        caseid_v3.caseid_admin62(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin63':
                        caseid_v3.caseid_admin63(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin64':
                        caseid_v3.caseid_admin64(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin65':
                        caseid_v3.caseid_admin65(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin66':
                        caseid_v3.caseid_admin66(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin67':
                        caseid_v3.caseid_admin67(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin68':
                        caseid_v3.caseid_admin68(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin69':
                        caseid_v3.caseid_admin69(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin70':
                        caseid_v3.caseid_admin70(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin71':
                        caseid_v3.caseid_admin71(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin72':
                        caseid_v3.caseid_admin72(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin73':
                        caseid_v3.caseid_admin73(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin74':
                        caseid_v3.caseid_admin74(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin75':
                        caseid_v3.caseid_admin75(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin76':
                        caseid_v3.caseid_admin76(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin77':
                        caseid_v3.caseid_admin77(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin78':
                        caseid_v3.caseid_admin78(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin79':
                        caseid_v3.caseid_admin79(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin80':
                        caseid_v3.caseid_admin80(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin81':
                        caseid_v3.caseid_admin81(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin82':
                        caseid_v3.caseid_admin82(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin83':
                        caseid_v3.caseid_admin83(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin84':
                        caseid_v3.caseid_admin84(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin85':
                        caseid_v3.caseid_admin85(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin86':
                        caseid_v3.caseid_admin86(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin87':
                        caseid_v3.caseid_admin87(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin88':
                        caseid_v3.caseid_admin88(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin89':
                        caseid_v3.caseid_admin89(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin90':
                        caseid_v3.caseid_admin90(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin91':
                        caseid_v3.caseid_admin91(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin92':
                        caseid_v3.caseid_admin92(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin93':
                        caseid_v3.caseid_admin93(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin94':
                        caseid_v3.caseid_admin94(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin95':
                        caseid_v3.caseid_admin95(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin96':
                        caseid_v3.caseid_admin96(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin97':
                        caseid_v3.caseid_admin97(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin98':
                        caseid_v3.caseid_admin98(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin99':
                        caseid_v3.caseid_admin99(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin100':
                        caseid_v3.caseid_admin100(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin101':
                        caseid_v3.caseid_admin101(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin102':
                        caseid_v3.caseid_admin102(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin103':
                        caseid_v3.caseid_admin103(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin104':
                        caseid_v3.caseid_admin104(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin105':
                        caseid_v3.caseid_admin105(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin106':
                        caseid_v3.caseid_admin106(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin107':
                        caseid_v3.caseid_admin107(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin108':
                        caseid_v3.caseid_admin108(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin109':
                        caseid_v3.caseid_admin109(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin110':
                        caseid_v3.caseid_admin110(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin111':
                        caseid_v3.caseid_admin111(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin112':
                        caseid_v3.caseid_admin112(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin113':
                        caseid_v3.caseid_admin113(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin114':
                        caseid_v3.caseid_admin114(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin115':
                        caseid_v3.caseid_admin115(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin116':
                        caseid_v3.caseid_admin116(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin117':
                        caseid_v3.caseid_admin117(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin118':
                        caseid_v3.caseid_admin118(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin119':
                        caseid_v3.caseid_admin119(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin120':
                        caseid_v3.caseid_admin120(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin121':
                        caseid_v3.caseid_admin121(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin122':
                        caseid_v3.caseid_admin122(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin123':
                        caseid_v3.caseid_admin123(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin124':
                        caseid_v3.caseid_admin124(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin125':
                        caseid_v3.caseid_admin125(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin126':
                        caseid_v3.caseid_admin126(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin127':
                        caseid_v3.caseid_admin127(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin128':
                        caseid_v3.caseid_admin128(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin129':
                        caseid_v3.caseid_admin129(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin130':
                        caseid_v3.caseid_admin130(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin131':
                        caseid_v3.caseid_admin131(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin132':
                        caseid_v3.caseid_admin132(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin133':
                        caseid_v3.caseid_admin133(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin134':
                        caseid_v3.caseid_admin134(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin135':
                        caseid_v3.caseid_admin135(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin136':
                        caseid_v3.caseid_admin136(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin137':
                        caseid_v3.caseid_admin137(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin138':
                        caseid_v3.caseid_admin138(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin139':
                        caseid_v3.caseid_admin139(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin140':
                        caseid_v3.caseid_admin140(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin141':
                        caseid_v3.caseid_admin141(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin142':
                        caseid_v3.caseid_admin142(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin143':
                        caseid_v3.caseid_admin143(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin144':
                        caseid_v3.caseid_admin144(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin145':
                        caseid_v3.caseid_admin145(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin146':
                        caseid_v3.caseid_admin146(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin147':
                        caseid_v3.caseid_admin147(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin148':
                        caseid_v3.caseid_admin148(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin149':
                        caseid_v3.caseid_admin149(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin150':
                        caseid_v3.caseid_admin150(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin151':
                        caseid_v3.caseid_admin151(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin152':
                        caseid_v3.caseid_admin152(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin153':
                        caseid_v3.caseid_admin153(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin154':
                        caseid_v3.caseid_admin154(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin155':
                        caseid_v3.caseid_admin155(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin156':
                        caseid_v3.caseid_admin156(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin157':
                        caseid_v3.caseid_admin157(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin158':
                        caseid_v3.caseid_admin158(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin159':
                        caseid_v3.caseid_admin159(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin160':
                        caseid_v3.caseid_admin160(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin161':
                        caseid_v3.caseid_admin161(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin162':
                        caseid_v3.caseid_admin162(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin163':
                        caseid_v3.caseid_admin163(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin164':
                        caseid_v3.caseid_admin164(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin165':
                        caseid_v3.caseid_admin165(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin166':
                        caseid_v3.caseid_admin166(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin167':
                        caseid_v3.caseid_admin167(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin168':
                        caseid_v3.caseid_admin168(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin169':
                        caseid_v3.caseid_admin169(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin170':
                        caseid_v3.caseid_admin170(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin171':
                        caseid_v3.caseid_admin171(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin172':
                        caseid_v3.caseid_admin172(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin173':
                        caseid_v3.caseid_admin173(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin174':
                        caseid_v3.caseid_admin174(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin175':
                        caseid_v3.caseid_admin175(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin176':
                        caseid_v3.caseid_admin176(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin177':
                        caseid_v3.caseid_admin177(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin178':
                        caseid_v3.caseid_admin178(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin179':
                        caseid_v3.caseid_admin179(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin180':
                        caseid_v3.caseid_admin180(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin181':
                        caseid_v3.caseid_admin181(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin182':
                        caseid_v3.caseid_admin182(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin183':
                        caseid_v3.caseid_admin183(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin184':
                        caseid_v3.caseid_admin184(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin185':
                        caseid_v3.caseid_admin185(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin186':
                        caseid_v3.caseid_admin186(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin187':
                        caseid_v3.caseid_admin187(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin188':
                        caseid_v3.caseid_admin188(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin189':
                        caseid_v3.caseid_admin189(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin190':
                        caseid_v3.caseid_admin190(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin191':
                        caseid_v3.caseid_admin191(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin192':
                        caseid_v3.caseid_admin192(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin193':
                        caseid_v3.caseid_admin193(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin194':
                        caseid_v3.caseid_admin194(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin195':
                        caseid_v3.caseid_admin195(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin196':
                        caseid_v3.caseid_admin196(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin197':
                        caseid_v3.caseid_admin197(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin198':
                        caseid_v3.caseid_admin198(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin199':
                        caseid_v3.caseid_admin199(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Admin01':
                        caseid_v3.caseid_admin01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin02':
                        caseid_v3.caseid_admin02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin03':
                        caseid_v3.caseid_admin03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin04':
                        caseid_v3.caseid_admin04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin05':
                        caseid_v3.caseid_admin05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin06':
                        caseid_v3.caseid_admin06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin07':
                        caseid_v3.caseid_admin07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin08':
                        caseid_v3.caseid_admin08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin09':
                        caseid_v3.caseid_admin09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin10':
                        caseid_v3.caseid_admin10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin11':
                        caseid_v3.caseid_admin11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin12':
                        caseid_v3.caseid_admin12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin13':
                        caseid_v3.caseid_admin13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin14':
                        caseid_v3.caseid_admin14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin15':
                        caseid_v3.caseid_admin15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin16':
                        caseid_v3.caseid_admin16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin17':
                        caseid_v3.caseid_admin17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin18':
                        caseid_v3.caseid_admin18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin19':
                        caseid_v3.caseid_admin19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin20':
                        caseid_v3.caseid_admin20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin21':
                        caseid_v3.caseid_admin21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin22':
                        caseid_v3.caseid_admin22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin23':
                        caseid_v3.caseid_admin23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin24':
                        caseid_v3.caseid_admin24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin25':
                        caseid_v3.caseid_admin25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin26':
                        caseid_v3.caseid_admin26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin27':
                        caseid_v3.caseid_admin27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin28':
                        caseid_v3.caseid_admin28(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin29':
                        caseid_v3.caseid_admin29(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin30':
                        caseid_v3.caseid_admin30(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin31':
                        caseid_v3.caseid_admin31(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin32':
                        caseid_v3.caseid_admin32(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin33':
                        caseid_v3.caseid_admin33(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin34':
                        caseid_v3.caseid_admin34(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin35':
                        caseid_v3.caseid_admin35(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin36':
                        caseid_v3.caseid_admin36(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin37':
                        caseid_v3.caseid_admin37(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin38':
                        caseid_v3.caseid_admin38(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin39':
                        caseid_v3.caseid_admin39(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin40':
                        caseid_v3.caseid_admin40(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin41':
                        caseid_v3.caseid_admin41(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin42':
                        caseid_v3.caseid_admin42(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin43':
                        caseid_v3.caseid_admin43(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin44':
                        caseid_v3.caseid_admin44(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin45':
                        caseid_v3.caseid_admin45(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin46':
                        caseid_v3.caseid_admin46(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin47':
                        caseid_v3.caseid_admin47(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin48':
                        caseid_v3.caseid_admin48(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin49':
                        caseid_v3.caseid_admin49(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin50':
                        caseid_v3.caseid_admin50(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin51':
                        caseid_v3.caseid_admin51(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin52':
                        caseid_v3.caseid_admin52(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin53':
                        caseid_v3.caseid_admin53(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin54':
                        caseid_v3.caseid_admin54(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin55':
                        caseid_v3.caseid_admin55(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin56':
                        caseid_v3.caseid_admin56(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin57':
                        caseid_v3.caseid_admin57(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin58':
                        caseid_v3.caseid_admin58(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin59':
                        caseid_v3.caseid_admin59(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin60':
                        caseid_v3.caseid_admin60(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin61':
                        caseid_v3.caseid_admin61(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin62':
                        caseid_v3.caseid_admin62(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin63':
                        caseid_v3.caseid_admin63(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin64':
                        caseid_v3.caseid_admin64(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin65':
                        caseid_v3.caseid_admin65(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin66':
                        caseid_v3.caseid_admin66(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin67':
                        caseid_v3.caseid_admin67(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin68':
                        caseid_v3.caseid_admin68(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin69':
                        caseid_v3.caseid_admin69(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin70':
                        caseid_v3.caseid_admin70(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin71':
                        caseid_v3.caseid_admin71(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin72':
                        caseid_v3.caseid_admin72(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin73':
                        caseid_v3.caseid_admin73(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin74':
                        caseid_v3.caseid_admin74(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin75':
                        caseid_v3.caseid_admin75(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin76':
                        caseid_v3.caseid_admin76(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin77':
                        caseid_v3.caseid_admin77(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin78':
                        caseid_v3.caseid_admin78(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin79':
                        caseid_v3.caseid_admin79(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin80':
                        caseid_v3.caseid_admin80(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin81':
                        caseid_v3.caseid_admin81(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin82':
                        caseid_v3.caseid_admin82(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin83':
                        caseid_v3.caseid_admin83(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin84':
                        caseid_v3.caseid_admin84(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin85':
                        caseid_v3.caseid_admin85(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin86':
                        caseid_v3.caseid_admin86(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin87':
                        caseid_v3.caseid_admin87(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin88':
                        caseid_v3.caseid_admin88(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin89':
                        caseid_v3.caseid_admin89(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin90':
                        caseid_v3.caseid_admin90(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin91':
                        caseid_v3.caseid_admin91(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin92':
                        caseid_v3.caseid_admin92(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin93':
                        caseid_v3.caseid_admin93(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin94':
                        caseid_v3.caseid_admin94(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin95':
                        caseid_v3.caseid_admin95(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin96':
                        caseid_v3.caseid_admin96(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin97':
                        caseid_v3.caseid_admin97(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin98':
                        caseid_v3.caseid_admin98(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin99':
                        caseid_v3.caseid_admin99(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin100':
                        caseid_v3.caseid_admin100(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin101':
                        caseid_v3.caseid_admin101(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin102':
                        caseid_v3.caseid_admin102(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin103':
                        caseid_v3.caseid_admin103(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin104':
                        caseid_v3.caseid_admin104(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin105':
                        caseid_v3.caseid_admin105(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin106':
                        caseid_v3.caseid_admin106(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin107':
                        caseid_v3.caseid_admin107(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin108':
                        caseid_v3.caseid_admin108(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin109':
                        caseid_v3.caseid_admin109(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin110':
                        caseid_v3.caseid_admin110(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin111':
                        caseid_v3.caseid_admin111(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin112':
                        caseid_v3.caseid_admin112(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin113':
                        caseid_v3.caseid_admin113(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin114':
                        caseid_v3.caseid_admin114(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin115':
                        caseid_v3.caseid_admin115(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin116':
                        caseid_v3.caseid_admin116(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin117':
                        caseid_v3.caseid_admin117(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin118':
                        caseid_v3.caseid_admin118(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin119':
                        caseid_v3.caseid_admin119(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin120':
                        caseid_v3.caseid_admin120(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin121':
                        caseid_v3.caseid_admin121(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin122':
                        caseid_v3.caseid_admin122(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin123':
                        caseid_v3.caseid_admin123(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin124':
                        caseid_v3.caseid_admin124(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin125':
                        caseid_v3.caseid_admin125(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin126':
                        caseid_v3.caseid_admin126(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin127':
                        caseid_v3.caseid_admin127(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin128':
                        caseid_v3.caseid_admin128(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin129':
                        caseid_v3.caseid_admin129(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin130':
                        caseid_v3.caseid_admin130(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin131':
                        caseid_v3.caseid_admin131(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin132':
                        caseid_v3.caseid_admin132(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin133':
                        caseid_v3.caseid_admin133(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin134':
                        caseid_v3.caseid_admin134(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin135':
                        caseid_v3.caseid_admin135(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin136':
                        caseid_v3.caseid_admin136(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin137':
                        caseid_v3.caseid_admin137(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin138':
                        caseid_v3.caseid_admin138(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin139':
                        caseid_v3.caseid_admin139(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin140':
                        caseid_v3.caseid_admin140(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin141':
                        caseid_v3.caseid_admin141(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin142':
                        caseid_v3.caseid_admin142(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin143':
                        caseid_v3.caseid_admin143(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin144':
                        caseid_v3.caseid_admin144(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin145':
                        caseid_v3.caseid_admin145(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin146':
                        caseid_v3.caseid_admin146(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin147':
                        caseid_v3.caseid_admin147(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin148':
                        caseid_v3.caseid_admin148(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin149':
                        caseid_v3.caseid_admin149(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin150':
                        caseid_v3.caseid_admin150(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin151':
                        caseid_v3.caseid_admin151(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin152':
                        caseid_v3.caseid_admin152(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin153':
                        caseid_v3.caseid_admin153(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin154':
                        caseid_v3.caseid_admin154(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin155':
                        caseid_v3.caseid_admin155(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin156':
                        caseid_v3.caseid_admin156(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin157':
                        caseid_v3.caseid_admin157(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin158':
                        caseid_v3.caseid_admin158(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin159':
                        caseid_v3.caseid_admin159(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin160':
                        caseid_v3.caseid_admin160(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin161':
                        caseid_v3.caseid_admin161(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin162':
                        caseid_v3.caseid_admin162(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin163':
                        caseid_v3.caseid_admin163(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin164':
                        caseid_v3.caseid_admin164(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin165':
                        caseid_v3.caseid_admin165(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin166':
                        caseid_v3.caseid_admin166(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin167':
                        caseid_v3.caseid_admin167(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin168':
                        caseid_v3.caseid_admin168(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin169':
                        caseid_v3.caseid_admin169(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin170':
                        caseid_v3.caseid_admin170(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin171':
                        caseid_v3.caseid_admin171(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin172':
                        caseid_v3.caseid_admin172(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin173':
                        caseid_v3.caseid_admin173(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin174':
                        caseid_v3.caseid_admin174(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin175':
                        caseid_v3.caseid_admin175(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin176':
                        caseid_v3.caseid_admin176(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin177':
                        caseid_v3.caseid_admin177(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin178':
                        caseid_v3.caseid_admin178(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin179':
                        caseid_v3.caseid_admin179(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin180':
                        caseid_v3.caseid_admin180(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin181':
                        caseid_v3.caseid_admin181(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin182':
                        caseid_v3.caseid_admin182(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin183':
                        caseid_v3.caseid_admin183(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin184':
                        caseid_v3.caseid_admin184(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin185':
                        caseid_v3.caseid_admin185(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin186':
                        caseid_v3.caseid_admin186(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin187':
                        caseid_v3.caseid_admin187(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin188':
                        caseid_v3.caseid_admin188(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin189':
                        caseid_v3.caseid_admin189(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin190':
                        caseid_v3.caseid_admin190(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin191':
                        caseid_v3.caseid_admin191(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin192':
                        caseid_v3.caseid_admin192(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin193':
                        caseid_v3.caseid_admin193(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin194':
                        caseid_v3.caseid_admin194(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin195':
                        caseid_v3.caseid_admin195(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin196':
                        caseid_v3.caseid_admin196(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin197':
                        caseid_v3.caseid_admin197(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin198':
                        caseid_v3.caseid_admin198(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Admin199':
                        caseid_v3.caseid_admin199(self)
                except:
                    module_other_v3.swich_tab_0()



#5 báo cáo
def report(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_v3.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 667
    while (rownum < 794):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print(list_mucdo1)

    modetest = ''.join(re.findall(r'\d+', var_v3.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Report01':
                        caseid_v3.caseid_report01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report02':
                        caseid_v3.caseid_report02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report03':
                        caseid_v3.caseid_report03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report04':
                        caseid_v3.caseid_report04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report05':
                        caseid_v3.caseid_report05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report06':
                        caseid_v3.caseid_report06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report07':
                        caseid_v3.caseid_report07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report08':
                        caseid_v3.caseid_report08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report09':
                        caseid_v3.caseid_report09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report10':
                        caseid_v3.caseid_report10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report11':
                        caseid_v3.caseid_report11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report12':
                        caseid_v3.caseid_report12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report13':
                        caseid_v3.caseid_report13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report14':
                        caseid_v3.caseid_report14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report15':
                        caseid_v3.caseid_report15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report16':
                        caseid_v3.caseid_report16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report17':
                        caseid_v3.caseid_report17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report18':
                        caseid_v3.caseid_report18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report19':
                        caseid_v3.caseid_report19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report20':
                        caseid_v3.caseid_report20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report21':
                        caseid_v3.caseid_report21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report22':
                        caseid_v3.caseid_report22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report23':
                        caseid_v3.caseid_report23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report24':
                        caseid_v3.caseid_report24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report25':
                        caseid_v3.caseid_report25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report26':
                        caseid_v3.caseid_report26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report27':
                        caseid_v3.caseid_report27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report28':
                        caseid_v3.caseid_report28(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report29':
                        caseid_v3.caseid_report29(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report30':
                        caseid_v3.caseid_report30(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report31':
                        caseid_v3.caseid_report31(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report32':
                        caseid_v3.caseid_report32(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report33':
                        caseid_v3.caseid_report33(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report34':
                        caseid_v3.caseid_report34(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report35':
                        caseid_v3.caseid_report35(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report36':
                        caseid_v3.caseid_report36(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report37':
                        caseid_v3.caseid_report37(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report38':
                        caseid_v3.caseid_report38(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report39':
                        caseid_v3.caseid_report39(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report40':
                        caseid_v3.caseid_report40(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report41':
                        caseid_v3.caseid_report41(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report42':
                        caseid_v3.caseid_report42(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report43':
                        caseid_v3.caseid_report43(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report44':
                        caseid_v3.caseid_report44(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report45':
                        caseid_v3.caseid_report45(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report46':
                        caseid_v3.caseid_report46(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report47':
                        caseid_v3.caseid_report47(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report48':
                        caseid_v3.caseid_report48(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report49':
                        caseid_v3.caseid_report49(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report50':
                        caseid_v3.caseid_report50(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report51':
                        caseid_v3.caseid_report51(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report52':
                        caseid_v3.caseid_report52(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report53':
                        caseid_v3.caseid_report53(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report54':
                        caseid_v3.caseid_report54(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report55':
                        caseid_v3.caseid_report55(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report56':
                        caseid_v3.caseid_report56(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report57':
                        caseid_v3.caseid_report57(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report58':
                        caseid_v3.caseid_report58(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report59':
                        caseid_v3.caseid_report59(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report60':
                        caseid_v3.caseid_report60(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report61':
                        caseid_v3.caseid_report61(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report62':
                        caseid_v3.caseid_report62(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report63':
                        caseid_v3.caseid_report63(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report64':
                        caseid_v3.caseid_report64(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report65':
                        caseid_v3.caseid_report65(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report66':
                        caseid_v3.caseid_report66(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report67':
                        caseid_v3.caseid_report67(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report68':
                        caseid_v3.caseid_report68(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report69':
                        caseid_v3.caseid_report69(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report70':
                        caseid_v3.caseid_report70(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report71':
                        caseid_v3.caseid_report71(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report72':
                        caseid_v3.caseid_report72(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report73':
                        caseid_v3.caseid_report73(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report74':
                        caseid_v3.caseid_report74(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report75':
                        caseid_v3.caseid_report75(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report76':
                        caseid_v3.caseid_report76(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report77':
                        caseid_v3.caseid_report77(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report78':
                        caseid_v3.caseid_report78(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report79':
                        caseid_v3.caseid_report79(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report80':
                        caseid_v3.caseid_report80(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report81':
                        caseid_v3.caseid_report81(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report82':
                        caseid_v3.caseid_report82(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report83':
                        caseid_v3.caseid_report83(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report84':
                        caseid_v3.caseid_report84(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report85':
                        caseid_v3.caseid_report85(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report86':
                        caseid_v3.caseid_report86(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report87':
                        caseid_v3.caseid_report87(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report88':
                        caseid_v3.caseid_report88(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report89':
                        caseid_v3.caseid_report89(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report90':
                        caseid_v3.caseid_report90(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report91':
                        caseid_v3.caseid_report91(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report92':
                        caseid_v3.caseid_report92(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report93':
                        caseid_v3.caseid_report93(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report94':
                        caseid_v3.caseid_report94(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report95':
                        caseid_v3.caseid_report95(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Report01':
                        caseid_v3.caseid_report01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report02':
                        caseid_v3.caseid_report02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report03':
                        caseid_v3.caseid_report03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report04':
                        caseid_v3.caseid_report04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report05':
                        caseid_v3.caseid_report05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report06':
                        caseid_v3.caseid_report06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report07':
                        caseid_v3.caseid_report07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report08':
                        caseid_v3.caseid_report08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report09':
                        caseid_v3.caseid_report09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report10':
                        caseid_v3.caseid_report10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report11':
                        caseid_v3.caseid_report11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report12':
                        caseid_v3.caseid_report12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report13':
                        caseid_v3.caseid_report13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report14':
                        caseid_v3.caseid_report14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report15':
                        caseid_v3.caseid_report15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report16':
                        caseid_v3.caseid_report16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report17':
                        caseid_v3.caseid_report17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report18':
                        caseid_v3.caseid_report18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report19':
                        caseid_v3.caseid_report19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report20':
                        caseid_v3.caseid_report20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report21':
                        caseid_v3.caseid_report21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report22':
                        caseid_v3.caseid_report22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report23':
                        caseid_v3.caseid_report23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report24':
                        caseid_v3.caseid_report24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report25':
                        caseid_v3.caseid_report25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report26':
                        caseid_v3.caseid_report26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report27':
                        caseid_v3.caseid_report27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report28':
                        caseid_v3.caseid_report28(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report29':
                        caseid_v3.caseid_report29(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report30':
                        caseid_v3.caseid_report30(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report31':
                        caseid_v3.caseid_report31(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report32':
                        caseid_v3.caseid_report32(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report33':
                        caseid_v3.caseid_report33(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report34':
                        caseid_v3.caseid_report34(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report35':
                        caseid_v3.caseid_report35(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report36':
                        caseid_v3.caseid_report36(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report37':
                        caseid_v3.caseid_report37(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report38':
                        caseid_v3.caseid_report38(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report39':
                        caseid_v3.caseid_report39(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report40':
                        caseid_v3.caseid_report40(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report41':
                        caseid_v3.caseid_report41(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report42':
                        caseid_v3.caseid_report42(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report43':
                        caseid_v3.caseid_report43(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report44':
                        caseid_v3.caseid_report44(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report45':
                        caseid_v3.caseid_report45(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report46':
                        caseid_v3.caseid_report46(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report47':
                        caseid_v3.caseid_report47(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report48':
                        caseid_v3.caseid_report48(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report49':
                        caseid_v3.caseid_report49(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report50':
                        caseid_v3.caseid_report50(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report51':
                        caseid_v3.caseid_report51(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report52':
                        caseid_v3.caseid_report52(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report53':
                        caseid_v3.caseid_report53(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report54':
                        caseid_v3.caseid_report54(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report55':
                        caseid_v3.caseid_report55(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report56':
                        caseid_v3.caseid_report56(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report57':
                        caseid_v3.caseid_report57(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report58':
                        caseid_v3.caseid_report58(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report59':
                        caseid_v3.caseid_report59(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report60':
                        caseid_v3.caseid_report60(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report61':
                        caseid_v3.caseid_report61(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report62':
                        caseid_v3.caseid_report62(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report63':
                        caseid_v3.caseid_report63(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report64':
                        caseid_v3.caseid_report64(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report65':
                        caseid_v3.caseid_report65(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report66':
                        caseid_v3.caseid_report66(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report67':
                        caseid_v3.caseid_report67(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report68':
                        caseid_v3.caseid_report68(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report69':
                        caseid_v3.caseid_report69(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report70':
                        caseid_v3.caseid_report70(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report71':
                        caseid_v3.caseid_report71(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report72':
                        caseid_v3.caseid_report72(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report73':
                        caseid_v3.caseid_report73(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report74':
                        caseid_v3.caseid_report74(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report75':
                        caseid_v3.caseid_report75(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report76':
                        caseid_v3.caseid_report76(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report77':
                        caseid_v3.caseid_report77(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report78':
                        caseid_v3.caseid_report78(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report79':
                        caseid_v3.caseid_report79(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report80':
                        caseid_v3.caseid_report80(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report81':
                        caseid_v3.caseid_report81(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report82':
                        caseid_v3.caseid_report82(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report83':
                        caseid_v3.caseid_report83(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report84':
                        caseid_v3.caseid_report84(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report85':
                        caseid_v3.caseid_report85(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report86':
                        caseid_v3.caseid_report86(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report87':
                        caseid_v3.caseid_report87(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report88':
                        caseid_v3.caseid_report88(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report89':
                        caseid_v3.caseid_report89(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report90':
                        caseid_v3.caseid_report90(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report91':
                        caseid_v3.caseid_report91(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report92':
                        caseid_v3.caseid_report92(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report93':
                        caseid_v3.caseid_report93(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report94':
                        caseid_v3.caseid_report94(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report95':
                        caseid_v3.caseid_report95(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Report01':
                        caseid_v3.caseid_report01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report02':
                        caseid_v3.caseid_report02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report03':
                        caseid_v3.caseid_report03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report04':
                        caseid_v3.caseid_report04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report05':
                        caseid_v3.caseid_report05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report06':
                        caseid_v3.caseid_report06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report07':
                        caseid_v3.caseid_report07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report08':
                        caseid_v3.caseid_report08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report09':
                        caseid_v3.caseid_report09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report10':
                        caseid_v3.caseid_report10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report11':
                        caseid_v3.caseid_report11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report12':
                        caseid_v3.caseid_report12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report13':
                        caseid_v3.caseid_report13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report14':
                        caseid_v3.caseid_report14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report15':
                        caseid_v3.caseid_report15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report16':
                        caseid_v3.caseid_report16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report17':
                        caseid_v3.caseid_report17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report18':
                        caseid_v3.caseid_report18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report19':
                        caseid_v3.caseid_report19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report20':
                        caseid_v3.caseid_report20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report21':
                        caseid_v3.caseid_report21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report22':
                        caseid_v3.caseid_report22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report23':
                        caseid_v3.caseid_report23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report24':
                        caseid_v3.caseid_report24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report25':
                        caseid_v3.caseid_report25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report26':
                        caseid_v3.caseid_report26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report27':
                        caseid_v3.caseid_report27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report28':
                        caseid_v3.caseid_report28(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report29':
                        caseid_v3.caseid_report29(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report30':
                        caseid_v3.caseid_report30(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report31':
                        caseid_v3.caseid_report31(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report32':
                        caseid_v3.caseid_report32(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report33':
                        caseid_v3.caseid_report33(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report34':
                        caseid_v3.caseid_report34(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report35':
                        caseid_v3.caseid_report35(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report36':
                        caseid_v3.caseid_report36(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report37':
                        caseid_v3.caseid_report37(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report38':
                        caseid_v3.caseid_report38(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report39':
                        caseid_v3.caseid_report39(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report40':
                        caseid_v3.caseid_report40(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report41':
                        caseid_v3.caseid_report41(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report42':
                        caseid_v3.caseid_report42(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report43':
                        caseid_v3.caseid_report43(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report44':
                        caseid_v3.caseid_report44(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report45':
                        caseid_v3.caseid_report45(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report46':
                        caseid_v3.caseid_report46(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report47':
                        caseid_v3.caseid_report47(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report48':
                        caseid_v3.caseid_report48(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report49':
                        caseid_v3.caseid_report49(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report50':
                        caseid_v3.caseid_report50(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report51':
                        caseid_v3.caseid_report51(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report52':
                        caseid_v3.caseid_report52(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report53':
                        caseid_v3.caseid_report53(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report54':
                        caseid_v3.caseid_report54(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report55':
                        caseid_v3.caseid_report55(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report56':
                        caseid_v3.caseid_report56(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report57':
                        caseid_v3.caseid_report57(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report58':
                        caseid_v3.caseid_report58(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report59':
                        caseid_v3.caseid_report59(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report60':
                        caseid_v3.caseid_report60(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report61':
                        caseid_v3.caseid_report61(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report62':
                        caseid_v3.caseid_report62(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report63':
                        caseid_v3.caseid_report63(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report64':
                        caseid_v3.caseid_report64(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report65':
                        caseid_v3.caseid_report65(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report66':
                        caseid_v3.caseid_report66(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report67':
                        caseid_v3.caseid_report67(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report68':
                        caseid_v3.caseid_report68(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report69':
                        caseid_v3.caseid_report69(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report70':
                        caseid_v3.caseid_report70(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report71':
                        caseid_v3.caseid_report71(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report72':
                        caseid_v3.caseid_report72(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report73':
                        caseid_v3.caseid_report73(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report74':
                        caseid_v3.caseid_report74(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report75':
                        caseid_v3.caseid_report75(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report76':
                        caseid_v3.caseid_report76(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report77':
                        caseid_v3.caseid_report77(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report78':
                        caseid_v3.caseid_report78(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report79':
                        caseid_v3.caseid_report79(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report80':
                        caseid_v3.caseid_report80(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report81':
                        caseid_v3.caseid_report81(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report82':
                        caseid_v3.caseid_report82(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report83':
                        caseid_v3.caseid_report83(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report84':
                        caseid_v3.caseid_report84(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report85':
                        caseid_v3.caseid_report85(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report86':
                        caseid_v3.caseid_report86(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report87':
                        caseid_v3.caseid_report87(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report88':
                        caseid_v3.caseid_report88(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report89':
                        caseid_v3.caseid_report89(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report90':
                        caseid_v3.caseid_report90(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report91':
                        caseid_v3.caseid_report91(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report92':
                        caseid_v3.caseid_report92(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report93':
                        caseid_v3.caseid_report93(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report94':
                        caseid_v3.caseid_report94(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report95':
                        caseid_v3.caseid_report95(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Report01':
                        caseid_v3.caseid_report01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report02':
                        caseid_v3.caseid_report02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report03':
                        caseid_v3.caseid_report03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report04':
                        caseid_v3.caseid_report04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report05':
                        caseid_v3.caseid_report05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report06':
                        caseid_v3.caseid_report06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report07':
                        caseid_v3.caseid_report07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report08':
                        caseid_v3.caseid_report08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report09':
                        caseid_v3.caseid_report09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report10':
                        caseid_v3.caseid_report10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report11':
                        caseid_v3.caseid_report11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report12':
                        caseid_v3.caseid_report12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report13':
                        caseid_v3.caseid_report13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report14':
                        caseid_v3.caseid_report14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report15':
                        caseid_v3.caseid_report15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report16':
                        caseid_v3.caseid_report16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report17':
                        caseid_v3.caseid_report17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report18':
                        caseid_v3.caseid_report18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report19':
                        caseid_v3.caseid_report19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report20':
                        caseid_v3.caseid_report20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report21':
                        caseid_v3.caseid_report21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report22':
                        caseid_v3.caseid_report22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report23':
                        caseid_v3.caseid_report23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report24':
                        caseid_v3.caseid_report24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report25':
                        caseid_v3.caseid_report25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report26':
                        caseid_v3.caseid_report26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report27':
                        caseid_v3.caseid_report27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report28':
                        caseid_v3.caseid_report28(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report29':
                        caseid_v3.caseid_report29(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report30':
                        caseid_v3.caseid_report30(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report31':
                        caseid_v3.caseid_report31(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report32':
                        caseid_v3.caseid_report32(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report33':
                        caseid_v3.caseid_report33(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report34':
                        caseid_v3.caseid_report34(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report35':
                        caseid_v3.caseid_report35(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report36':
                        caseid_v3.caseid_report36(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report37':
                        caseid_v3.caseid_report37(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report38':
                        caseid_v3.caseid_report38(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report39':
                        caseid_v3.caseid_report39(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report40':
                        caseid_v3.caseid_report40(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report41':
                        caseid_v3.caseid_report41(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report42':
                        caseid_v3.caseid_report42(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report43':
                        caseid_v3.caseid_report43(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report44':
                        caseid_v3.caseid_report44(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report45':
                        caseid_v3.caseid_report45(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report46':
                        caseid_v3.caseid_report46(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report47':
                        caseid_v3.caseid_report47(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report48':
                        caseid_v3.caseid_report48(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report49':
                        caseid_v3.caseid_report49(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report50':
                        caseid_v3.caseid_report50(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report51':
                        caseid_v3.caseid_report51(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report52':
                        caseid_v3.caseid_report52(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report53':
                        caseid_v3.caseid_report53(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report54':
                        caseid_v3.caseid_report54(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report55':
                        caseid_v3.caseid_report55(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report56':
                        caseid_v3.caseid_report56(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report57':
                        caseid_v3.caseid_report57(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report58':
                        caseid_v3.caseid_report58(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report59':
                        caseid_v3.caseid_report59(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report60':
                        caseid_v3.caseid_report60(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report61':
                        caseid_v3.caseid_report61(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report62':
                        caseid_v3.caseid_report62(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report63':
                        caseid_v3.caseid_report63(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report64':
                        caseid_v3.caseid_report64(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report65':
                        caseid_v3.caseid_report65(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report66':
                        caseid_v3.caseid_report66(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report67':
                        caseid_v3.caseid_report67(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report68':
                        caseid_v3.caseid_report68(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report69':
                        caseid_v3.caseid_report69(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report70':
                        caseid_v3.caseid_report70(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report71':
                        caseid_v3.caseid_report71(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report72':
                        caseid_v3.caseid_report72(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report73':
                        caseid_v3.caseid_report73(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report74':
                        caseid_v3.caseid_report74(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report75':
                        caseid_v3.caseid_report75(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report76':
                        caseid_v3.caseid_report76(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report77':
                        caseid_v3.caseid_report77(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report78':
                        caseid_v3.caseid_report78(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report79':
                        caseid_v3.caseid_report79(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report80':
                        caseid_v3.caseid_report80(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report81':
                        caseid_v3.caseid_report81(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report82':
                        caseid_v3.caseid_report82(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report83':
                        caseid_v3.caseid_report83(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report84':
                        caseid_v3.caseid_report84(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report85':
                        caseid_v3.caseid_report85(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report86':
                        caseid_v3.caseid_report86(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report87':
                        caseid_v3.caseid_report87(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report88':
                        caseid_v3.caseid_report88(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report89':
                        caseid_v3.caseid_report89(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report90':
                        caseid_v3.caseid_report90(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report91':
                        caseid_v3.caseid_report91(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report92':
                        caseid_v3.caseid_report92(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report93':
                        caseid_v3.caseid_report93(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report94':
                        caseid_v3.caseid_report94(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Report95':
                        caseid_v3.caseid_report95(self)
                except:
                    module_other_v3.swich_tab_0()




#6 video clip
def videoclip(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_v3.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 796
    while (rownum < 839):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print(list_mucdo4)

    modetest = ''.join(re.findall(r'\d+', var_v3.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Video01':
                        caseid_v3.caseid_video01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video02':
                        caseid_v3.caseid_video02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video03':
                        caseid_v3.caseid_video03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video04':
                        caseid_v3.caseid_video04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video05':
                        caseid_v3.caseid_video05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video06':
                        caseid_v3.caseid_video06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video07':
                        caseid_v3.caseid_video07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video08':
                        caseid_v3.caseid_video08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video09':
                        caseid_v3.caseid_video09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video10':
                        caseid_v3.caseid_video10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video11':
                        caseid_v3.caseid_video11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video12':
                        caseid_v3.caseid_video12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video13':
                        caseid_v3.caseid_video13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video14':
                        caseid_v3.caseid_video14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video15':
                        caseid_v3.caseid_video15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video16':
                        caseid_v3.caseid_video16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video17':
                        caseid_v3.caseid_video17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video18':
                        caseid_v3.caseid_video18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video19':
                        caseid_v3.caseid_video19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video20':
                        caseid_v3.caseid_video20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video21':
                        caseid_v3.caseid_video21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video22':
                        caseid_v3.caseid_video22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video23':
                        caseid_v3.caseid_video23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video24':
                        caseid_v3.caseid_video24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video25':
                        caseid_v3.caseid_video25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video26':
                        caseid_v3.caseid_video26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video27':
                        caseid_v3.caseid_video27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video28':
                        caseid_v3.caseid_video28(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Video01':
                        caseid_v3.caseid_video01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video02':
                        caseid_v3.caseid_video02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video03':
                        caseid_v3.caseid_video03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video04':
                        caseid_v3.caseid_video04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video05':
                        caseid_v3.caseid_video05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video06':
                        caseid_v3.caseid_video06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video07':
                        caseid_v3.caseid_video07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video08':
                        caseid_v3.caseid_video08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video09':
                        caseid_v3.caseid_video09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video10':
                        caseid_v3.caseid_video10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video11':
                        caseid_v3.caseid_video11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video12':
                        caseid_v3.caseid_video12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video13':
                        caseid_v3.caseid_video13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video14':
                        caseid_v3.caseid_video14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video15':
                        caseid_v3.caseid_video15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video16':
                        caseid_v3.caseid_video16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video17':
                        caseid_v3.caseid_video17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video18':
                        caseid_v3.caseid_video18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video19':
                        caseid_v3.caseid_video19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video20':
                        caseid_v3.caseid_video20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video21':
                        caseid_v3.caseid_video21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video22':
                        caseid_v3.caseid_video22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video23':
                        caseid_v3.caseid_video23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video24':
                        caseid_v3.caseid_video24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video25':
                        caseid_v3.caseid_video25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video26':
                        caseid_v3.caseid_video26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video27':
                        caseid_v3.caseid_video27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video28':
                        caseid_v3.caseid_video28(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Video01':
                        caseid_v3.caseid_video01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video02':
                        caseid_v3.caseid_video02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video03':
                        caseid_v3.caseid_video03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video04':
                        caseid_v3.caseid_video04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video05':
                        caseid_v3.caseid_video05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video06':
                        caseid_v3.caseid_video06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video07':
                        caseid_v3.caseid_video07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video08':
                        caseid_v3.caseid_video08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video09':
                        caseid_v3.caseid_video09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video10':
                        caseid_v3.caseid_video10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video11':
                        caseid_v3.caseid_video11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video12':
                        caseid_v3.caseid_video12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video13':
                        caseid_v3.caseid_video13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video14':
                        caseid_v3.caseid_video14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video15':
                        caseid_v3.caseid_video15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video16':
                        caseid_v3.caseid_video16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video17':
                        caseid_v3.caseid_video17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video18':
                        caseid_v3.caseid_video18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video19':
                        caseid_v3.caseid_video19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video20':
                        caseid_v3.caseid_video20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video21':
                        caseid_v3.caseid_video21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video22':
                        caseid_v3.caseid_video22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video23':
                        caseid_v3.caseid_video23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video24':
                        caseid_v3.caseid_video24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video25':
                        caseid_v3.caseid_video25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video26':
                        caseid_v3.caseid_video26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video27':
                        caseid_v3.caseid_video27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video28':
                        caseid_v3.caseid_video28(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Video01':
                        caseid_v3.caseid_video01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video02':
                        caseid_v3.caseid_video02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video03':
                        caseid_v3.caseid_video03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video04':
                        caseid_v3.caseid_video04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video05':
                        caseid_v3.caseid_video05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video06':
                        caseid_v3.caseid_video06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video07':
                        caseid_v3.caseid_video07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video08':
                        caseid_v3.caseid_video08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video09':
                        caseid_v3.caseid_video09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video10':
                        caseid_v3.caseid_video10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video11':
                        caseid_v3.caseid_video11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video12':
                        caseid_v3.caseid_video12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video13':
                        caseid_v3.caseid_video13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video14':
                        caseid_v3.caseid_video14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video15':
                        caseid_v3.caseid_video15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video16':
                        caseid_v3.caseid_video16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video17':
                        caseid_v3.caseid_video17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video18':
                        caseid_v3.caseid_video18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video19':
                        caseid_v3.caseid_video19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video20':
                        caseid_v3.caseid_video20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video21':
                        caseid_v3.caseid_video21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video22':
                        caseid_v3.caseid_video22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video23':
                        caseid_v3.caseid_video23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video24':
                        caseid_v3.caseid_video24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video25':
                        caseid_v3.caseid_video25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video26':
                        caseid_v3.caseid_video26(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video27':
                        caseid_v3.caseid_video27(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Video28':
                        caseid_v3.caseid_video28(self)
                except:
                    module_other_v3.swich_tab_0()




#7 hình ảnh
def image(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_v3.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 840
    while (rownum < 874):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print("muc do 1", list_mucdo1)
    print("muc do 2", list_mucdo2)
    print("muc do 3", list_mucdo3)
    print("muc do 4", list_mucdo4)

    modetest = ''.join(re.findall(r'\d+', var_v3.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Image01':
                        caseid_v3.caseid_image01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image02':
                        caseid_v3.caseid_image02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image03':
                        caseid_v3.caseid_image03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image04':
                        caseid_v3.caseid_image04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image05':
                        caseid_v3.caseid_image05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image06':
                        caseid_v3.caseid_image06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image07':
                        caseid_v3.caseid_image07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image08':
                        caseid_v3.caseid_image08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image09':
                        caseid_v3.caseid_image09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image10':
                        caseid_v3.caseid_image10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image11':
                        caseid_v3.caseid_image11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image12':
                        caseid_v3.caseid_image12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image13':
                        caseid_v3.caseid_image13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image14':
                        caseid_v3.caseid_image14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image15':
                        caseid_v3.caseid_image15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image16':
                        caseid_v3.caseid_image16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image17':
                        caseid_v3.caseid_image17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image18':
                        caseid_v3.caseid_image18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image19':
                        caseid_v3.caseid_image19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image20':
                        caseid_v3.caseid_image20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image21':
                        caseid_v3.caseid_image21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image22':
                        caseid_v3.caseid_image22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image23':
                        caseid_v3.caseid_image23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image24':
                        caseid_v3.caseid_image24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image25':
                        caseid_v3.caseid_image25(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Image01':
                        caseid_v3.caseid_image01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image02':
                        caseid_v3.caseid_image02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image03':
                        caseid_v3.caseid_image03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image04':
                        caseid_v3.caseid_image04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image05':
                        caseid_v3.caseid_image05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image06':
                        caseid_v3.caseid_image06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image07':
                        caseid_v3.caseid_image07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image08':
                        caseid_v3.caseid_image08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image09':
                        caseid_v3.caseid_image09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image10':
                        caseid_v3.caseid_image10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image11':
                        caseid_v3.caseid_image11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image12':
                        caseid_v3.caseid_image12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image13':
                        caseid_v3.caseid_image13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image14':
                        caseid_v3.caseid_image14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image15':
                        caseid_v3.caseid_image15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image16':
                        caseid_v3.caseid_image16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image17':
                        caseid_v3.caseid_image17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image18':
                        caseid_v3.caseid_image18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image19':
                        caseid_v3.caseid_image19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image20':
                        caseid_v3.caseid_image20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image21':
                        caseid_v3.caseid_image21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image22':
                        caseid_v3.caseid_image22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image23':
                        caseid_v3.caseid_image23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image24':
                        caseid_v3.caseid_image24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image25':
                        caseid_v3.caseid_image25(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Image01':
                        caseid_v3.caseid_image01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image02':
                        caseid_v3.caseid_image02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image03':
                        caseid_v3.caseid_image03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image04':
                        caseid_v3.caseid_image04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image05':
                        caseid_v3.caseid_image05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image06':
                        caseid_v3.caseid_image06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image07':
                        caseid_v3.caseid_image07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image08':
                        caseid_v3.caseid_image08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image09':
                        caseid_v3.caseid_image09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image10':
                        caseid_v3.caseid_image10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image11':
                        caseid_v3.caseid_image11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image12':
                        caseid_v3.caseid_image12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image13':
                        caseid_v3.caseid_image13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image14':
                        caseid_v3.caseid_image14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image15':
                        caseid_v3.caseid_image15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image16':
                        caseid_v3.caseid_image16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image17':
                        caseid_v3.caseid_image17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image18':
                        caseid_v3.caseid_image18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image19':
                        caseid_v3.caseid_image19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image20':
                        caseid_v3.caseid_image20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image21':
                        caseid_v3.caseid_image21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image22':
                        caseid_v3.caseid_image22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image23':
                        caseid_v3.caseid_image23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image24':
                        caseid_v3.caseid_image24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image25':
                        caseid_v3.caseid_image25(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Image01':
                        caseid_v3.caseid_image01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image02':
                        caseid_v3.caseid_image02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image03':
                        caseid_v3.caseid_image03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image04':
                        caseid_v3.caseid_image04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image05':
                        caseid_v3.caseid_image05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image06':
                        caseid_v3.caseid_image06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image07':
                        caseid_v3.caseid_image07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image08':
                        caseid_v3.caseid_image08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image09':
                        caseid_v3.caseid_image09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image10':
                        caseid_v3.caseid_image10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image11':
                        caseid_v3.caseid_image11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image12':
                        caseid_v3.caseid_image12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image13':
                        caseid_v3.caseid_image13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image14':
                        caseid_v3.caseid_image14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image15':
                        caseid_v3.caseid_image15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image16':
                        caseid_v3.caseid_image16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image17':
                        caseid_v3.caseid_image17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image18':
                        caseid_v3.caseid_image18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image19':
                        caseid_v3.caseid_image19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image20':
                        caseid_v3.caseid_image20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image21':
                        caseid_v3.caseid_image21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image22':
                        caseid_v3.caseid_image22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image23':
                        caseid_v3.caseid_image23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image24':
                        caseid_v3.caseid_image24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Image25':
                        caseid_v3.caseid_image25(self)
                except:
                    module_other_v3.swich_tab_0()




#8 tien ích
def utility(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_v3.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 875
    while (rownum < 911):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)
    print("muc do 1", list_mucdo1)
    print("muc do 2", list_mucdo2)
    print("muc do 3", list_mucdo3)
    print("muc do 4", list_mucdo4)


    modetest = ''.join(re.findall(r'\d+', var_v3.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Utility01':
                        caseid_v3.caseid_utility01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility02':
                        caseid_v3.caseid_utility02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility03':
                        caseid_v3.caseid_utility03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility04':
                        caseid_v3.caseid_utility04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility05':
                        caseid_v3.caseid_utility05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility06':
                        caseid_v3.caseid_utility06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility07':
                        caseid_v3.caseid_utility07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility08':
                        caseid_v3.caseid_utility08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility09':
                        caseid_v3.caseid_utility09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility10':
                        caseid_v3.caseid_utility10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility11':
                        caseid_v3.caseid_utility11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility12':
                        caseid_v3.caseid_utility12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility13':
                        caseid_v3.caseid_utility13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility14':
                        caseid_v3.caseid_utility14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility15':
                        caseid_v3.caseid_utility15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility16':
                        caseid_v3.caseid_utility16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility17':
                        caseid_v3.caseid_utility17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility18':
                        caseid_v3.caseid_utility18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility19':
                        caseid_v3.caseid_utility19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility20':
                        caseid_v3.caseid_utility20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility21':
                        caseid_v3.caseid_utility21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility22':
                        caseid_v3.caseid_utility22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility23':
                        caseid_v3.caseid_utility23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility24':
                        caseid_v3.caseid_utility24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility25':
                        caseid_v3.caseid_utility25(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Utility01':
                        caseid_v3.caseid_utility01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility02':
                        caseid_v3.caseid_utility02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility03':
                        caseid_v3.caseid_utility03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility04':
                        caseid_v3.caseid_utility04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility05':
                        caseid_v3.caseid_utility05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility06':
                        caseid_v3.caseid_utility06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility07':
                        caseid_v3.caseid_utility07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility08':
                        caseid_v3.caseid_utility08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility09':
                        caseid_v3.caseid_utility09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility10':
                        caseid_v3.caseid_utility10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility11':
                        caseid_v3.caseid_utility11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility12':
                        caseid_v3.caseid_utility12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility13':
                        caseid_v3.caseid_utility13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility14':
                        caseid_v3.caseid_utility14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility15':
                        caseid_v3.caseid_utility15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility16':
                        caseid_v3.caseid_utility16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility17':
                        caseid_v3.caseid_utility17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility18':
                        caseid_v3.caseid_utility18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility19':
                        caseid_v3.caseid_utility19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility20':
                        caseid_v3.caseid_utility20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility21':
                        caseid_v3.caseid_utility21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility22':
                        caseid_v3.caseid_utility22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility23':
                        caseid_v3.caseid_utility23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility24':
                        caseid_v3.caseid_utility24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility25':
                        caseid_v3.caseid_utility25(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Utility01':
                        caseid_v3.caseid_utility01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility02':
                        caseid_v3.caseid_utility02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility03':
                        caseid_v3.caseid_utility03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility04':
                        caseid_v3.caseid_utility04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility05':
                        caseid_v3.caseid_utility05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility06':
                        caseid_v3.caseid_utility06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility07':
                        caseid_v3.caseid_utility07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility08':
                        caseid_v3.caseid_utility08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility09':
                        caseid_v3.caseid_utility09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility10':
                        caseid_v3.caseid_utility10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility11':
                        caseid_v3.caseid_utility11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility12':
                        caseid_v3.caseid_utility12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility13':
                        caseid_v3.caseid_utility13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility14':
                        caseid_v3.caseid_utility14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility15':
                        caseid_v3.caseid_utility15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility16':
                        caseid_v3.caseid_utility16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility17':
                        caseid_v3.caseid_utility17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility18':
                        caseid_v3.caseid_utility18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility19':
                        caseid_v3.caseid_utility19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility20':
                        caseid_v3.caseid_utility20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility21':
                        caseid_v3.caseid_utility21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility22':
                        caseid_v3.caseid_utility22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility23':
                        caseid_v3.caseid_utility23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility24':
                        caseid_v3.caseid_utility24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility25':
                        caseid_v3.caseid_utility25(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Utility01':
                        caseid_v3.caseid_utility01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility02':
                        caseid_v3.caseid_utility02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility03':
                        caseid_v3.caseid_utility03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility04':
                        caseid_v3.caseid_utility04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility05':
                        caseid_v3.caseid_utility05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility06':
                        caseid_v3.caseid_utility06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility07':
                        caseid_v3.caseid_utility07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility08':
                        caseid_v3.caseid_utility08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility09':
                        caseid_v3.caseid_utility09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility10':
                        caseid_v3.caseid_utility10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility11':
                        caseid_v3.caseid_utility11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility12':
                        caseid_v3.caseid_utility12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility13':
                        caseid_v3.caseid_utility13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility14':
                        caseid_v3.caseid_utility14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility15':
                        caseid_v3.caseid_utility15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility16':
                        caseid_v3.caseid_utility16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility17':
                        caseid_v3.caseid_utility17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility18':
                        caseid_v3.caseid_utility18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility19':
                        caseid_v3.caseid_utility19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility20':
                        caseid_v3.caseid_utility20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility21':
                        caseid_v3.caseid_utility21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility22':
                        caseid_v3.caseid_utility22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility23':
                        caseid_v3.caseid_utility23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility24':
                        caseid_v3.caseid_utility24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Utility25':
                        caseid_v3.caseid_utility25(self)
                except:
                    module_other_v3.swich_tab_0()



#9 AI
def ai(self):
    list_mucdo1 = []
    list_mucdo2 = []
    list_mucdo3 = []
    list_mucdo4 = []
    wordbook = openpyxl.load_workbook(var_v3.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 912
    while (rownum < 944):
        rownum += 1
        rownum = str(rownum)
        if sheet["H"+rownum].value == "x":
            muc1 = sheet["A"+rownum].value
            list_mucdo1.append(muc1)
        if sheet["I"+rownum].value == "x":
            muc2 = sheet["A"+rownum].value
            list_mucdo2.append(muc2)
        if sheet["J"+rownum].value == "x":
            muc3 = sheet["A"+rownum].value
            list_mucdo3.append(muc3)
        if sheet["K"+rownum].value == "x":
            muc4 = sheet["A"+rownum].value
            list_mucdo4.append(muc4)
        rownum = int(rownum)

    modetest = ''.join(re.findall(r'\d+', var_v3.modetest))
    for i in modetest:
        if i == "1":
            for case in list_mucdo1:
                try:
                    if case == 'Ai01':
                        caseid_v3.caseid_ai01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai02':
                        caseid_v3.caseid_ai02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai03':
                        caseid_v3.caseid_ai03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai04':
                        caseid_v3.caseid_ai04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai05':
                        caseid_v3.caseid_ai05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai06':
                        caseid_v3.caseid_ai06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai07':
                        caseid_v3.caseid_ai07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai08':
                        caseid_v3.caseid_ai08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai09':
                        caseid_v3.caseid_ai09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai10':
                        caseid_v3.caseid_ai10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai11':
                        caseid_v3.caseid_ai11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai12':
                        caseid_v3.caseid_ai12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai13':
                        caseid_v3.caseid_ai13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai14':
                        caseid_v3.caseid_ai14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai15':
                        caseid_v3.caseid_ai15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai16':
                        caseid_v3.caseid_ai16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai17':
                        caseid_v3.caseid_ai17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai18':
                        caseid_v3.caseid_ai18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai19':
                        caseid_v3.caseid_ai19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai20':
                        caseid_v3.caseid_ai20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai21':
                        caseid_v3.caseid_ai21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai22':
                        caseid_v3.caseid_ai22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai23':
                        caseid_v3.caseid_ai23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai24':
                        caseid_v3.caseid_ai24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai25':
                        caseid_v3.caseid_ai25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai26':
                        caseid_v3.caseid_ai26(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "2":
            for case in list_mucdo2:
                try:
                    if case == 'Ai01':
                        caseid_v3.caseid_ai01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai02':
                        caseid_v3.caseid_ai02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai03':
                        caseid_v3.caseid_ai03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai04':
                        caseid_v3.caseid_ai04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai05':
                        caseid_v3.caseid_ai05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai06':
                        caseid_v3.caseid_ai06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai07':
                        caseid_v3.caseid_ai07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai08':
                        caseid_v3.caseid_ai08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai09':
                        caseid_v3.caseid_ai09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai10':
                        caseid_v3.caseid_ai10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai11':
                        caseid_v3.caseid_ai11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai12':
                        caseid_v3.caseid_ai12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai13':
                        caseid_v3.caseid_ai13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai14':
                        caseid_v3.caseid_ai14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai15':
                        caseid_v3.caseid_ai15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai16':
                        caseid_v3.caseid_ai16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai17':
                        caseid_v3.caseid_ai17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai18':
                        caseid_v3.caseid_ai18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai19':
                        caseid_v3.caseid_ai19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai20':
                        caseid_v3.caseid_ai20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai21':
                        caseid_v3.caseid_ai21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai22':
                        caseid_v3.caseid_ai22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai23':
                        caseid_v3.caseid_ai23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai24':
                        caseid_v3.caseid_ai24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai25':
                        caseid_v3.caseid_ai25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai26':
                        caseid_v3.caseid_ai26(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "3":
            for case in list_mucdo3:
                try:
                    if case == 'Ai01':
                        caseid_v3.caseid_ai01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai02':
                        caseid_v3.caseid_ai02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai03':
                        caseid_v3.caseid_ai03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai04':
                        caseid_v3.caseid_ai04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai05':
                        caseid_v3.caseid_ai05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai06':
                        caseid_v3.caseid_ai06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai07':
                        caseid_v3.caseid_ai07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai08':
                        caseid_v3.caseid_ai08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai09':
                        caseid_v3.caseid_ai09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai10':
                        caseid_v3.caseid_ai10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai11':
                        caseid_v3.caseid_ai11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai12':
                        caseid_v3.caseid_ai12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai13':
                        caseid_v3.caseid_ai13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai14':
                        caseid_v3.caseid_ai14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai15':
                        caseid_v3.caseid_ai15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai16':
                        caseid_v3.caseid_ai16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai17':
                        caseid_v3.caseid_ai17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai18':
                        caseid_v3.caseid_ai18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai19':
                        caseid_v3.caseid_ai19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai20':
                        caseid_v3.caseid_ai20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai21':
                        caseid_v3.caseid_ai21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai22':
                        caseid_v3.caseid_ai22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai23':
                        caseid_v3.caseid_ai23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai24':
                        caseid_v3.caseid_ai24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai25':
                        caseid_v3.caseid_ai25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai26':
                        caseid_v3.caseid_ai26(self)
                except:
                    module_other_v3.swich_tab_0()
        if i == "4":
            for case in list_mucdo4:
                try:
                    if case == 'Ai01':
                        caseid_v3.caseid_ai01(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai02':
                        caseid_v3.caseid_ai02(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai03':
                        caseid_v3.caseid_ai03(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai04':
                        caseid_v3.caseid_ai04(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai05':
                        caseid_v3.caseid_ai05(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai06':
                        caseid_v3.caseid_ai06(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai07':
                        caseid_v3.caseid_ai07(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai08':
                        caseid_v3.caseid_ai08(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai09':
                        caseid_v3.caseid_ai09(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai10':
                        caseid_v3.caseid_ai10(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai11':
                        caseid_v3.caseid_ai11(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai12':
                        caseid_v3.caseid_ai12(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai13':
                        caseid_v3.caseid_ai13(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai14':
                        caseid_v3.caseid_ai14(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai15':
                        caseid_v3.caseid_ai15(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai16':
                        caseid_v3.caseid_ai16(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai17':
                        caseid_v3.caseid_ai17(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai18':
                        caseid_v3.caseid_ai18(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai19':
                        caseid_v3.caseid_ai19(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai20':
                        caseid_v3.caseid_ai20(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai21':
                        caseid_v3.caseid_ai21(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai22':
                        caseid_v3.caseid_ai22(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai23':
                        caseid_v3.caseid_ai23(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai24':
                        caseid_v3.caseid_ai24(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai25':
                        caseid_v3.caseid_ai25(self)
                except:
                    module_other_v3.swich_tab_0()
                try:
                    if case == 'Ai26':
                        caseid_v3.caseid_ai26(self)
                except:
                    module_other_v3.swich_tab_0()





















