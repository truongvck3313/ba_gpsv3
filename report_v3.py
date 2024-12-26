import logging
import random
import var_v3
import time
import json
from retry import retry
import administration
from selenium.webdriver.common.by import By
import module_other_v3
import login_v3
import os
import shutil
import re
import openpyxl
from xls2xlsx import XLS2XLSX
from retry import retry
from selenium.webdriver.common.keys import Keys
import minitor_v3
import ai_v3




def write_from_date(fromdate_day_input):
    var_v3.driver.implicitly_wait(1)

    from_date = time.strftime("%d/%m/%Y, ", time.localtime())
    from_date = str(from_date)

    print(from_date)
    from_date_day = from_date[0:2]
    from_date_day = int(from_date_day)
    if from_date_day > 3:
        from_date_day = from_date_day - 3


    if from_date_day == 10:
        from_date_day = "09"
    if from_date_day == 9:
        from_date_day = "08"
    if from_date_day == 8:
        from_date_day = "07"
    if from_date_day == 7:
        from_date_day = "06"
    if from_date_day == 6:
        from_date_day = "05"
    if from_date_day == 5:
        from_date_day = "04"
    if from_date_day == 4:
        from_date_day = "03"
    if from_date_day == 3:
        from_date_day = "02"
    if from_date_day == 2:
        from_date_day = "01"
    if from_date_day == 1:
        from_date_day = "01"
    if from_date_day == 0:
        from_date_day = "01"

    print(from_date_day)
    fromdate_moth_year = from_date[3::]
    fromdate_moth_year = ''.join(re.findall(r'\d+', fromdate_moth_year))

    from_date = str(from_date_day) + fromdate_moth_year
    print(from_date)
    time.sleep(1)
    delete = var_v3.driver.find_element(By.XPATH, fromdate_day_input)
    delete.send_keys(Keys.CONTROL, "a")
    var_v3.driver.find_element(By.XPATH, fromdate_day_input).send_keys(from_date)



def write_from_date_month(fromdate_month_input):
    var_v3.driver.implicitly_wait(1)
    var_v3.driver.find_element(By.XPATH, fromdate_month_input).click()
    time.sleep(1)
    var_v3.driver.find_element(By.XPATH, var_v3.from_day_icon_left).click()
    time.sleep(1.5)

    var_v3.driver.implicitly_wait(0.2)
    n = 0
    while (n < 8):
        n += 1
        n = str(n)
        path_color = "//*[@class='k-content k-calendar-table ng-star-inserted']/tbody/tr[6]/td[" + n + "]"
        try:
            value = var_v3.driver.find_element(By.XPATH, path_color)
            date = value.text
            color = value.value_of_css_property("color")
            print(path_color)
            print(date)
            print(color)
            if color == "rgba(66, 66, 66, 1)":
                value.click()
                time.sleep(1)
                break
        except:
            pass
        n = int(n)

    m = 8
    while (m > 0):
        m -= 1
        m = str(m)
        path_color = "//*[@class='k-content k-calendar-table ng-star-inserted']/tbody/tr[5]/td[" + m + "]"
        try:
            value = var_v3.driver.find_element(By.XPATH, path_color)
            date = value.text
            color = value.value_of_css_property("color")
            print(path_color)
            print(date)
            print(color)
            if color == "rgba(66, 66, 66, 1)":
                value.click()
                time.sleep(1)
                break
        except:
            pass
        m = int(m)




def write_from_date_month1(fromdate_month_input):
    var_v3.driver.implicitly_wait(1)
    var_v3.driver.find_element(By.XPATH, fromdate_month_input).click()
    time.sleep(1)
    var_v3.driver.find_element(By.XPATH, var_v3.from_day_icon_left).click()
    time.sleep(1.5)

    var_v3.driver.implicitly_wait(0.2)
    n = 0
    while (n < 8):
        n += 1
        n = str(n)
        path_color = "//*[@class='k-content k-calendar-table']/tbody/tr[6]/td[" + n + "]"
        try:
            value = var_v3.driver.find_element(By.XPATH, path_color)
            date = value.text
            color = value.value_of_css_property("color")
            print(path_color)
            print(date)
            print(color)
            if color == "rgba(66, 66, 66, 1)":
                value.click()
                time.sleep(1)
                break
        except:
            pass
        n = int(n)

    m = 8
    while (m > 0):
        m -= 1
        m = str(m)
        path_color = "//*[@class='k-content k-calendar-table']/tbody/tr[5]/td[" + n + "]"
        try:
            value = var_v3.driver.find_element(By.XPATH, path_color)
            date = value.text
            color = value.value_of_css_property("color")
            print(path_color)
            print(date)
            print(color)
            if color == "rgba(66, 66, 66, 1)":
                value.click()
                time.sleep(1)
                break
        except:
            pass
        m = int(m)




def clearData_luutamthoi_checkexcel(file, sheetName, column1, column2, column3, column4):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 119
    while (i < 150):
        i += 1
        i = str(i)
        sheet["A"+i] = column1
        sheet["B"+i] = column2
        sheet["C"+i] = column3
        sheet["D"+i] = column4
        i = int(i)
    wordbook.save(file)




def get_info_web():
    var_v3.driver.implicitly_wait(0.05)
    row = 119
    n = 0
    while (n < 50):
        n += 1
        n = str(n)
        row += 1
        path_column = "//*[@aria-label='Data table']/div[1]/div/table/thead/tr/th[" + n + "]"
        path_data = "//*[@class='k-table-tbody']/tr[1]/td[" + n + "]"
        print(n)
        try:
            name_colum = var_v3.driver.find_element(By.XPATH, path_column).text
            name_data = var_v3.driver.find_element(By.XPATH, path_data).text
            print("ten cot web:" .format(name_colum))
            print("data cot web:" .format(name_data))
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", row, 1, name_colum)
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", row, 2, name_data)
        except:
            pass
        n = int(n)



def get_info_excel(row, sheet):
    row2 = row + 1

    try:
        filename = max([var_v3.excelpath + "\\" + f for f in os.listdir(var_v3.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_v3.excelpath, r"baocao_v3.xlsx"))
    except:
        var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
        time.sleep(15)
        filename = max([var_v3.excelpath + "\\" + f for f in os.listdir(var_v3.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_v3.excelpath, r"baocao_v3.xlsx"))

    # #Đọc check file excel
    bangchucai = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

    print("r0")
    try:
        wordbook = openpyxl.load_workbook(var_v3.excelpath + "/baocao_v3.xlsx")
        sheet = wordbook.get_sheet_by_name(sheet)
    except:
        var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
        time.sleep(15)
        filename = max([var_v3.excelpath + "\\" + f for f in os.listdir(var_v3.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_v3.excelpath, r"baocao_v3.xlsx"))
        wordbook = openpyxl.load_workbook(var_v3.excelpath + "/baocao_v3.xlsX")
        sheet = wordbook.get_sheet_by_name(sheet)

    print("r1")
    row_tamthoi = 119
    for i in bangchucai:
        row_tamthoi += 1
        if str(sheet[str(i + str(row))].value) == "None":
            break

        cloumn = str(i + str(row))
        print("vị trí tên cột excel: ".format(cloumn))

        cloumn2 = str(i + str(row2))
        print("vị trí data cột excel: ".format(cloumn2))

        name_column = str(sheet[cloumn].value)
        print("Tên cột excel: ".format(name_column))

        data_column = str(sheet[cloumn2].value)
        print(data_column)
        print("Data cột excel: ".format(data_column))

        var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", row_tamthoi, 3, name_column)
        var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", row_tamthoi, 4, data_column)
        var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", row_tamthoi, 5, cloumn)
        var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", row_tamthoi, 6, cloumn2)



def check_info_web_excel(code, eventname, result, path_module):
    row = 119
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)


    while (row < 150):
        row += 1
        name_column_web = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', row, 1))
        data_column_web = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', row, 2))
        name_column_excel = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', row, 3))
        data_column_excel = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', row, 4))
        location_name_coloumn = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', row, 5))
        location_data_coloumn = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', row, 6))

        if name_column_web == "None":
            break
        logging.info("-------------------------")
        logging.info("Tên cột web: " + name_column_web)
        logging.info("Tên cột excel: " + name_column_excel)
        if name_column_web == name_column_excel:
            logging.info("True")
            module_other_v3.writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")
        else:
            if name_column_web in ['Video', 'Lộ trình', 'Hình ảnh', 'Chi tiết', 'Biểu đồ']:
                pass
            else:
                logging.error("False")
                module_other_v3.writeData(var_v3.checklistpath, "Checklist", code, 6, "Tên cột web: {}\nTên cột excel: {} \nDòng: {}"
                                          .format(name_column_web, name_column_excel, location_name_coloumn))
                module_other_v3.writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
                logging.error("Dòng: {}".format(location_name_coloumn))





        logging.info("Dữ liệu web: " + data_column_web)
        logging.info("Dữ liệu excel: " + data_column_excel)
        if name_column_web in ["STT", "Phương tiện", "Thời điểm bắt đầu", "Thời điểm kết thúc", "Thời gian lăn bánh (hh:mm)", "% Thời gian lăn bánh",
                "Thời gian dừng đỗ (hh:mm)", "Thời gian làm thêm tăng ca (hh:mm)", "Số lần dừng đỗ", "Vận tốc trung bình", "Số lần vi phạm lái xe liên tục", "Số lần quá tốc độ",
                               "Địa chỉ đi", "Địa chỉ đến", "Giờ đi", "Giờ đến", "TG hoạt động (giờ:phút)", "Giấy phép lái xe", "Thời gian đăng nhập", "Địa điểm đăng nhập",
                                "Định mức NL/km", "NL tiêu thụ định mức", "Tên trạm", "Thời điểm vào trạm", "Thời điểm ra trạm", "Thời gian trong trạm (HH:mm:ss)",
                               "Lái xe", "Mã nhân viên", "Điện thoại", "Thời điểm dừng đỗ", "Thời gian dừng đỗ (phút)", "Thời gian dừng đỗ (HH:mm:ss)", "Nổ máy khi dừng (phút)",
                               "Bật điều hòa khi dừng (phút)", "Vị trí dừng", "Định mức nhiên liệu trên 1 km", "Thời gian hoạt động (phút)", "Ngày tháng", "Giờ đến - Giờ đi",
                               "Thời gian hoạt động (phút)", "Nhiên liệu tiêu thụ định mức", "Địa điểm bắt đầu", "Địa điểm kết thúc", "Thời gian (s)", "Vận tốc cực đại (km/h)",
                               "Tốc độ cho phép (km/h)", "Thời điểm bắt đầu bật", "Thời gian bật điều hòa khi dừng đỗ (hh:mm)", "Thời gian đăng xuất", "Địa điểm đăng xuất",
                               "Thời gian di chuyển (hh:mm)", "Thời gian dừng đỗ (hh:mm)", "Ngày", "Số lần đăng nhập", "Số lần đăng xuất", "Tổng số lần", "Số phút",
                               "Thời gian bật máy (giờ:phút)", "Thời gian tắt máy (giờ:phút)", "Thời gian mất tín hiệu (giờ:phút)", "Tổng thời gian", "Số lần bật máy",
                               "Số lần tắt máy", "TG bật máy (hh:mm)", "TG lăn bánh (hh:mm)", "TG dừng đỗ bật máy (hh:mm)", "Số lần đổ", "Số lần hút", "Định mức TT bật máy",
                               "Chênh lệch TT và QĐ", "Hành động", "Địa chỉ", "Thời gian"]:
            print("name vao 1" + name_column_web)
            if data_column_web == data_column_excel:
                logging.info("True")
                module_other_v3.writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.error("False")
                module_other_v3.writeData(var_v3.checklistpath, "Checklist", code, 6, "Dữ liệu web: {}\nDữ liệu excel: {}\n Dòng: {}"
                                          .format(data_column_web, data_column_excel, location_data_coloumn))
                module_other_v3.writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
                logging.error("Dòng: {}".format(location_data_coloumn))


        if name_column_web in ["Km GPS", "Số lít đầu kỳ", "Số lít đổ trong kỳ", "Số lít hút trong kỳ", "Số lít cuối kỳ", "Nhiên liệu tiêu thụ thực tế (lít)", "Định mức thực tế",
                               "Nhiên liệu tiêu thụ", "Quãng đường (m)", "Km cơ", "Vận tốc hành trình (km/h)", "Vận tốc trung bình (km/h)", 'Thời gian dừng đỗ bật điều hòa (hh:mm)',
                               "Tổng thời gian bật điều hòa (hh:mm:ss)", "Từ ngày", "Đến ngày", "Số lít tồn", "Số lít hút", "Số lít đổ", "Số lít tiêu hao", "Định mức QĐ",
                               "Định mức TT km", "Số lít trước", "Số lít sau", "Số lít thay đổi", "NL tiêu thụ"]:
            print("name vao 2" + name_column_web)
            try:
                data_column_web = ''.join(re.findall(r'\d+', data_column_web))[:3]
                data_column_excel = ''.join(re.findall(r'\d+', data_column_excel))[:3]
            except Exception as e:
                logging.error(f"Lỗi khi xử lý dữ liệu: {e}")
            if data_column_web == data_column_excel:
                logging.info("True")
                module_other_v3.writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.error("False")
                module_other_v3.writeData(var_v3.checklistpath, "Checklist", code, 6, "Dữ liệu web: {}\nDữ liệu excel: {}\n Dòng: {}"
                                          .format(data_column_web, data_column_excel, location_data_coloumn))
                module_other_v3.writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
                logging.error("Dòng: {}".format(location_data_coloumn))








class activity_report:

    def summary_report_of_activities(self, code, eventname, result):
        var_v3.driver.implicitly_wait(5)
        login_v3.login.login_v2(self, var_v3.data['login']['conhom_quantri_tk'], var_v3.data['login']['conhom_quantri_mk'])

        administration.hover(var_v3.business_reports)
        administration.hover(var_v3.activity_report)
        var_v3.driver.find_element(By.XPATH, var_v3.summary_report_of_activities).click()
        time.sleep(2.5)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo tổng hợp hoạt động",
                                              var_v3.title_page1, "BÁO CÁO TỔNG HỢP HOẠT ĐỘNG", "_BaoCaoTongHopHoatDong.png")


    def summary_report_of_activities_search(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.time_run)
        except:
            activity_report.summary_report_of_activities(self, "", "", "")

        # write_from_date_month(var_v3.from_day)
        write_from_date(var_v3.from_day)

        time.sleep(0.5)
        var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
        time.sleep(7)
        module_other_v3.write_result_text_try_if_other(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo tổng hợp hoạt động",
                                              var_v3.check_report_data2, "", "_BaoCaoTongHopHoatDong_TimKiem.png")


    def summary_report_of_activities_excel(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.time_run)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            activity_report.summary_report_of_activities_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
        time.sleep(15)
        get_info_web()
        try:
            get_info_excel(5, "BAOCAOTONGHOPHOATDONG")
        except:
            var_v3.driver.refresh()
            time.sleep(7)
            activity_report.summary_report_of_activities_search(self, "", "", "")
            var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
            time.sleep(15)
            get_info_web()
            get_info_excel(5, "BAOCAOTONGHOPHOATDONG")

        check_info_web_excel(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo tổng hợp hoạt động")


    def summary_report_of_activities_pdf(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        try:
            var_v3.driver.refresh()
            time.sleep(7)
        except:
            pass

        activity_report.summary_report_of_activities_search(self, "", "", "")


        var_v3.driver.find_element(By.XPATH, var_v3.export_pdf1).click()
        time.sleep(15)
        module_other_v3.write_result_dowload_file(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo tổng hợp hoạt động",
                                               "baocaotonghophoatdong.pdf", "_BaoCaoTongHopHoatDong_XuatPDF.png")


    def summary_report_of_activities_hide_column(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.time_run)
        except:
            activity_report.summary_report_of_activities(self, "", "", "")


        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column).click()
        except:
            button = var_v3.driver.find_element(By.XPATH, var_v3.hide_column1)
            var_v3.driver.execute_script("arguments[0].click();", button)


        time.sleep(2)
        if var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).is_selected() == False:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).click()
            time.sleep(0.5)
            var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
        else:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(2.3)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo tổng hợp hoạt động",
                                                       var_v3.update_success, "Cập nhật thành công", "_BaoCaoTongHopHoatDong_AnHienCot_Message.png")

        module_other_v3.write_result_displayed_try(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo tổng hợp hoạt động",
                                              var_v3.group_vehicle, "_BaoCaoTongHopHoatDong_AnHienCot.png")


        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column1).click()
            time.sleep(2)
            if var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).is_selected() == True:
                var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).click()
                time.sleep(0.5)
                var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
            else:
                var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        except:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(3)






    def detailed_activity_reports(self, code, eventname, result):
        var_v3.driver.implicitly_wait(5)
        login_v3.login.login_v2(self, var_v3.data['login']['conhom_quantri_tk'], var_v3.data['login']['conhom_quantri_mk'])

        administration.hover(var_v3.business_reports)
        administration.hover(var_v3.activity_report)
        var_v3.driver.find_element(By.XPATH, var_v3.detailed_activity_reports).click()
        time.sleep(2.5)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo chi tiết hoạt động",
                                              var_v3.title_page1, "BÁO CÁO CHI TIẾT HOẠT ĐỘNG", "_BaoCaoChiTietHoatDong.png")


    def detailed_activity_reports_search(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.fuel_1km)
        except:
            activity_report.detailed_activity_reports(self, "", "", "")

        # write_from_date_month(var_v3.from_day)
        write_from_date(var_v3.from_day)

        time.sleep(0.5)
        var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
        time.sleep(7)
        module_other_v3.write_result_text_try_if_other(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo chi tiết hoạt động",
                                              var_v3.check_report_data2, "", "_BaoCaoTongHopHoatDong_TimKiem.png")


    def detailed_activity_reports_excel(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.fuel_1km)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            activity_report.detailed_activity_reports_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
        time.sleep(15)
        get_info_web()
        try:
            get_info_excel(4, "BAOCAOCHITIETHOATDONG")
        except:
            var_v3.driver.refresh()
            time.sleep(7)
            activity_report.detailed_activity_reports_search(self, "", "", "")
            var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
            time.sleep(15)
            get_info_web()
            get_info_excel(4, "BAOCAOCHITIETHOATDONG")
        check_info_web_excel(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo chi tiết hoạt động")


    def detailed_activity_reports_pdf(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        try:
            var_v3.driver.refresh()
            time.sleep(7)
        except:
            pass

        activity_report.detailed_activity_reports_search(self, "", "", "")


        var_v3.driver.find_element(By.XPATH, var_v3.export_pdf1).click()
        time.sleep(15)
        module_other_v3.write_result_dowload_file(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo chi tiết hoạt động",
                                               "baocaochitiethoatdong.pdf", "_BaoCaoChiTietHoatDong_XuatPDF.png")


    def detailed_activity_reports_hide_column(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.fuel_1km)
        except:
            activity_report.detailed_activity_reports(self, "", "", "")


        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column).click()
        except:
            button = var_v3.driver.find_element(By.XPATH, var_v3.hide_column1)
            var_v3.driver.execute_script("arguments[0].click();", button)


        time.sleep(2)
        if var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).is_selected() == False:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).click()
            time.sleep(0.5)
            var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
        else:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(2.3)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo chi tiết hoạt động",
                                                       var_v3.update_success, "Cập nhật thành công", "_BaoCaoChiTietHoatDong_AnHienCot_Message.png")

        module_other_v3.write_result_displayed_try(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo chi tiết hoạt động",
                                              var_v3.group_vehicle, "_BaoCaoChiTietHoatDong_AnHienCot.png")


        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column1).click()
            time.sleep(2)
            if var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).is_selected() == True:
                var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).click()
                time.sleep(0.5)
                var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
            else:
                var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        except:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(3)





    def stop_report(self, code, eventname, result):
        var_v3.driver.implicitly_wait(5)
        login_v3.login.login_v2(self, var_v3.data['login']['conhom_quantri_tk'], var_v3.data['login']['conhom_quantri_mk'])

        administration.hover(var_v3.business_reports)
        administration.hover(var_v3.activity_report)
        var_v3.driver.find_element(By.XPATH, var_v3.stop_report).click()
        time.sleep(2.5)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo dừng đỗ",
                                              var_v3.title_page1, "BÁO CÁO DỪNG ĐỖ", "_BaoCaoDungDo.png")


    def stop_report_search(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.employee_code)
        except:
            activity_report.stop_report(self, "", "", "")

        # write_from_date_month(var_v3.from_day)
        ai_v3.write_from_moth1(var_v3.from_day_b)

        time.sleep(0.5)
        var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
        time.sleep(10)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            var_v3.driver.refresh()
            time.sleep(5)
            var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
            time.sleep(10)
        module_other_v3.write_result_text_try_if_other(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo dừng đỗ",
                                              var_v3.check_report_data2, "", "_BaoCaoDungDo_TimKiem.png")


    def stop_report_excel(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.employee_code)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            activity_report.stop_report_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
        time.sleep(15)
        get_info_web()
        try:
            get_info_excel(5, "BAOCAODUNGDO")
        except:
            var_v3.driver.refresh()
            time.sleep(7)
            activity_report.stop_report_search(self, "", "", "")
            var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
            time.sleep(15)
            get_info_web()
            get_info_excel(5, "BAOCAODUNGDO")

        check_info_web_excel(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo dừng đỗ")


    def stop_report_pdf(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        try:
            var_v3.driver.refresh()
            time.sleep(7)
        except:
            pass

        activity_report.stop_report_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_pdf1).click()
        time.sleep(15)
        module_other_v3.write_result_dowload_file(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo dừng đỗ",
                                               "baocaoddungdo.pdf", "_BaoCaoDungDo_XuatPDF.png")


    def stop_report_hide_column(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.employee_code)
        except:
            activity_report.stop_report(self, "", "", "")


        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column).click()
        except:
            button = var_v3.driver.find_element(By.XPATH, var_v3.hide_column1)
            var_v3.driver.execute_script("arguments[0].click();", button)


        time.sleep(2)
        if var_v3.driver.find_element(By.XPATH, var_v3.hide_field3).is_selected() == False:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_field3).click()
            time.sleep(0.5)
            var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
        else:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(2.3)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo dừng đỗ",
                                                       var_v3.update_success, "Cập nhật thành công", "_BaoCaoDungDo_AnHienCot_Message.png")

        module_other_v3.write_result_displayed_try(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo dừng đỗ",
                                              var_v3.group_vehicle, "_BaoCaoDungDo_AnHienCot.png")


        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column1).click()
            time.sleep(2)
            if var_v3.driver.find_element(By.XPATH, var_v3.hide_field3).is_selected() == True:
                var_v3.driver.find_element(By.XPATH, var_v3.hide_field3).click()
                time.sleep(0.5)
                var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
            else:
                var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        except:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(3)


    def stop_report_icon(self, code, eventname, result, button1, button2, type_check):
        var_v3.driver.implicitly_wait(4)

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.employee_code)
        except:
            activity_report.stop_report(self, "", "", "")


        var_v3.driver.find_element(By.XPATH, button1).click()
        time.sleep(2)
        if type_check == "0":
            module_other_v3.write_result_displayed_try(code, eventname, result,
                                                     "Báo cáo - Báo cáo hoạt động - Báo cáo dừng đỗ",
                                                     var_v3.check_stop_report_icon_see_location, "_BaoCaoDungDo_IconXemViTri.png")
        if type_check == "1":
            module_other_v3.write_result_text_try_if(code, eventname, result,
                                                 "Báo cáo - Báo cáo hoạt động - Báo cáo dừng đỗ",
                                                 var_v3.check_stop_report_icon_see_select_location, "Chọn vị trí hiển thị bản đồ", "_BaoCaoDungDo_IconChonViTri.png")

        var_v3.driver.find_element(By.XPATH, button2).click()
        time.sleep(2)





    def station_report(self, code, eventname, result):
        var_v3.driver.implicitly_wait(5)
        login_v3.login.login_v2(self, var_v3.data['login']['conhom_quantri_tk'], var_v3.data['login']['conhom_quantri_mk'])

        administration.hover(var_v3.business_reports)
        administration.hover(var_v3.activity_report)
        var_v3.driver.find_element(By.XPATH, var_v3.station_report).click()
        time.sleep(2.5)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo ra vào trạm",
                                                 var_v3.title_page1, "BÁO CÁO RA VÀO TRẠM", "_BaoCaoRaVaoTram.png")


    def station_report_search(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.name_station)
        except:
            activity_report.station_report(self, "", "", "")

        # write_from_date_month(var_v3.from_day)
        ai_v3.write_from_moth1(var_v3.from_day_b)

        time.sleep(0.5)
        var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
        time.sleep(10)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            var_v3.driver.refresh()
            time.sleep(5)
            var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
            time.sleep(10)
        module_other_v3.write_result_text_try_if_other(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo ra vào trạm",
                                              var_v3.check_report_data2, "", "__BaoCaoRaVaoTram_TimKiem.png")


    def station_report_excel(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.name_station)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            activity_report.station_report_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
        time.sleep(15)
        get_info_web()
        try:
            get_info_excel(4, "BAOCAORAVAOTRAM")
        except:
            var_v3.driver.refresh()
            time.sleep(7)
            activity_report.station_report_search(self, "", "", "")
            var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
            time.sleep(15)
            get_info_web()
            get_info_excel(4, "BAOCAORAVAOTRAM")

        check_info_web_excel(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo ra vào trạm")


    def station_report_pdf(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        try:
            var_v3.driver.refresh()
            time.sleep(7)
        except:
            pass

        activity_report.station_report_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_pdf1).click()
        time.sleep(15)
        module_other_v3.write_result_dowload_file(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo ra vào trạm",
                                               "baocaoravaotram.pdf", "_BaoCaoRaVaoTram_XuatPDF.png")


    def station_report_hide_column(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.name_station)
        except:
            activity_report.station_report(self, "", "", "")


        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column).click()
        except:
            button = var_v3.driver.find_element(By.XPATH, var_v3.hide_column1)
            var_v3.driver.execute_script("arguments[0].click();", button)


        time.sleep(2)
        if var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).is_selected() == False:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).click()
            time.sleep(0.5)
            var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
        else:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(2.3)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo ra vào trạm",
                                                       var_v3.update_success, "Cập nhật thành công", "_BaoCaoRaVaoTram_AnHienCot_Message.png")

        module_other_v3.write_result_displayed_try(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo ra vào trạm",
                                              var_v3.group_vehicle, "_BaoCaoRaVaoTram_AnHienCot.png")


        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column1).click()
            time.sleep(2)
            if var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).is_selected() == True:
                var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).click()
                time.sleep(0.5)
                var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
            else:
                var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        except:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(3)






    def report_business_trip(self, code, eventname, result):
        var_v3.driver.implicitly_wait(5)
        login_v3.login.login_v2(self, var_v3.data['login']['conhom_quantri_tk'], var_v3.data['login']['conhom_quantri_mk'])

        administration.hover(var_v3.business_reports)
        administration.hover(var_v3.activity_report)
        var_v3.driver.find_element(By.XPATH, var_v3.report_business_trip).click()
        time.sleep(2.5)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo chuyến kinh doanh",
                                                 var_v3.title_page1, "BÁO CÁO CHUYẾN KINH DOANH", "_BaoCaoChuyenKinhDoanh.png")


    def report_business_trip_search(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.report_image)
        except:
            activity_report.report_business_trip(self, "", "", "")

        # write_from_date_month(var_v3.from_day)
        ai_v3.write_from_moth1(var_v3.from_day_b)

        time.sleep(0.5)
        var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
        time.sleep(10)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            var_v3.driver.refresh()
            time.sleep(5)
            var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
            time.sleep(10)
        module_other_v3.write_result_text_try_if_other(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo chuyến kinh doanh",
                                              var_v3.check_report_data2, "", "_BaoCaoChuyenKinhDoanh_TimKiem.png")


    def report_business_trip_excel(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.report_image)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            activity_report.report_business_trip_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
        time.sleep(15)
        get_info_web()
        try:
            get_info_excel(5, "BAOCAOCHUYENKINHDOANH")
        except:
            var_v3.driver.refresh()
            time.sleep(7)
            activity_report.report_business_trip_search(self, "", "", "")
            var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
            time.sleep(15)
            get_info_web()
            get_info_excel(5, "BAOCAOCHUYENKINHDOANH")
        #
        check_info_web_excel(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo chuyến kinh doanh")


    def report_business_trip_pdf(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        try:
            var_v3.driver.refresh()
            time.sleep(7)
        except:
            pass

        activity_report.report_business_trip_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_pdf1).click()
        time.sleep(15)
        module_other_v3.write_result_dowload_file(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo chuyến kinh doanh",
                                               "baocaochuyenkinhdoanh.pdf", "_BaoCaoChuyenKinhDoanh_XuatPDF.png")


    def report_business_trip_hide_column(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.report_image)
        except:
            activity_report.report_business_trip(self, "", "", "")


        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column).click()
        except:
            button = var_v3.driver.find_element(By.XPATH, var_v3.hide_column1)
            var_v3.driver.execute_script("arguments[0].click();", button)


        time.sleep(2)
        if var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).is_selected() == False:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).click()
            time.sleep(0.5)
            var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
        else:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(2.3)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo chuyến kinh doanh",
                                                       var_v3.update_success, "Cập nhật thành công", "_BaoCaoChuyenKinhDoanh_AnHienCot_Message.png")

        module_other_v3.write_result_displayed_try(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo chuyến kinh doanh",
                                              var_v3.group_vehicle, "_BaoCaoChuyenKinhDoanh_AnHienCot.png")


        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column1).click()
            time.sleep(2)
            if var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).is_selected() == True:
                var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).click()
                time.sleep(0.5)
                var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
            else:
                var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        except:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(3)






    def too_speed_report(self, code, eventname, result):
        var_v3.driver.implicitly_wait(5)
        login_v3.login.login_v2(self, var_v3.data['login']['conhom_quantri_tk'], var_v3.data['login']['conhom_quantri_mk'])

        administration.hover(var_v3.business_reports)
        administration.hover(var_v3.activity_report)
        var_v3.driver.find_element(By.XPATH, var_v3.too_speed_report).click()
        time.sleep(2.5)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo quá tốc độ",
                                                 var_v3.title_page1, "BÁO CÁO QUÁ TỐC ĐỘ", "_BaoCaoQuaTocDo.png")


    def too_speed_report_search(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.v_max)
        except:
            activity_report.too_speed_report(self, "", "", "")

        # write_from_date_month(var_v3.from_day)
        ai_v3.write_from_moth1(var_v3.from_day_b)
        time.sleep(0.5)
        var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
        time.sleep(10)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            var_v3.driver.refresh()
            time.sleep(5)
            var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
            time.sleep(10)
        module_other_v3.write_result_text_try_if_other(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo quá tốc độ",
                                              var_v3.check_report_data2, "", "_BaoCaoQuaTocDo_TimKiem.png")


    def too_speed_report_excel(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.v_max)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            activity_report.too_speed_report_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
        time.sleep(15)
        get_info_web()
        try:
            get_info_excel(5, "BAOCAOQUATOCDO")
        except:
            var_v3.driver.refresh()
            time.sleep(7)
            activity_report.too_speed_report_search(self, "", "", "")
            var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
            time.sleep(15)
            get_info_web()
            get_info_excel(5, "BAOCAOQUATOCDO")

        check_info_web_excel(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo quá tốc độ")


    def too_speed_report_pdf(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        # try:
        #     login_v3.login.logout_v2(self)
        # except:
        #     pass

        activity_report.too_speed_report_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_pdf1).click()
        time.sleep(15)
        module_other_v3.write_result_dowload_file(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo quá tốc độ",
                                               "baocaoquatocdo.pdf", "_BaoCaoQuaTocDo_XuatPDF.png")


    def too_speed_report_hide_column(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.v_max)
        except:
            activity_report.too_speed_report(self, "", "", "")


        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column).click()
        except:
            button = var_v3.driver.find_element(By.XPATH, var_v3.hide_column1)
            var_v3.driver.execute_script("arguments[0].click();", button)


        time.sleep(2)
        if var_v3.driver.find_element(By.XPATH, var_v3.hide_field3).is_selected() == False:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_field3).click()
            time.sleep(0.5)
            var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
        else:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(2.3)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo quá tốc độ",
                                                       var_v3.update_success, "Cập nhật thành công", "_BaoCaoQuaTocDo_AnHienCot_Message.png")

        module_other_v3.write_result_displayed_try(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo quá tốc độ",
                                              var_v3.group_vehicle, "_BaoCaoQuaTocDo_AnHienCot.png")


        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column1).click()
            time.sleep(2)
            if var_v3.driver.find_element(By.XPATH, var_v3.hide_field3).is_selected() == True:
                var_v3.driver.find_element(By.XPATH, var_v3.hide_field3).click()
                time.sleep(0.5)
                var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
            else:
                var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        except:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(3)







    def trip_report(self, code, eventname, result):
        var_v3.driver.implicitly_wait(5)
        login_v3.login.login_v2(self, var_v3.data['login']['conhom_quantri_tk'],  var_v3.data['login']['conhom_quantri_mk'])

        administration.hover(var_v3.business_reports)
        administration.hover(var_v3.activity_report)
        var_v3.driver.find_element(By.XPATH, var_v3.trip_report).click()
        time.sleep(2.5)
        module_other_v3.write_result_text_try_if(code, eventname, result,
                                                 "Báo cáo - Báo cáo hoạt động - Báo cáo hành trình",
                                                 var_v3.title_page1, "BÁO CÁO HÀNH TRÌNH", "_BaoCaoHanhTrinh.png")


    def trip_report_search(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.space_2location)
        except:
            activity_report.trip_report(self, "", "", "")

        write_from_date(var_v3.from_day1)
        time.sleep(0.5)

        var_v3.driver.find_element(By.XPATH, var_v3.report_select_vehicle).click()
        time.sleep(1)
        var_v3.driver.find_element(By.XPATH, var_v3.select_list1).click()
        time.sleep(1)
        var_v3.driver.find_element(By.XPATH, var_v3.see_report1).click()
        time.sleep(10)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.icon_check_trip_report)
        except:
            var_v3.driver.find_element(By.XPATH, var_v3.report_select_vehicle).click()
            time.sleep(1)
            var_v3.driver.find_element(By.XPATH, var_v3.select_list2).click()
            time.sleep(1)
            var_v3.driver.find_element(By.XPATH, var_v3.see_report1).click()
            time.sleep(10)

        logging.info("-------------------------")
        logging.info("Báo cáo - Báo cáo hoạt động - Báo cáo hành trình")
        logging.info("Mã - " + code)
        logging.info("Tên sự kiện - " + eventname)
        logging.info("Kết quả - " + result)
        try:
            trip_report_km = var_v3.driver.find_element(By.XPATH, var_v3.trip_report_km).text
            trip_report_stop = var_v3.driver.find_element(By.XPATH, var_v3.trip_report_stop).text
            trip_report_location = var_v3.driver.find_element(By.XPATH, var_v3.trip_report_location).text
            trip_report_time = var_v3.driver.find_element(By.XPATH, var_v3.trip_report_time).text
            trip_report_lost_singal = var_v3.driver.find_element(By.XPATH, var_v3.trip_report_lost_singal).text
            logging.info(trip_report_km)
            logging.info(trip_report_stop)
            logging.info(trip_report_location)
            logging.info(trip_report_time)
            logging.info(trip_report_lost_singal)
            module_other_v3.writeData(var_v3.checklistpath, "Checklist", code, 6, "{}\n{}\n{}\n{}\n{}".format(trip_report_km, trip_report_stop, trip_report_location,
                                                                                                                  trip_report_time, trip_report_lost_singal))

            if trip_report_km and trip_report_stop and trip_report_location and trip_report_time and trip_report_lost_singal != "":
                logging.info("True")
                module_other_v3.writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")
            else:
                logging.info("False")
                var_v3.driver.save_screenshot(var_v3.imagepath + code + "_BaoCaoHanhTrinh_TimKiem.png")
                module_other_v3.writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
                module_other_v3.writeData(var_v3.checklistpath, "Checklist", code, 13, code + "_BaoCaoHanhTrinh_TimKiem.png")
        except:
            logging.info("False")
            var_v3.driver.save_screenshot(var_v3.imagepath + code + "_BaoCaoHanhTrinh_TimKiem.png")
            module_other_v3.writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
            module_other_v3.writeData(var_v3.checklistpath, "Checklist", code, 13, code + "_BaoCaoHanhTrinh_TimKiem.png")


    def trip_report_dowload(self, code, eventname, result, button, file, name_image):
        module_other_v3.delete_excel()
        var_v3.driver.implicitly_wait(4)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.see)
        except:
            activity_report.trip_report_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, button).click()
        time.sleep(15)
        module_other_v3.write_result_dowload_file(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo hành trình",
                                               file, name_image)


    def trip_report_icon_see(self, code, eventname, result):
        module_other_v3.delete_excel()
        var_v3.driver.implicitly_wait(4)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.trip_report_icon_see)
        except:
            activity_report.trip_report_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.trip_report_icon_see).click()
        time.sleep(7)
        var_v3.driver.switch_to.window(var_v3.driver.window_handles[1])

        module_other_v3.write_result_text_try_if(code, eventname, result, "Quản trị - Quản trị phương tiện - Danh sách phương tiện",
                                                       var_v3.check_trip_report_icon_see, "Tổng quan về lộ trình", "_BaoCaoHanhTrinh_IconXem.png")

        time.sleep(1)
        module_other_v3.close_tab()
        time.sleep(1)
        var_v3.driver.switch_to.window(var_v3.driver.window_handles[0])
        time.sleep(1)





    def enable_control_table_report(self, code, eventname, result):
        var_v3.driver.implicitly_wait(5)
        login_v3.login.login_v2(self, var_v3.data['login']['conhom_quantri_tk'], var_v3.data['login']['conhom_quantri_mk'])

        administration.hover(var_v3.business_reports)
        administration.hover(var_v3.activity_report)
        var_v3.driver.find_element(By.XPATH, var_v3.enable_control_table_report).click()
        time.sleep(2.5)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo quá tổng hợp bật điều hòa",
                                                 var_v3.title_page1, "BÁO CÁO TỔNG HỢP BẬT ĐIỀU HÒA", "_BaoCaoTongHopBatDieuHoa.png")


    def enable_control_table_report_search(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.time_on_dh)
        except:
            activity_report.enable_control_table_report(self, "", "", "")

        write_from_date(var_v3.from_day2)

        time.sleep(0.5)
        var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
        time.sleep(10)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            var_v3.driver.refresh()
            time.sleep(5)
            var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
            time.sleep(10)
        module_other_v3.write_result_text_try_if_other(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo quá tổng hợp bật điều hòa",
                                              var_v3.check_report_data2, "", "_BaoCaoTongHopBatDieuHoa_TimKiem.png")


    def enable_control_table_report_excel(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.time_on_dh)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            activity_report.enable_control_table_report_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
        time.sleep(15)
        get_info_web()
        try:
            get_info_excel(4, "BAOCAOTONGHOPBATDIEUHOA")
        except:
            var_v3.driver.refresh()
            time.sleep(7)
            activity_report.too_speed_report_search(self, "", "", "")
            var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
            time.sleep(15)
            get_info_web()
            get_info_excel(4, "BAOCAOTONGHOPBATDIEUHOA")

        check_info_web_excel(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo quá tổng hợp bật điều hòa")


    def enable_control_table_report_pdf(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        # try:
        #     login_v3.login.logout_v2(self)
        # except:
        #     pass
        #
        # activity_report.too_speed_report_search(self, "", "", "")

        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.time_on_dh)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            activity_report.enable_control_table_report_search(self, "", "", "")


        var_v3.driver.find_element(By.XPATH, var_v3.export_pdf1).click()
        time.sleep(15)
        module_other_v3.write_result_dowload_file(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo quá tổng hợp bật điều hòa",
                                               "baocaotonghopbatdieuhoa.pdf", "_BaoCaoTongHopBatDieuHoa_XuatPDF.png")







    def report_on_the_ari_condition_naer(self, code, eventname, result):
        var_v3.driver.implicitly_wait(5)
        login_v3.login.login_v2(self, var_v3.data['login']['conhom_quantri_tk'], var_v3.data['login']['conhom_quantri_mk'])

        administration.hover(var_v3.business_reports)
        administration.hover(var_v3.activity_report)
        var_v3.driver.find_element(By.XPATH, var_v3.report_on_the_ari_condition_naer).click()
        time.sleep(2.5)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo bật điều hòa",
                                                 var_v3.title_page1, "BÁO CÁO BẬT ĐIỀU HÒA", "_BaoCaoBatDieuHoa.png")


    def report_on_the_ari_condition_naer_search(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.time_on_dh_stop)
        except:
            activity_report.report_on_the_ari_condition_naer(self, "", "", "")

        write_from_date(var_v3.from_day2)

        time.sleep(0.5)
        var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
        time.sleep(10)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            var_v3.driver.refresh()
            time.sleep(5)
            var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
            time.sleep(10)
        module_other_v3.write_result_text_try_if_other(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo bật điều hòa",
                                              var_v3.check_report_data2, "", "_BaoCaoBatDieuHoa_TimKiem.png")


    def report_on_the_ari_condition_naer_excel(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.time_on_dh_stop)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            activity_report.report_on_the_ari_condition_naer_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
        time.sleep(15)
        get_info_web()
        try:
            get_info_excel(4, "BAOCAOBATDIEUHOA")
        except:
            var_v3.driver.refresh()
            time.sleep(7)
            activity_report.report_on_the_ari_condition_naer_search(self, "", "", "")
            var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
            time.sleep(15)
            get_info_web()
            get_info_excel(4, "BAOCAOBATDIEUHOA")

        check_info_web_excel(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo bật điều hòa")


    def report_on_the_ari_condition_naer_pdf(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()

        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.time_on_dh_stop)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            activity_report.report_on_the_ari_condition_naer_search(self, "", "", "")


        var_v3.driver.find_element(By.XPATH, var_v3.export_pdf1).click()
        time.sleep(15)
        module_other_v3.write_result_dowload_file(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo quá bật điều hòa",
                                               "baocaobatdieuhoa.pdf", "_BaoCaoBatDieuHoa_XuatPDF.png")


    def report_on_the_ari_condition_naer_hide_column(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.time_on_dh_stop)
        except:
            activity_report.report_on_the_ari_condition_naer(self, "", "", "")


        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column).click()
        except:
            button = var_v3.driver.find_element(By.XPATH, var_v3.hide_column1)
            var_v3.driver.execute_script("arguments[0].click();", button)


        time.sleep(2)
        if var_v3.driver.find_element(By.XPATH, var_v3.hide_field3).is_selected() == False:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_field3).click()
            time.sleep(0.5)
            var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
        else:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(2.3)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo quá bật điều hòa",
                                                       var_v3.update_success, "Cập nhật thành công", "_BaoCaoBatDieuHoa_AnHienCot_Message.png")

        module_other_v3.write_result_displayed_try(code, eventname, result, "Báo cáo - Báo cáo hoạt động - Báo cáo quá bật điều hòa",
                                              var_v3.group_vehicle, "_BaoCaoBatDieuHoa_AnHienCot.png")

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column1).click()
            time.sleep(2)
            if var_v3.driver.find_element(By.XPATH, var_v3.hide_field3).is_selected() == True:
                var_v3.driver.find_element(By.XPATH, var_v3.hide_field3).click()
                time.sleep(0.5)
                var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
            else:
                var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        except:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(3)






class summary_report:



    def report_detail_checkin_checkout(self, code, eventname, result):
        var_v3.driver.implicitly_wait(5)
        login_v3.login.login_v2(self, var_v3.data['login']['conhom_quantri_tk'], var_v3.data['login']['conhom_quantri_mk'])

        administration.hover(var_v3.business_reports)
        administration.hover(var_v3.summary_report)
        var_v3.driver.find_element(By.XPATH, var_v3.report_detail_checkin_checkout).click()
        time.sleep(2.5)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo tổng hợp - Báo cáo chi tiết lái xe đăng nhập đăng xuất",
                                              var_v3.title_page1, "BÁO CÁO CHI TIẾT LÁI XE ĐĂNG NHẬP ĐĂNG XUẤT", "_BaoCaoChiTietLaiXeDangNhapDangXuat.png")


    def report_detail_checkin_checkout_search(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.time_login)
        except:
            summary_report.report_detail_checkin_checkout(self, "", "", "")

        # write_from_date_month(var_v3.from_day)
        write_from_date(var_v3.from_day2)

        time.sleep(0.5)
        var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
        time.sleep(7)
        module_other_v3.write_result_text_try_if_other(code, eventname, result, "Báo cáo - Báo cáo tổng hợp - Báo cáo chi tiết lái xe đăng nhập đăng xuất",
                                              var_v3.check_report_data2, "", "_BaoCaoChiTietLaiXeDangNhapDangXuat_TimKiem.png")


    def report_detail_checkin_checkout_excel(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.time_login)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            summary_report.report_detail_checkin_checkout_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
        time.sleep(15)
        get_info_web()
        try:
            get_info_excel(4, "BAOCAOCHITIETLAIXEDANGNHAPDANGX")
        except:
            var_v3.driver.refresh()
            time.sleep(7)
            summary_report.report_detail_checkin_checkout_search(self, "", "", "")
            var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
            time.sleep(15)
            get_info_web()
            get_info_excel(4, "BAOCAOCHITIETLAIXEDANGNHAPDANGX")

        check_info_web_excel(code, eventname, result, "Báo cáo - Báo cáo tổng hợp - Báo cáo chi tiết lái xe đăng nhập đăng xuất")


    def report_detail_checkin_checkout_pdf(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()

        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.time_login)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            summary_report.report_detail_checkin_checkout_search(self, "", "", "")


        var_v3.driver.find_element(By.XPATH, var_v3.export_pdf1).click()
        time.sleep(15)
        module_other_v3.write_result_dowload_file(code, eventname, result, "Báo cáo - Báo cáo tổng hợp - Báo cáo chi tiết lái xe đăng nhập đăng xuất",
                                               "baocaochitietlaixedangnhapdangxuat.pdf", "_BaoCaoChiTietLaiXeDangNhapDangXuat_XuatPDF.png")






    def report_summary_checkin_checkout(self, code, eventname, result):
        var_v3.driver.implicitly_wait(5)
        login_v3.login.login_v2(self, var_v3.data['login']['conhom_quantri_tk'], var_v3.data['login']['conhom_quantri_mk'])

        administration.hover(var_v3.business_reports)
        administration.hover(var_v3.summary_report)
        var_v3.driver.find_element(By.XPATH, var_v3.report_summary_checkin_checkout).click()
        time.sleep(2.5)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo tổng hợp - Báo cáo tổng hợp lái xe đăng nhập đăng xuất",
                                              var_v3.title_page1, "BÁO CÁO TỔNG HỢP LÁI XE ĐĂNG NHẬP ĐĂNG XUẤT", "_BaoCaoTongHopLaiXeDangNhapDangXuat.png")


    def report_summary_checkin_checkout_search(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.summary_time)
        except:
            summary_report.report_summary_checkin_checkout(self, "", "", "")

        # write_from_date_month(var_v3.from_day)
        write_from_date(var_v3.from_day2)

        time.sleep(0.5)
        var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
        time.sleep(7)
        module_other_v3.write_result_text_try_if_other(code, eventname, result, "Báo cáo - Báo cáo tổng hợp - Báo cáo tổng hợp lái xe đăng nhập đăng xuất",
                                              var_v3.check_report_data2, "", "_BaoCaoTongHopLaiXeDangNhapDangXuat_TimKiem.png")


    def report_summary_checkin_checkout_excel(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.summary_time)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            summary_report.report_summary_checkin_checkout_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
        time.sleep(15)
        get_info_web()
        try:
            get_info_excel(5, "BAOCAOTONGHOPLAIXEDANGNHAPDANGX")
        except:
            var_v3.driver.refresh()
            time.sleep(7)
            summary_report.report_summary_checkin_checkout_search(self, "", "", "")
            var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
            time.sleep(15)
            get_info_web()
            get_info_excel(5, "BAOCAOTONGHOPLAIXEDANGNHAPDANGX")

        check_info_web_excel(code, eventname, result, "Báo cáo - Báo cáo tổng hợp - Báo cáo tổng hợp lái xe đăng nhập đăng xuất")


    def report_summary_checkin_checkout_pdf(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()

        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.summary_time)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            summary_report.report_summary_checkin_checkout_search(self, "", "", "")


        var_v3.driver.find_element(By.XPATH, var_v3.export_pdf1).click()
        time.sleep(15)
        module_other_v3.write_result_dowload_file(code, eventname, result, "Báo cáo - Báo cáo tổng hợp - Báo cáo tổng hợp lái xe đăng nhập đăng xuất",
                                               "baocaotonghoplaixedangnhapdangxuat.pdf", "_BaoCaoTongHopLaiXeDangNhapDangXuat_XuatPDF.png")





class engine_report:


    def engine_report(self, code, eventname, result):
        var_v3.driver.implicitly_wait(5)
        login_v3.login.login_v2(self, var_v3.data['login']['conhom_quantri_tk'], var_v3.data['login']['conhom_quantri_mk'])

        administration.hover(var_v3.business_reports)
        administration.hover(var_v3.engine_report)
        var_v3.driver.find_element(By.XPATH, var_v3.engine_report1).click()
        time.sleep(2.5)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo động cơ - Báo cáo động cơ",
                                              var_v3.title_page1, "BÁO CÁO ĐỘNG CƠ", "_BaoCaoDongCo.png")


    def engine_report_search(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.number_minute)
        except:
            engine_report.engine_report(self, "", "", "")

        # write_from_date_month(var_v3.from_day3)
        ai_v3.write_from_moth1(var_v3.from_day_b)
        time.sleep(0.5)
        var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
        time.sleep(10)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            var_v3.driver.refresh()
            time.sleep(5)
            var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
            time.sleep(10)
        module_other_v3.write_result_text_try_if_other(code, eventname, result, "Báo cáo - Báo cáo động cơ - Báo cáo động cơ",
                                              var_v3.check_report_data2, "", "_BaoCaoDongCo_TimKiem.png")


    def engine_report_excel(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.number_minute)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            engine_report.engine_report_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
        time.sleep(15)
        get_info_web()
        try:
            get_info_excel(5, "BAOCAODONGCO")
        except:
            var_v3.driver.refresh()
            time.sleep(7)
            engine_report.engine_report_search(self, "", "", "")
            var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
            time.sleep(15)
            get_info_web()
            get_info_excel(5, "BAOCAODONGCO")

        check_info_web_excel(code, eventname, result, "Báo cáo - Báo cáo động cơ - Báo cáo động cơ")


    def engine_report_pdf(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        # try:
        #     login_v3.login.logout_v2(self)
        # except:
        #     pass

        engine_report.engine_report_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_pdf1).click()
        time.sleep(15)
        module_other_v3.write_result_dowload_file(code, eventname, result, "Báo cáo - Báo cáo động cơ - Báo cáo động cơ",
                                               "baocaodongco.pdf", "_BaoCaoDongCo_XuatPDF.png")


    def engine_report_hide_column(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.engine_report)
        except:
            engine_report.engine_report(self, "", "", "")


        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column).click()
        except:
            button = var_v3.driver.find_element(By.XPATH, var_v3.hide_column1)
            var_v3.driver.execute_script("arguments[0].click();", button)


        time.sleep(2)
        if var_v3.driver.find_element(By.XPATH, var_v3.hide_field3).is_selected() == False:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_field3).click()
            time.sleep(0.5)
            var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
        else:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(2.3)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo động cơ - Báo cáo động cơ",
                                                       var_v3.update_success, "Cập nhật thành công", "_BaoCaoDongCo_AnHienCot_Message.png")

        module_other_v3.write_result_displayed_try(code, eventname, result, "Báo cáo - Báo cáo động cơ - Báo cáo động cơ",
                                              var_v3.group_vehicle, "_BaoCaoDongCo_AnHienCot.png")

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column1).click()
            time.sleep(2)
            if var_v3.driver.find_element(By.XPATH, var_v3.hide_field3).is_selected() == True:
                var_v3.driver.find_element(By.XPATH, var_v3.hide_field3).click()
                time.sleep(0.5)
                var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
            else:
                var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        except:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(3)





    def report_status_engine(self, code, eventname, result):
        var_v3.driver.implicitly_wait(5)
        login_v3.login.login_v2(self, var_v3.data['login']['conhom_quantri_tk'], var_v3.data['login']['conhom_quantri_mk'])

        administration.hover(var_v3.business_reports)
        administration.hover(var_v3.engine_report)
        var_v3.driver.find_element(By.XPATH, var_v3.report_status_engine).click()
        time.sleep(2.5)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo động cơ - Báo cáo trạng thái động cơ",
                                              var_v3.title_page1, "BÁO CÁO TRẠNG THÁI ĐỘNG CƠ", "_BaoCaoTrangThaiDongCo.png")


    def report_status_engine_search(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.day)
        except:
            engine_report.report_status_engine(self, "", "", "")

        # write_from_date_month(var_v3.from_day4)
        ai_v3.write_from_moth1(var_v3.from_day_b)
        time.sleep(0.5)
        var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
        time.sleep(10)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            var_v3.driver.refresh()
            time.sleep(5)
            var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
            time.sleep(10)
        module_other_v3.write_result_text_try_if_other(code, eventname, result, "Báo cáo - Báo cáo động cơ - Báo cáo trạng thái động cơ",
                                              var_v3.check_report_data2, "", "_BaoCaoTrangThaiDongCo_TimKiem.png")


    def report_status_engine_excel(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.day)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            engine_report.report_status_engine_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
        time.sleep(15)
        get_info_web()
        try:
            get_info_excel(4, "BAOCAOTRANGTHAIDONGCO")
        except:
            var_v3.driver.refresh()
            time.sleep(7)
            engine_report.report_status_engine_search(self, "", "", "")
            var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
            time.sleep(15)
            get_info_web()
            get_info_excel(4, "BAOCAOTRANGTHAIDONGCO")

        check_info_web_excel(code, eventname, result, "Báo cáo - Báo cáo động cơ - Báo cáo trạng thái động cơ")


    def report_status_engine_pdf(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        try:
            login_v3.login.logout_v2(self)
        except:
            pass

        engine_report.report_status_engine_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_pdf1).click()
        time.sleep(15)
        module_other_v3.write_result_dowload_file(code, eventname, result, "Báo cáo - Báo cáo động cơ - Báo cáo trạng thái động cơ",
                                               "baocaotrangthaidongco.pdf", "_BaoCaoTrangThaiDongCo_XuatPDF.png")


    def report_status_engine_hide_column(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.day)
        except:
            engine_report.report_status_engine(self, "", "", "")


        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column).click()
        except:
            button = var_v3.driver.find_element(By.XPATH, var_v3.hide_column1)
            var_v3.driver.execute_script("arguments[0].click();", button)


        time.sleep(2)
        if var_v3.driver.find_element(By.XPATH, var_v3.hide_field1).is_selected() == True:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_field1).click()
            time.sleep(0.5)
            var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
        else:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(2.3)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo động cơ - Báo cáo trạng thái động cơ",
                                                       var_v3.update_success, "Cập nhật thành công", "_BaoCaoTrangThaiDongCo_AnHienCot_Message.png")

        module_other_v3.write_result_not_displayed_try(code, eventname, result, "Báo cáo - Báo cáo động cơ - Báo cáo trạng thái động cơ",
                                              var_v3.day, "_BaoCaoTrangThaiDongCo_AnHienCot.png")


        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column1).click()
            time.sleep(2)
            if var_v3.driver.find_element(By.XPATH, var_v3.hide_field1).is_selected() == True:
                var_v3.driver.find_element(By.XPATH, var_v3.hide_field1).click()
                time.sleep(0.5)
                var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
            else:
                var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        except:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(3)





class fuel_report:


    def fuel_consumtion_report(self, code, eventname, result):
        var_v3.driver.implicitly_wait(5)
        login_v3.login.login_v2(self, var_v3.data['login']['conhom_quantri_tk'], var_v3.data['login']['conhom_quantri_mk'])

        administration.hover(var_v3.business_reports)
        administration.hover(var_v3.fuel_report)
        var_v3.driver.find_element(By.XPATH, var_v3.fuel_consumtion_report).click()
        time.sleep(2.5)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo nhiên liệu - Báo cáo tiêu hao nhiên liệu",
                                              var_v3.title_page1, "BÁO CÁO TIÊU HAO NHIÊN LIỆU", "_BaoCaoTieuHaoNhienLieu.png")


    def fuel_consumtion_report_search(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.dinhmuc_qd)
        except:
            fuel_report.fuel_consumtion_report(self, "", "", "")

        # write_from_date_month(var_v3.from_day)
        write_from_date(var_v3.from_day2)

        time.sleep(0.5)
        var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
        time.sleep(10)
        module_other_v3.write_result_text_try_if_other(code, eventname, result, "Báo cáo - Báo cáo nhiên liệu - Báo cáo tiêu hao nhiên liệu",
                                              var_v3.check_report_data2, "", "_BaoCaoTieuHaoNhienLieu_TimKiem.png")


    def fuel_consumtion_report_excel(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.dinhmuc_qd)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            fuel_report.fuel_consumtion_report(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
        time.sleep(15)
        get_info_web()
        try:
            get_info_excel(5, "BAOCAOTIEUHAONHIENLIEU")
        except:
            var_v3.driver.refresh()
            time.sleep(7)
            fuel_report.fuel_consumtion_report_search(self, "", "", "")
            var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
            time.sleep(15)
            get_info_web()
            get_info_excel(5, "BAOCAOTIEUHAONHIENLIEU")

        check_info_web_excel(code, eventname, result, "Báo cáo - Báo cáo nhiên liệu - Báo cáo tiêu hao nhiên liệu")


    def report_detail_checkin_checkout_pdf(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()

        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.dinhmuc_qd)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            fuel_report.fuel_consumtion_report_search(self, "", "", "")


        var_v3.driver.find_element(By.XPATH, var_v3.export_pdf1).click()
        time.sleep(15)
        module_other_v3.write_result_dowload_file(code, eventname, result, "Báo cáo - Báo cáo nhiên liệu - Báo cáo tiêu hao nhiên liệu",
                                               "baocaochitieuhaonhienlieu.pdf", "_BaoCaoTieuHaoNhienLieu_XuatPDF.png")


    def summary_report_of_activities_hide_column(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.dinhmuc_qd)
        except:
            fuel_report.fuel_consumtion_report(self, "", "", "")


        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column).click()
        except:
            button = var_v3.driver.find_element(By.XPATH, var_v3.hide_column1)
            var_v3.driver.execute_script("arguments[0].click();", button)


        time.sleep(2)
        if var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).is_selected() == False:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).click()
            time.sleep(0.5)
            var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
        else:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(2.3)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo nhiên liệu - Báo cáo tiêu hao nhiên liệu",
                                                       var_v3.update_success, "Cập nhật thành công", "_BaoCaoTieuHaoNhienLieu_AnHienCot_Message.png")

        module_other_v3.write_result_displayed_try(code, eventname, result, "Báo cáo - Báo cáo nhiên liệu - Báo cáo tiêu hao nhiên liệu",
                                              var_v3.group_vehicle, "_BaoCaoTieuHaoNhienLieu_AnHienCot.png")

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column1).click()
            time.sleep(2)
            if var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).is_selected() == True:
                var_v3.driver.find_element(By.XPATH, var_v3.hide_field4).click()
                time.sleep(0.5)
                var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
            else:
                var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        except:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(3)









    def report_fuel_spillage(self, code, eventname, result):
        var_v3.driver.implicitly_wait(5)
        login_v3.login.login_v2(self, var_v3.data['login']['conhom_quantri_tk'], var_v3.data['login']['conhom_quantri_mk'])

        administration.hover(var_v3.business_reports)
        administration.hover(var_v3.fuel_report)
        var_v3.driver.find_element(By.XPATH, var_v3.report_fuel_spillage).click()
        time.sleep(2.5)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo nhiên liệu - Báo cáo đổ hút nhiên liệu",
                                              var_v3.title_page1, "BÁO CÁO ĐỔ HÚT NHIÊN LIỆU", "_BaoCaoDoHutNhienLieu.png")


    def report_fuel_spillage_search(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.fuel_change)
        except:
            fuel_report.report_fuel_spillage(self, "", "", "")

        # write_from_date_month(var_v3.from_day)
        write_from_date(var_v3.from_day2)

        time.sleep(0.5)
        var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
        time.sleep(10)
        module_other_v3.write_result_text_try_if_other(code, eventname, result, "Báo cáo - Báo cáo nhiên liệu - Báo cáo đổ hút nhiên liệu",
                                              var_v3.check_report_data2, "", "_BaoCaoDoHutNhienLieu_TimKiem.png")


    def report_fuel_spillage_excel(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()
        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.fuel_change)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            fuel_report.report_fuel_spillage_search(self, "", "", "")

        var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
        time.sleep(15)
        get_info_web()
        try:
            get_info_excel(4, "BAOCAODOHUTNHIENLIEU")
        except:
            var_v3.driver.refresh()
            time.sleep(7)
            fuel_report.report_fuel_spillage_search(self, "", "", "")
            var_v3.driver.find_element(By.XPATH, var_v3.export_excel).click()
            time.sleep(15)
            get_info_web()
            get_info_excel(4, "BAOCAODOHUTNHIENLIEU")

        check_info_web_excel(code, eventname, result, "Báo cáo - Báo cáo nhiên liệu - Báo cáo đổ hút nhiên liệu")


    def report_fuel_spillage_pdf(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        module_other_v3.delete_excel()

        clearData_luutamthoi_checkexcel(var_v3.path_luutamthoi, "Sheet1", "", "", "", "")
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.fuel_change)
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            fuel_report.report_fuel_spillage_search(self, "", "", "")


        var_v3.driver.find_element(By.XPATH, var_v3.export_pdf1).click()
        time.sleep(15)
        module_other_v3.write_result_dowload_file(code, eventname, result, "Báo cáo - Báo cáo nhiên liệu - Báo cáo đổ hút nhiên liệu",
                                               "baocaodohutnhienlieu.pdf", "_BaoCaoDoHutNhienLieu_XuatPDF.png")


    def report_fuel_spillage_hide_column(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.fuel_change)
        except:
            fuel_report.report_fuel_spillage(self, "", "", "")


        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column).click()
        except:
            button = var_v3.driver.find_element(By.XPATH, var_v3.hide_column1)
            var_v3.driver.execute_script("arguments[0].click();", button)


        time.sleep(2)
        if var_v3.driver.find_element(By.XPATH, var_v3.hide_field1).is_selected() == False:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_field1).click()
            time.sleep(0.5)
            var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
        else:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(2.3)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo - Báo cáo nhiên liệu - Báo cáo đổ hút nhiên liệu",
                                                       var_v3.update_success, "Cập nhật thành công", "_BaoCaoDoHutNhienLieu_AnHienCot_Message.png")

        module_other_v3.write_result_displayed_try(code, eventname, result, "Báo cáo - Báo cáo nhiên liệu - Báo cáo đổ hút nhiên liệu",
                                              var_v3.liscense_plate, "_BaoCaoDoHutNhienLieu_AnHienCot.png")

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column1).click()
            time.sleep(2)
            if var_v3.driver.find_element(By.XPATH, var_v3.hide_field1).is_selected() == True:
                var_v3.driver.find_element(By.XPATH, var_v3.hide_field1).click()
                time.sleep(0.5)
                var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
            else:
                var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        except:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(3)




class report_bgt:


    def over_speed_detail_report(self, code, eventname, result):
        var_v3.driver.implicitly_wait(5)
        minitor_v3.goto(self, "Mã XN", "950", " Công ty: Viconship Đà Nẵng - 950 - viconshipdanang1 ")

        administration.hover(var_v3.report_bgt)
        administration.hover(var_v3.tt09_2015)
        var_v3.driver.find_element(By.XPATH, var_v3.over_speed_detail_report).click()
        time.sleep(2.5)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo BGT - Báo cáo TT09/2015/TT - BGTVT -  Chi tiết vi phạm tốc độ xe chạy",
                                              var_v3.title_page, "CHI TIẾT VI PHẠM TỐC ĐỘ XE CHẠY", "_ChiTietViPhamTocDoXeChay.png")


    def over_speed_detail_report_search(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.violation_time)
        except:
            report_bgt.over_speed_detail_report(self, "", "", "")

        # write_from_date_month1(var_v3.from_day3)
        ai_v3.write_from_moth1(var_v3.from_day_b)

        time.sleep(0.5)
        var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
        time.sleep(10)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
        except:
            var_v3.driver.refresh()
            time.sleep(5)
            var_v3.driver.find_element(By.XPATH, var_v3.see_report).click()
            time.sleep(10)
        module_other_v3.write_result_text_try_if_other(code, eventname, result, "Báo cáo BGT - Báo cáo TT09/2015/TT - BGTVT -  Chi tiết vi phạm tốc độ xe chạy",
                                              var_v3.check_report_data2, "", "_ChiTietViPhamTocDoXeChay_TimKiem.png")


    def over_speed_detail_report_dowload(self, code, eventname, result, button, file, name_image):
        module_other_v3.delete_excel()
        var_v3.driver.implicitly_wait(4)
        try:
            var_v3.driver.find_element(By.XPATH, var_v3.check_report_data2)
            var_v3.driver.find_element(By.XPATH, var_v3.violation_time)
        except:
            report_bgt.over_speed_detail_report_search(self, "", "", "")


        var_v3.driver.find_element(By.XPATH, button).click()
        time.sleep(15)
        module_other_v3.write_result_dowload_file(code, eventname, result, "Báo cáo BGT - Báo cáo TT09/2015/TT - BGTVT -  Chi tiết vi phạm tốc độ xe chạy",
                                               file, name_image)


    def over_speed_detail_report_hide_column(self, code, eventname, result):
        var_v3.driver.implicitly_wait(4)

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.violation_time)
        except:
            report_bgt.over_speed_detail_report(self, "", "", "")


        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column).click()
        except:
            button = var_v3.driver.find_element(By.XPATH, var_v3.hide_column1)
            var_v3.driver.execute_script("arguments[0].click();", button)


        time.sleep(2)
        if var_v3.driver.find_element(By.XPATH, var_v3.hide_field1).is_selected() == False:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_field1).click()
            time.sleep(0.5)
            var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
        else:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(2.3)
        module_other_v3.write_result_text_try_if(code, eventname, result, "Báo cáo BGT - Báo cáo TT09/2015/TT - BGTVT -  Chi tiết vi phạm tốc độ xe chạy",
                                                       var_v3.update_success, "Cập nhật thành công", "_ChiTietViPhamTocDoXeChay_AnHienCot_Message.png")

        module_other_v3.write_result_displayed_try(code, eventname, result, "Báo cáo BGT - Báo cáo TT09/2015/TT - BGTVT -  Chi tiết vi phạm tốc độ xe chạy",
                                              var_v3.liscense_plate, "_ChiTietViPhamTocDoXeChay_AnHienCot.png")

        try:
            var_v3.driver.find_element(By.XPATH, var_v3.hide_column1).click()
            time.sleep(2)
            if var_v3.driver.find_element(By.XPATH, var_v3.hide_field1).is_selected() == True:
                var_v3.driver.find_element(By.XPATH, var_v3.hide_field1).click()
                time.sleep(0.5)
                var_v3.driver.find_element(By.XPATH, var_v3.save1).click()
            else:
                var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        except:
            var_v3.driver.find_element(By.XPATH, var_v3.close2).click()
        time.sleep(3)
