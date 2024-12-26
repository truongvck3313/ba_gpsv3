import logging
from selenium.common.exceptions import NoSuchElementException
import var_v3
import time
import openpyxl
import subprocess
from selenium.webdriver.common.keys import Keys
import os
from selenium.webdriver.common.by import By
from retry import retry
import os
import shutil
import module_gpsv3










def timerun():
    while True:
        time.sleep(3)
        timerun = time.strftime("%H:%M:%S", time.localtime())
        print("Thời gian hiện tại:", timerun)
        print("Thời gian chạy tool:", var_v3.timerun)
        var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 47, 2, timerun)
        if var_v3.timerun == "":
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 47, 2, timerun)
        else:
            var_v3.writeData(var_v3.path_luutamthoi, "Sheet1", 47, 2, var_v3.timerun)


        if var_v3.timerun == time.strftime("%H:%M", time.localtime()):
            break
        if var_v3.timerun == "":
            break



def clear_log():
    logging.basicConfig(handlers=[logging.FileHandler(filename=var_v3.logpath,
                                                      encoding='utf-8', mode='w')],  # mode='a+', w
                        format="%(asctime)s %(name)s:%(levelname)s:%(message)s",
                        datefmt="%F %A %T",
                        level=logging.INFO)




def clearData(file, sheetName, ketqua, trangthai, tenanh):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 9
    while (i < 1000):
        i += 1
        i = str(i)
        sheet["F"+i] = ketqua
        sheet["G"+i] = trangthai
        sheet["M"+i] = tenanh
        i = int(i)
    wordbook.save(file)






def clearData_luutamthoi(file,sheetName, web, api, popup):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 6
    while (i < 100):
        i += 1
        i = str(i)
        sheet["B"+i] = web
        sheet["C"+i] = api
        sheet["D"+i] = popup
        i = int(i)
    wordbook.save(file)




def clearData_luutamthoi1(file,sheetName, column1, column2, column3, column4):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 119
    while (i < 150):
        i += 1
        i = str(i)
        sheet["A"+i] = column1
        sheet["B"+i] = column2
        sheet["C"+i] = column3
        sheet["D"+i] = column4
        i = int(i)
    wordbook.save(file)





def close_tab():
    var_v3.driver.implicitly_wait(5)
    # đóng tab không liên quan
    try:
        var_v3.driver.switch_to.window(var_v3.driver.window_handles[0])
        curr = var_v3.driver.current_window_handle
        for handle in var_v3.driver.window_handles:
            var_v3.driver.switch_to.window(handle)
            if handle != curr:
                var_v3.driver.close()
    except:
        pass




def getRowCount(file, sheetName):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)




def getColumnCount(file, sheetName):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)




def readData(file,sheetName,rownum,columnno):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum,column=columnno).value




def writeData(file,sheetName,caseid,columnno,data):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 0
    while (i < 5000):
        i += 1
        i = str(i)
        if sheet["A"+i].value == caseid:
            i = int(i)
            sheet.cell(row=i, column=columnno).value = data
            break
        i = int(i)
    wordbook.save(file)




@retry(tries=2, delay=2, backoff=1, jitter=5, )
def notification_telegram():
    from DrissionPage import ChromiumPage, ChromiumOptions
    do1 = ChromiumOptions().set_paths(local_port=9201, user_data_path=r""+var_v3.uploadpath+"Profile 30""")
    driver2 = ChromiumPage(addr_or_opts=do1)


    wordbook = openpyxl.load_workbook(var_v3.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    module_gpsv3.check_casenone()
    module_gpsv3.check_casefail()
    module_gpsv3.check_casepass()

    mucnghiemtrong = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 65, 2))
    tong_case_trong = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 66, 2))

    case_fail = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 77, 2))
    case_pass = str(var_v3.readData(var_v3.path_luutamthoi, 'Sheet1', 87, 2))




    time_string1 = time.strftime("%m/%d/%Y, ", time.localtime())
    time_string1 = str(time_string1)
    time_string2 = time.strftime("%H:%M", time.localtime())
    time_string2 = str(time_string2)
    thoigianbatdauchay = str(var_v3.readData(var_v3.path_luutamthoi , 'Sheet1', 47, 2))

    print("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2+
                                              "\n- LinkTest: " + var_v3.linktest+
                                              "\n- ModeTest: " + var_v3.modetest+
                                              "\n- Số case Pass: " + case_pass+
                                              "\n- Số case False: "+ case_fail+
                                              "\n- Số case trống: "+ tong_case_trong+
                                              "\n- Số case False nghiêm trọng: "+ mucnghiemtrong)



    thoigianbatdauchay = str(var_v3.readData(var_v3.path_luutamthoi , 'Sheet1', 47, 2))


    if int(case_fail) >= 1 or int(tong_case_trong) >= 1:
        print("đã vào tele")
        driver2.get("https://web.telegram.org/a/")
        time.sleep(2)
        case_pass = str(case_pass)
        case_fail = str(case_fail)
        try:
            driver2.ele("//*[text()='OK']").click()
        except:
            pass

        if var_v3.linktest[0:25] == "http://192.168.45.48:8000" or var_v3.linktest[0:25] == "http://192.168.45.82:8000":
            driver2.ele(var_v3.hopthoai1).click()
        else:
            driver2.ele(var_v3.hopthoai2).click()
        time.sleep(1)
        time_string1 = time.strftime("%m/%d/%Y, ", time.localtime())
        time_string1 = str(time_string1)
        time_string2 = time.strftime("%H:%M", time.localtime())
        time_string2 = str(time_string2)
        driver2.ele(var_v3.hopthoai_input).input("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2+
                                                  "\n- LinkTest: " + var_v3.linktest+
                                                  "\n- ModeTest: " + var_v3.modetest+
                                                  "\n- Số case Pass: " + case_pass+
                                                  "\n- Số case False: "+ case_fail+
                                                  "\n- Số case trống: "+ tong_case_trong+
                                                  "\n- Số case False nghiêm trọng: "+ mucnghiemtrong)
        driver2.ele(var_v3.hopthoai_input).input(Keys.ENTER)
        time.sleep(1)
        driver2.ele(var_v3.hopthoai_iconlink).click()
        time.sleep(1)
        driver2.ele(var_v3.hopthoai_iconlink_file).click()
        time.sleep(1)
        subprocess.Popen(var_v3.uploadpath+"checklist_bagps.exe")
        time.sleep(1)
        driver2.ele(var_v3.hopthoai_send).click()
        time.sleep(2)
        driver2.ele(var_v3.hopthoai_iconlink).click()
        time.sleep(1)
        driver2.ele(var_v3.hopthoai_iconlink_file).click()
        time.sleep(1)
        subprocess.Popen(var_v3.uploadpath+"ba_log.exe")
        time.sleep(1)
        driver2.ele(var_v3.hopthoai_send).click()
        time.sleep(1)

        time.sleep(30)
        driver2.close()



def delete_image():
    path = os.path.join(var_v3.imagepath)
    l = os.listdir(path)
    print(l)
    for i in l:
        print(os.path.join(path, i))
        os.remove(os.path.join(path, i))




def delete_excel():
    path = os.path.join(var_v3.excelpath)
    l = os.listdir(path)
    print(l)
    for i in l:
        print(os.path.join(path, i))
        os.remove(os.path.join(path, i))




def get_datachecklist(ma):
    wordbook = openpyxl.load_workbook(var_v3.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    while (rownum < 3000):
        rownum += 1
        rownum = str(rownum)
        if sheet["A" + rownum].value == ma:
            tensukien = sheet["B" + rownum].value
            ketqua = sheet["E" + rownum].value
            print(ma)
            print(tensukien)
            print(ketqua)
            logging.info("đang chạy case: " + ma)
        rownum = int(rownum)




@retry(tries=3, delay=2, backoff=1, jitter=5, )
def swich_tab_0():
    var_v3.driver.implicitly_wait(15)

    try:
        var_v3.driver.implicitly_wait(1)
        var_v3.driver.switch_to.alert.accept()
        time.sleep(1)
    except:
        pass


    try:
        var_v3.driver.implicitly_wait(1)
        subprocess.Popen(var_v3.uploadpath+"cancel.exe")
    except:
        pass


    time.sleep(1)
    try:
        # Chuyển sang tab hiện tại
        var_v3.driver.switch_to.window(var_v3.driver.window_handles[0])
        time.sleep(0.5)
        # Mở tab mới và chuyển sang tab mới trước khi truy cập URL
        var_v3.driver.execute_script("window.open('');")
        time.sleep(2)
        var_v3.driver.switch_to.window(var_v3.driver.window_handles[1])
        var_v3.driver.get("https://gps3.binhanh.vn/auth/login")
        time.sleep(3)
    except:
        # Trong trường hợp có lỗi, thử lại quy trình
        var_v3.driver.execute_script("window.open('');")
        time.sleep(2)
        var_v3.driver.switch_to.window(var_v3.driver.window_handles[1])
        var_v3.driver.get("https://gps3.binhanh.vn/auth/login")
        time.sleep(5)
        # Chuyển lại về tab đầu tiên
        var_v3.driver.switch_to.window(var_v3.driver.window_handles[0])


    time.sleep(1)
    try:
        # Chuyển sang tab hiện tại
        var_v3.driver.switch_to.window(var_v3.driver.window_handles[0])
        time.sleep(0.5)
        # Mở tab mới và chuyển sang tab mới trước khi truy cập URL
        var_v3.driver.execute_script("window.open('');")
        time.sleep(2)
        var_v3.driver.switch_to.window(var_v3.driver.window_handles[1])
        var_v3.driver.get("https://gps3.binhanh.vn/auth/login")
        time.sleep(3)
    except:
        # Trong trường hợp có lỗi, thử lại quy trình
        var_v3.driver.execute_script("window.open('');")
        time.sleep(2)
        var_v3.driver.switch_to.window(var_v3.driver.window_handles[1])
        var_v3.driver.get("https://gps3.binhanh.vn/auth/login")
        time.sleep(5)
        # Chuyển lại về tab đầu tiên
        var_v3.driver.switch_to.window(var_v3.driver.window_handles[0])



    # try:
    #     var_v3.driver.switch_to.window(var_v3.driver.window_handles[2])
    #     curr = var_v3.driver.current_window_handle
    #     for handle in var_v3.driver.window_handles:
    #         var_v3.driver.switch_to.window(handle)
    #         if handle != curr:
    #             var_v3.driver.close()
    #         time.sleep(1)
    #         var_v3.driver.switch_to.window(var_v3.driver.window_handles[0])
    #         time.sleep(0.5)
    # except:
    #     var_v3.driver.execute_script("window.open('');")
    #     time.sleep(2)
    #     var_v3.driver.switch_to.window(var_v3.driver.window_handles[0])
    #     var_v3.driver.get("https://gps3.binhanh.vn/auth/login")
    #     time.sleep(5)

    try:
        var_v3.driver.switch_to.window(var_v3.driver.window_handles[2])
        curr = var_v3.driver.current_window_handle
        for handle in var_v3.driver.window_handles:
            if handle != curr:
                var_v3.driver.switch_to.window(handle)
                var_v3.driver.close()
                time.sleep(1)
        var_v3.driver.switch_to.window(curr)
        time.sleep(0.5)

    except:
        var_v3.driver.execute_script("window.open('');")
        var_v3.driver.switch_to.window(var_v3.driver.window_handles[-1])  # Chuyển đến tab mới nhất
        var_v3.driver.get("https://gps3.binhanh.vn/auth/login")
        time.sleep(5)
        var_v3.driver.switch_to.window(var_v3.driver.window_handles[0])









def write_caseif():
    n = 0
    while (n < 26):
        var_v3.driver.implicitly_wait(1)
        n += 1
        n = str(n)
        print("try:\n   if case == 'Ai"+n+"':\n       caseid_v3.caseid_ai"+n+"(self)\nexcept:\n    module_other_v3.swich_tab_0()")
        n = int(n)



def write_caseif1():
    n = 53
    while (n < 70):
        var_v3.driver.implicitly_wait(1)
        n += 1
        n = str(n)
        print("try:\n   caseid.caseid_report"+n+"(self)\nexcept:\n     pass")
        n = int(n)



def write_caseif2():
    n = 53
    while (n < 70):
        var_v3.driver.implicitly_wait(1)
        n += 1
        n = str(n)
        print("caseid.caseid_report"+n+"(self)")
        n = int(n)



def write_result_text_try_if_color(code, eventname, result, path_module, path_color, check_result, name_image):
    var_v3.driver.implicitly_wait(1)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        element1 = var_v3.driver.find_element(By.XPATH, path_color)
        check_color = element1.value_of_css_property("color")
        logging.info(check_color)
        logging.info(check_result)
        writeData(var_v3.checklistpath, "Checklist", code, 6, check_color)

        if check_color == check_result:
            logging.info("True")
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_v3.driver.save_screenshot(var_v3.imagepath + code + name_image)
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_v3.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var_v3.driver.save_screenshot(var_v3.imagepath + code + name_image)
        writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_v3.checklistpath, "Checklist", code, 13, code + name_image)



def write_result_text_try_if_boolean(code, eventname, result, path_module, path_selector, check_result, name_image):
    var_v3.driver.implicitly_wait(1)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_selector = var_v3.driver.find_element(By.XPATH, path_selector).is_selected()
        logging.info(check_selector)
        logging.info(check_result)
        writeData(var_v3.checklistpath, "Checklist", code, 6, check_selector)

        if check_selector == check_result:
            logging.info("True")
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_v3.driver.save_screenshot(var_v3.imagepath + code + name_image)
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_v3.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var_v3.driver.save_screenshot(var_v3.imagepath + code + name_image)
        writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_v3.checklistpath, "Checklist", code, 13, code + name_image)




def write_result_dowload_file(code, eventname, result, path_module, file, name_image):
    var_v3.driver.implicitly_wait(1)
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        filename = max([var_v3.excelpath + "\\" + f for f in os.listdir(var_v3.excelpath)], key=os.path.getctime)
        shutil.move(filename, os.path.join(var_v3.excelpath, r""+file+""))
        logging.info("True")
        writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")
    except:
        logging.info("False")
        var_v3.driver.save_screenshot(var_v3.imagepath + code + name_image)
        writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_v3.checklistpath, "Checklist", code, 13, code + name_image)




def write_result_text_try_if_status_vehicle(code, eventname, result, path_module, all_on, status, path_count_vehicle, check_result, name_image):
    var_v3.driver.implicitly_wait(1)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        count_vehicle = var_v3.driver.find_element(By.XPATH, path_count_vehicle).text
        logging.info(count_vehicle)
        logging.info(check_result)
        writeData(var_v3.checklistpath, "Checklist", code, 6, "{}: {}\nTất cả trên: {}".format(status, count_vehicle, all_on))

        if count_vehicle == check_result:
            logging.info("True")
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_v3.driver.save_screenshot(var_v3.imagepath + code + name_image)
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_v3.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var_v3.driver.save_screenshot(var_v3.imagepath + code + name_image)
        writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_v3.checklistpath, "Checklist", code, 13, code + name_image)
    # module_other_v3.write_result_text_try_if(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_v3.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")





def write_result_text_try_if(code, eventname, result, path_module, path_text, check_result, name_image):
    var_v3.driver.implicitly_wait(1)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_text = var_v3.driver.find_element(By.XPATH, path_text).text
        logging.info(check_text)
        logging.info(check_result)
        writeData(var_v3.checklistpath, "Checklist", code, 6, check_text)

        if check_text == check_result:
            logging.info("True")
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_v3.driver.save_screenshot(var_v3.imagepath + code + name_image)
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_v3.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var_v3.driver.save_screenshot(var_v3.imagepath + code + name_image)
        writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_v3.checklistpath, "Checklist", code, 13, code + name_image)
    # module_other_v3.write_result_text_try_if(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_v3.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")




def write_result_text_try_if_src(code, eventname, result, path_module, path_text, check_result, name_image, src_to):
    var_v3.driver.implicitly_wait(1)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_text = var_v3.driver.find_element(By.XPATH, path_text).get_attribute("src")
        logging.info(check_text)
        logging.info(check_text[src_to::])
        logging.info(check_result)
        writeData(var_v3.checklistpath, "Checklist", code, 6, check_text)

        if check_text[src_to::] == check_result:
            logging.info("True")
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_v3.driver.save_screenshot(var_v3.imagepath + code + name_image)
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_v3.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var_v3.driver.save_screenshot(var_v3.imagepath + code + name_image)
        writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_v3.checklistpath, "Checklist", code, 13, code + name_image)
    # module_other_v3.write_result_text_try_if(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_v3.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")




def write_result_text_try_if_title(code, eventname, result, path_module, check_result, name_image):
    var_v3.driver.implicitly_wait(1)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        title = str(var_v3.driver.title)
        logging.info(title)
        logging.info(check_result)
        writeData(var_v3.checklistpath, "Checklist", code, 6, title)
        if title == check_result:
            logging.info("True")
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_v3.driver.save_screenshot(var_v3.imagepath + code + name_image)
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_v3.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var_v3.driver.save_screenshot(var_v3.imagepath + code + name_image)
        writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_v3.checklistpath, "Checklist", code, 13, code + name_image)
    # module_other_v3.write_result_text_try_if_title(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")





def write_result_text_try_if_other(code, eventname, result, path_module, path_text, check_result, name_image):
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_text = var_v3.driver.find_element(By.XPATH, path_text).text
        logging.info(check_text)
        writeData(var_v3.checklistpath, "Checklist", code, 6, check_text)
        if check_text != check_result:
            logging.info("True")
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            var_v3.driver.save_screenshot(var_v3.imagepath + code + name_image)
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_v3.checklistpath, "Checklist", code, 13, code + name_image)
    except:
        logging.info("False")
        var_v3.driver.save_screenshot(var_v3.imagepath + code + name_image)
        writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_v3.checklistpath, "Checklist", code, 13, code + name_image)
    # module_other_v3.write_result_text_try_if_other(code, eventname, result, "Quản trị - Danh sách xe",
    #                                       var_v3.check_open_car_quickly, "Mở xe thành công", "_QuanTri_DsXe_MoXeNhanh.png")




def write_result_displayed_try(code, eventname, result, path_module, path_text, name_image):
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_displayed = var_v3.driver.find_element(By.XPATH, path_text).is_displayed()
        logging.info("True")
        writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")
    except NoSuchElementException:
        logging.info("False")
        var_v3.driver.save_screenshot(var_v3.imagepath + code + name_image)
        writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_v3.checklistpath, "Checklist", code, 13, code + name_image)

    # logging.info("Tìm biển kiểm soát - " + data['quantri']['bienkiemsoat'])
    # module_other_v3.write_result_displayed_try(code, eventname, result, "Quản trị - Danh sách xe",
    #                                         var_v3.check_hide_car, "_QuanTri_DsXe_AnXe.png")




def write_result_not_displayed_try(code, eventname, result, path_module, path_text, name_image):
    var_v3.driver.implicitly_wait(2)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        check_not_displayed = var_v3.driver.find_element(By.XPATH, path_text).is_displayed()
        logging.info("False")
        var_v3.driver.save_screenshot(var_v3.imagepath + code + name_image)
        writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_v3.checklistpath, "Checklist", code, 13, code + name_image)
    except NoSuchElementException:
        logging.info("True")
        writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")
    # module_other_v3.write_result_displayed_try(code, eventname, result, "Quản trị - Danh sách xe",
    #                                         var_v3.check_hide_car, "_QuanTri_DsXe_AnXe.png")




def write_result_close_popup(code, eventname, result, path_module, path_text, name_image):
    var_v3.driver.implicitly_wait(2)
    logging.info("-------------------------")
    logging.info(path_module)
    logging.info("Mã - " + code)
    logging.info("Tên sự kiện - " + eventname)
    logging.info("Kết quả - " + result)
    try:
        var_v3.driver.find_element(By.XPATH, path_text).click()
        logging.info("False")
        var_v3.driver.save_screenshot(var_v3.imagepath + code + name_image)
        writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_v3.checklistpath, "Checklist", code, 13, code + name_image)
    except NoSuchElementException:
        logging.info("True")
        writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")


    # module_other_v3.write_result_close_popup(code, eventname, result, "Quản trị - Danh sách xe",
    #                                         var_v3.check_hide_car, "_QuanTri_DsXe_AnXe.png")




def write_result_excelreport(code, sheet, column, name_sheet, number_column, number_row, output):
    if str(sheet[column + number_column]) == "<Cell '"+name_sheet+"'." + number_row + ">":
        logging.info("Check vị trí: "+number_row+":  "+output+"")
        if str(sheet[column + number_column].value) == output:
            logging.info("True")
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_v3.checklistpath, "Checklist", code, 6, "File Báo cáo tổng hợp hoạt động sai ô " + number_row)
    # module_other_v3.write_result_excelreport(code, sheet, column, 'BC Tổng hợp', "5", "C5", "STT")




def write_result_excelreport1(code, sheet, column, name_sheet, number_column, number_row, output, number_row2, output2):
    data_excel = str(sheet[number_row2].value)
    output2 = str(output2)

    print("Dữ liệu web: " +output2)
    print("Dữ liệu excel: " +data_excel)
    if str(sheet[column + number_column]) == "<Cell '"+name_sheet+"'." + number_row + ">":
        logging.info("Check vị trí: "+number_row+":  "+output+"")
        logging.info("Dữ liệu excel: "+ data_excel)
        logging.info("Dữ liệu web: "+ output2)
        if str(sheet[column + number_column].value) == output and data_excel == output2:
            logging.info("True")
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_v3.checklistpath, "Checklist", code, 6, "File Báo cáo tổng hợp hoạt động sai ô " + number_row2)
    # if str(sheet[column + number_column]) == "<Cell '"+name_sheet+"'." + number_row + ">":
    # chucnangkhac.write_result_excelreport1(code, sheet, column, 'BC Tổng hợp', "5", "D5", "Ngày tháng", "D10", activity_synthesis_group_report_day_month)




def write_result_excelreport2(code, output_web, output_excel, name_output):
    logging.info(name_output + " web: " + output_web)
    logging.info(name_output + " excel: " + output_excel)
    if output_web == output_excel:
        logging.info("True")
        writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")
    else:
        logging.info("False")
        writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_v3.checklistpath, "Checklist", code, 6, "File Báo cáo tổng hợp hoạt động sai trường " + output_web)






def write_result_excel_checkweb(code, data_web, desire):
    logging.info("Dữ liệu web: " + data_web)
    logging.info("Dữ liệu mong muốn: " + desire)
    if data_web == desire:
        logging.info("True")
        writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")
    else:
        logging.info("False")
        writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
        writeData(var_v3.checklistpath, "Checklist", code, 6, "File Báo cáo tổng hợp hoạt động mất sai trường" + desire)








def write_result_excelreport_clear_data(code, sheet, column, name_sheet, number_column, number_row, output):
    if str(sheet[column + number_column]) == "<Cell '"+name_sheet+"'." + number_row + ">":
        logging.info("Check vị trí: "+number_row+": "+output+"")
        if str(sheet[column + number_column].value) == output:
            logging.info("True")
            writeData(var_v3.checklistpath, "Checklist", code, 6, " ")
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("False")
            writeData(var_v3.checklistpath, "Checklist", code, 7, "Fail")
            writeData(var_v3.checklistpath, "Checklist", code, 6, "File Báo cáo tổng hợp hoạt động sai ô " + number_row)
    # chucnangkhac.write_result_excelreport(code, sheet, column, 'BC Tổng hợp', "5", "C5", "STT")





